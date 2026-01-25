from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count, Avg, Max
from django.db import models
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.models import User
from django.utils import timezone

from .models import Property, PropertyType, Region, District, Amenity, PropertyImage, PropertyFavorite, Booking, Room, Customer, Payment
from documents.models import Booking as DocumentBooking
from .utils import get_management_context, get_property_filtered_queryset, clear_property_selection, set_property_selection
from .forms import (
    PropertyForm, PropertyImageFormSet, PropertySearchForm,
    RegionForm, DistrictForm, PropertyTypeForm, AmenityForm,
    HotelPropertyForm, HousePropertyForm, LodgePropertyForm, VenuePropertyForm
)

# Import rent reminder views
from .house_rent_reminder_views import (
    house_rent_reminders_dashboard,
    house_rent_reminders_list,
    house_rent_reminder_detail,
    house_rent_reminder_settings,
    house_rent_reminder_templates,
    house_rent_reminder_template_detail,
    house_rent_reminder_analytics,
    send_manual_reminder,
    cancel_reminder
)


# Helper function to wrap DocumentBooking for template compatibility
class DocumentBookingWrapper:
    """Wrapper to make DocumentBooking compatible with properties.Booking for template"""
    def __init__(self, doc_booking):
        self.doc_booking = doc_booking
        self.id = doc_booking.id
        self.booking_reference = f"MOB-{doc_booking.id:06d}"
        self.property_obj = doc_booking.property_ref
        self.check_in_date = doc_booking.check_in
        self.check_out_date = doc_booking.check_out
        self.total_amount = doc_booking.total_amount
        self.paid_amount = 0  # DocumentBooking doesn't track this
        self.deposit_amount = 0  # DocumentBooking doesn't track this
        self.booking_status = doc_booking.status
        self.payment_status = 'pending'  # Default
        self.number_of_guests = 1  # Default
        self.room_number = None  # DocumentBooking doesn't have room_number
        self.room_type = None  # DocumentBooking doesn't have room_type
        self.special_requests = None  # DocumentBooking doesn't have special_requests
        self.created_at = doc_booking.created_at
        self.updated_at = doc_booking.created_at  # Use created_at as fallback
        self.created_by = doc_booking.tenant
        self.assigned_staff = None  # DocumentBooking doesn't have assigned_staff
        self.confirmed_at = None
        self.checked_in_at = None
        self.checked_out_at = None
        self.is_deleted = False  # DocumentBooking doesn't support soft delete
        self.deleted_at = None
        self.deleted_by = None
        
        # Create a customer-like object from the tenant user
        # Get phone from profile if available
        phone = ''
        try:
            if hasattr(doc_booking.tenant, 'profile'):
                phone = doc_booking.tenant.profile.phone or ''
        except:
            pass
        
        # Get full name
        first_name = doc_booking.tenant.first_name or ''
        last_name = doc_booking.tenant.last_name or ''
        if not first_name and not last_name:
            # Fallback to username
            username = doc_booking.tenant.username
            if '@' in username:
                first_name = username.split('@')[0]
            else:
                first_name = username
        
        full_name = f"{first_name} {last_name}".strip() or doc_booking.tenant.username
        
        self.customer = type('Customer', (), {
            'id': doc_booking.tenant.id,
            'first_name': first_name,
            'last_name': last_name,
            'full_name': full_name,
            'email': doc_booking.tenant.email or '',
            'phone': phone,
        })()
        self.is_mobile_booking = True  # Flag to identify mobile bookings
    
    @property
    def duration_days(self):
        """Calculate duration in days"""
        return (self.check_out_date - self.check_in_date).days
    
    @property
    def remaining_amount(self):
        """Calculate remaining amount"""
        return self.total_amount - self.paid_amount
    
    def get_booking_status_display(self):
        """Return display name for booking status"""
        status_map = {
            'pending': 'Pending',
            'confirmed': 'Confirmed',
            'cancelled': 'Cancelled',
            'completed': 'Completed',
        }
        return status_map.get(self.booking_status, self.booking_status.title())
    
    def get_payment_status_display(self):
        """Return display name for payment status"""
        status_map = {
            'pending': 'Pending',
            'partial': 'Partial',
            'paid': 'Paid',
            'refunded': 'Refunded',
        }
        return status_map.get(self.payment_status, self.payment_status.title())


class LeaseWrapper:
    """Wrapper to make Lease compatible with properties.Booking for template"""
    def __init__(self, lease):
        self.lease = lease
        self.id = lease.id
        self.booking_reference = f"LEASE-{lease.id:06d}"
        self.property_obj = lease.property_ref
        self.check_in_date = lease.start_date
        self.check_out_date = lease.end_date
        self.total_amount = lease.rent_amount
        self.paid_amount = 0  # Lease doesn't track this separately
        self.deposit_amount = 0  # Lease doesn't track this
        self.booking_status = 'confirmed' if lease.status == 'active' else lease.status
        self.payment_status = 'pending'  # Default
        self.number_of_guests = 1  # Default
        self.room_number = None
        self.room_type = None
        self.special_requests = None
        self.created_at = lease.created_at
        self.updated_at = lease.created_at
        self.created_by = lease.tenant
        self.assigned_staff = None
        self.confirmed_at = None
        self.checked_in_at = None
        self.checked_out_at = None
        self.is_deleted = False  # Lease doesn't support soft delete
        self.deleted_at = None
        self.deleted_by = None
        
        # Create a customer-like object from the tenant user
        phone = ''
        try:
            if hasattr(lease.tenant, 'profile'):
                phone = lease.tenant.profile.phone or ''
        except:
            pass
        
        first_name = lease.tenant.first_name or ''
        last_name = lease.tenant.last_name or ''
        if not first_name and not last_name:
            username = lease.tenant.username
            if '@' in username:
                first_name = username.split('@')[0]
            else:
                first_name = username
        
        full_name = f"{first_name} {last_name}".strip() or lease.tenant.username
        
        self.customer = type('Customer', (), {
            'id': lease.tenant.id,
            'first_name': first_name,
            'last_name': last_name,
            'full_name': full_name,
            'email': lease.tenant.email or '',
            'phone': phone,
        })()
        self.is_mobile_lease = True  # Flag to identify mobile leases
    
    @property
    def duration_days(self):
        """Calculate duration in days"""
        return (self.check_out_date - self.check_in_date).days
    
    @property
    def remaining_amount(self):
        """Calculate remaining amount"""
        return self.total_amount - self.paid_amount
    
    def get_booking_status_display(self):
        """Return display name for booking status"""
        status_map = {
            'pending': 'Pending',
            'active': 'Active',
            'terminated': 'Terminated',
            'expired': 'Expired',
            'rejected': 'Rejected',
            'confirmed': 'Confirmed',
        }
        return status_map.get(self.booking_status, self.booking_status.title())
    
    def get_payment_status_display(self):
        """Return display name for payment status"""
        return 'Pending'  # Leases don't track payment status separately


def auto_complete_venue_bookings(user=None):
    """
    Automatically mark venue bookings as completed when their checkout date has passed.
    Returns the number of bookings updated.
    """
    today = timezone.now().date()
    filters = {
        'property_obj__property_type__name__iexact': 'venue',
        'check_out_date__lt': today,
        'booking_status__in': ['pending', 'confirmed', 'checked_in'],
    }

    if user and user.is_authenticated and not (user.is_staff or user.is_superuser):
        filters['property_obj__owner'] = user

    return Booking.objects.filter(**filters).update(booking_status='checked_out')


def combine_bookings(prop_bookings_query, doc_bookings_query, property_type_name, selected_property=None, search_query='', status_filter='', payment_filter='', user=None):
    """
    Helper function to combine bookings from properties.Booking and documents.Booking models.
    
    Args:
        prop_bookings_query: QuerySet of properties.Booking
        doc_bookings_query: QuerySet of documents.Booking
        property_type_name: Property type name ('hotel', 'lodge', 'venue', 'house')
        selected_property: Optional Property instance to filter by
        search_query: Optional search string
        status_filter: Optional status filter
        payment_filter: Optional payment status filter
        user: Optional user for multi-tenancy filtering
    
    Returns:
        Combined list of bookings sorted by created_at
    """
    # Apply search filter if provided
    if search_query:
        prop_bookings_query = prop_bookings_query.filter(
            Q(customer__first_name__icontains=search_query) |
            Q(customer__last_name__icontains=search_query) |
            Q(customer__phone__icontains=search_query) |
            Q(property_obj__title__icontains=search_query) |
            Q(property_obj__address__icontains=search_query)
        )
        # Also filter document bookings
        doc_bookings_query = doc_bookings_query.filter(
            Q(tenant__first_name__icontains=search_query) |
            Q(tenant__last_name__icontains=search_query) |
            Q(tenant__email__icontains=search_query) |
            Q(tenant__username__icontains=search_query) |
            Q(property_ref__title__icontains=search_query) |
            Q(property_ref__address__icontains=search_query)
        )
    
    # Apply status filter if provided
    if status_filter:
        # Map status values between the two models
        status_mapping = {
            'pending': 'pending',
            'confirmed': 'confirmed',
            'cancelled': 'cancelled',
            'checked_in': 'confirmed',  # DocumentBooking doesn't have checked_in, use confirmed
            'checked_out': 'completed',  # Map to completed
        }
        doc_status = status_mapping.get(status_filter, status_filter)
        prop_bookings_query = prop_bookings_query.filter(booking_status=status_filter)
        doc_bookings_query = doc_bookings_query.filter(status=doc_status)
    
    # Apply payment status filter if provided (only for properties.Booking)
    if payment_filter:
        prop_bookings_query = prop_bookings_query.filter(payment_status=payment_filter)
        # DocumentBooking doesn't have payment_status, so we skip it
    
    # Filter by selected property if specified
    if selected_property:
        prop_bookings_query = prop_bookings_query.filter(property_obj=selected_property)
        doc_bookings_query = doc_bookings_query.filter(property_ref=selected_property)
    
    # Get all bookings as lists
    prop_bookings_list = list(prop_bookings_query)
    doc_bookings_list = list(doc_bookings_query)
    
    # Wrap document bookings
    wrapped_doc_bookings = [DocumentBookingWrapper(db) for db in doc_bookings_list]
    
    # Combine both lists and sort by created_at
    all_bookings = sorted(
        list(prop_bookings_list) + wrapped_doc_bookings,
        key=lambda x: x.created_at,
        reverse=True
    )
    
    return all_bookings


def property_list(request):
    """
    View for listing all properties with search functionality and pagination
    MULTI-TENANCY: Property owners only see their own properties, admins see all
    """
    # MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        # Property owner: only see own properties (including unapproved ones)
        properties = Property.objects.filter(owner=request.user).select_related(
            'property_type', 'region', 'owner'
        ).prefetch_related('images', 'amenities').order_by('-created_at')
    elif request.user.is_staff or request.user.is_superuser:
        # Admin/staff: see all properties (including unapproved ones)
        properties = Property.objects.select_related(
            'property_type', 'region', 'owner'
        ).prefetch_related('images', 'amenities').order_by('-created_at')
    else:
        # Unauthenticated/public: only see approved and active properties
        properties = Property.objects.filter(
            is_active=True,
            is_approved=True
        ).select_related(
            'property_type', 'region', 'owner'
        ).prefetch_related('images', 'amenities').order_by('-created_at')
    
    # Get all property types and regions for filter dropdowns
    property_types = PropertyType.objects.all()
    regions = Region.objects.all()
    
    # Apply filters based on GET parameters
    search_query = request.GET.get('search', '').strip()
    if search_query:
        properties = properties.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(address__icontains=search_query)
        )
    
    property_type_id = request.GET.get('property_type')
    if property_type_id:
        properties = properties.filter(property_type_id=property_type_id)
    
    region_id = request.GET.get('region')
    if region_id:
        properties = properties.filter(region_id=region_id)
    
    # District filter
    district_id = request.GET.get('district')
    if district_id:
        try:
            properties = properties.filter(district_id=int(district_id))
        except (ValueError, TypeError):
            pass  # Invalid district_id, ignore
    
    bedrooms = request.GET.get('bedrooms')
    if bedrooms:
        if bedrooms == '4':  # 4+ bedrooms
            properties = properties.filter(bedrooms__gte=4)
        else:
            properties = properties.filter(bedrooms=bedrooms)
    
    status = request.GET.get('status')
    if status:
        properties = properties.filter(status=status)
    
    # Pagination
    page_size = request.GET.get('page_size', '5')  # Default to 5 items per page
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50, 100]:
            page_size = 5
    except (ValueError, TypeError):
        page_size = 5
    
    paginator = Paginator(properties, page_size)
    page_number = request.GET.get('page')
    
    try:
        page_obj = paginator.get_page(page_number)
    except:
        page_obj = paginator.get_page(1)
    
    # Calculate statistics
    total_properties = Property.objects.count()
    available_count = Property.objects.filter(status='available').count()
    rented_count = Property.objects.filter(status='rented').count()
    maintenance_count = Property.objects.filter(status='under_maintenance').count()
    
    context = {
        'properties': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'property_types': property_types,
        'regions': regions,
        'total_properties': total_properties,
        'available_count': available_count,
        'rented_count': rented_count,
        'maintenance_count': maintenance_count,
        'current_page_size': page_size,
        'search_query': search_query,
        'current_property_type': property_type_id or '',
        'current_region': region_id or '',
        'current_status': status or '',
    }
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'properties/property_list_table.html', context)
    
    return render(request, 'properties/property_list.html', context)


def property_detail(request, pk):
    """View for displaying property details"""
    property_obj = get_object_or_404(
        Property.objects.select_related('property_type', 'region', 'owner')
        .prefetch_related(
            models.Prefetch('images', queryset=PropertyImage.objects.order_by('-is_primary', 'order')),
            'amenities'
        ),
        pk=pk
    )
    
    # Track property view (if user is authenticated)
    if request.user.is_authenticated:
        from .models import PropertyView
        PropertyView.objects.get_or_create(
            property=property_obj,
            user=request.user,
            defaults={
                'ip_address': request.META.get('REMOTE_ADDR', ''),
                'user_agent': request.META.get('HTTP_USER_AGENT', '')
            }
        )
    
    # Check if property is favorited by current user
    is_favorited = False
    if request.user.is_authenticated:
        is_favorited = PropertyFavorite.objects.filter(
            user=request.user,
            property=property_obj
        ).exists()
    
    # Get related properties (same region or property type)
    related_properties = Property.objects.filter(
        Q(region=property_obj.region) | Q(property_type=property_obj.property_type)
    ).exclude(pk=property_obj.pk).select_related(
        'property_type', 'region'
    ).prefetch_related('images')[:6]
    
    # Get active lease/occupant if property is rented or occupied
    active_lease = None
    current_occupant = None
    active_booking = None
    
    # Check for active lease (for long-term rentals)
    if property_obj.status in ['rented', 'occupied']:
        try:
            from documents.models import Lease
            from django.utils import timezone
            today = timezone.now().date()
            # Check for active lease - also check leases that might have expired but status is still active
            active_lease = Lease.objects.filter(
                property_ref=property_obj,
                status='active'
            ).select_related('tenant').order_by('-start_date').first()
            
            # If lease exists, check if it's currently active (within date range)
            if active_lease:
                if active_lease.start_date <= today <= active_lease.end_date:
                    current_occupant = active_lease.tenant
                # If property is rented but lease dates don't match, still show the tenant
                # (in case dates weren't updated properly)
                elif property_obj.status == 'rented':
                    current_occupant = active_lease.tenant
        except ImportError as e:
            import sys
            print(f"Error importing Lease: {e}", file=sys.stderr)
        except Exception as e:
            import sys
            print(f"Error checking lease: {e}", file=sys.stderr)
    
    # Also check for active bookings (for hotels/lodges/venues)
    if not current_occupant and property_obj.status in ['rented', 'occupied', 'unavailable']:
        try:
            from django.utils import timezone
            from .models import Booking
            today = timezone.now().date()
            active_booking = Booking.objects.filter(
                property_obj=property_obj,
                booking_status__in=['confirmed', 'checked_in'],
                check_in_date__lte=today,
                check_out_date__gte=today
            ).select_related('customer').first()
            
            if active_booking:
                current_occupant = active_booking.customer
        except Exception as e:
            import sys
            print(f"Error checking booking: {e}", file=sys.stderr)
    
    # Check if user has paid for visit access (for house properties only)
    has_visit_access = False
    if request.user.is_authenticated and property_obj.property_type.name.lower() == 'house':
        from .models import PropertyVisitPayment
        visit_payment = PropertyVisitPayment.objects.filter(
            property=property_obj,
            user=request.user,
            status='completed'
        ).first()
        # Check if payment exists and is still active (not expired)
        has_visit_access = visit_payment is not None and visit_payment.is_active()
    
    # Check if user can see owner contact (admin, owner, or has paid for visit)
    can_see_owner_contact = False
    owner_phone = None
    if request.user.is_authenticated:
        if (request.user == property_obj.owner or 
            request.user.is_staff or 
            request.user.is_superuser or 
            has_visit_access):
            can_see_owner_contact = True
            # Get owner phone from profile
            if hasattr(property_obj.owner, 'profile') and property_obj.owner.profile.phone:
                owner_phone = property_obj.owner.profile.phone
    
    context = {
        'property': property_obj,
        'is_favorited': is_favorited,
        'related_properties': related_properties,
        'active_lease': active_lease,
        'active_booking': active_booking,
        'current_occupant': current_occupant,
        'has_visit_access': has_visit_access,
        'can_see_owner_contact': can_see_owner_contact,
        'owner_phone': owner_phone,
    }
    
    return render(request, 'properties/property_detail.html', context)


@login_required
def property_create(request):
    """View for creating a new property"""
    property_type = request.GET.get('type', '').lower()
    
    # Get PropertyType object if type is specified
    property_type_obj = None
    if property_type:
        # Use first() to handle potential duplicates gracefully (returns None if not found)
        property_type_obj = PropertyType.objects.filter(name__iexact=property_type).first()
    
    # Select appropriate form based on property type
    if property_type == 'hotel':
        form_class = HotelPropertyForm
        title = 'Create Hotel'
    elif property_type == 'house':
        form_class = HousePropertyForm
        title = 'Create House'
    elif property_type == 'lodge':
        form_class = LodgePropertyForm
        title = 'Create Lodge'
    elif property_type == 'venue':
        form_class = VenuePropertyForm
        title = 'Create Venue'
    else:
        form_class = PropertyForm
        title = 'Create Property'
    
    if request.method == 'POST':
        # Get property_type from POST data or URL parameter (for form re-rendering after errors)
        post_property_type = request.POST.get('type') or request.GET.get('type', '').lower()
        if post_property_type and not property_type:
            property_type = post_property_type
            # Use first() to handle potential duplicates gracefully
            property_type_obj = PropertyType.objects.filter(name__iexact=property_type).first()
        
        # Create a mutable copy of POST data
        post_data = request.POST.copy()
        
        # Ensure property_type is set if it was specified in URL
        if property_type_obj and ('property_type' not in post_data or not post_data.get('property_type')):
            post_data['property_type'] = str(property_type_obj.pk)
        
        form = form_class(post_data, owner=request.user)
        image_formset = PropertyImageFormSet(request.POST, request.FILES)
        
        # Debug: Log form and formset validation
        if not form.is_valid():
            for field, errors in form.errors.items():
                messages.error(request, f"Form error - {field}: {', '.join(errors)}")
        
        if not image_formset.is_valid():
            for error in image_formset.non_form_errors():
                messages.error(request, f"Image formset error: {error}")
            for form in image_formset:
                if form.errors:
                    for field, errors in form.errors.items():
                        messages.error(request, f"Image form error - {field}: {', '.join(errors)}")
        
        if form.is_valid() and image_formset.is_valid():
            property_obj = form.save(commit=False)
            property_obj.owner = request.user
            # Ensure property_type is set
            if property_type_obj and not property_obj.property_type:
                property_obj.property_type = property_type_obj
            # Admin/staff created properties are auto-approved, others require approval
            if request.user.is_staff or request.user.is_superuser:
                property_obj.is_active = True
                property_obj.is_approved = True
                property_obj.approved_by = request.user
                property_obj.approved_at = timezone.now()
                messages.success(request, f'{property_obj.property_type.name} created successfully!')
            else:
                # Non-admin owners require approval
                property_obj.is_active = False
                property_obj.is_approved = False
                messages.success(request, f'{property_obj.property_type.name} created successfully! It is pending admin approval before being visible to the public.')
            property_obj.save()
            form.save_m2m()  # Save many-to-many relationships (amenities)
            
            # Save images
            image_formset.instance = property_obj
            image_formset.save()
            
            # Redirect based on property type and user role
            property_type_name = property_obj.property_type.name.lower() if property_obj.property_type else ''
            
            # For property owners, redirect to property selection page for that type
            if not (request.user.is_staff or request.user.is_superuser):
                if property_type_name == 'hotel':
                    return redirect('properties:hotel_select_property')
                elif property_type_name == 'lodge':
                    return redirect('properties:lodge_select_property')
                elif property_type_name == 'venue':
                    return redirect('properties:venue_select_property')
                elif property_type_name == 'house':
                    return redirect('properties:house_select_property')
            
            # For admins, redirect to property detail
            return redirect('properties:property_detail', pk=property_obj.pk)
    else:
        # Set initial property_type if specified in URL
        initial_data = {}
        if property_type_obj:
            initial_data['property_type'] = property_type_obj.pk  # Pass the ID, not the object
        form = form_class(owner=request.user, initial=initial_data)
        image_formset = PropertyImageFormSet()
    
    context = {
        'form': form,
        'image_formset': image_formset,
        'title': title,
        'property_type': property_type,
        'property': None,  # Explicitly set to None when creating to avoid template errors
    }
    
    return render(request, 'properties/property_form.html', context)


@login_required
def property_edit(request, pk):
    """View for editing an existing property"""
    # Allow admins and staff to edit any property, otherwise only owner can edit
    if request.user.is_staff or request.user.is_superuser:
        property_obj = get_object_or_404(Property, pk=pk)
    else:
        property_obj = get_object_or_404(Property, pk=pk, owner=request.user)
    
    # Select appropriate form based on property type
    property_type_name = property_obj.property_type.name.lower() if property_obj.property_type else ''
    if property_type_name == 'hotel':
        form_class = HotelPropertyForm
    elif property_type_name == 'house':
        form_class = HousePropertyForm
    elif property_type_name == 'lodge':
        form_class = LodgePropertyForm
    elif property_type_name == 'venue':
        form_class = VenuePropertyForm
    else:
        form_class = PropertyForm
    
    if request.method == 'POST':
        import sys
        print("\n" + "=" * 70, file=sys.stderr)
        print("🔥 POST REQUEST RECEIVED FOR PROPERTY EDIT 🔥", file=sys.stderr)
        print("=" * 70, file=sys.stderr)
        print(f"Property ID: {pk}", file=sys.stderr)
        print(f"Title from POST: '{request.POST.get('title', 'NOT FOUND')}'", file=sys.stderr)
        print("=" * 70 + "\n", file=sys.stderr)
        
        # Ensure required fields are included in POST data if missing
        post_data = request.POST.copy()
        
        # size_sqft is required
        if 'size_sqft' not in post_data or not post_data.get('size_sqft'):
            if property_obj.size_sqft:
                post_data['size_sqft'] = str(property_obj.size_sqft)
        
        # booking_expiration_hours is required
        if 'booking_expiration_hours' not in post_data or not post_data.get('booking_expiration_hours'):
            if property_obj.booking_expiration_hours:
                post_data['booking_expiration_hours'] = str(property_obj.booking_expiration_hours)
            else:
                # Set default value if property doesn't have it
                post_data['booking_expiration_hours'] = '0'  # 0 means no auto-expiration
        
        form = form_class(post_data, request.FILES, instance=property_obj, owner=request.user)
        
        # Create formset - always include POST and FILES data
        image_formset = PropertyImageFormSet(
            request.POST, 
            request.FILES, 
            instance=property_obj
        )
        
        # Validate form first - this is the critical part
        import sys
        is_valid = form.is_valid()
        print(f"Form is_valid(): {is_valid}", file=sys.stderr)
        
        if not is_valid:
            print("=" * 70, file=sys.stderr)
            print("FORM VALIDATION ERRORS:", file=sys.stderr)
            print("=" * 70, file=sys.stderr)
            for field, errors in form.errors.items():
                print(f"  {field}: {errors}", file=sys.stderr)
            print("=" * 70, file=sys.stderr)
        
        # Save form if valid, don't let formset block it
        if is_valid:
            # Save the main form FIRST - don't wait for formset validation
            property_obj = form.save()
            print(f"✓ Property saved! New title: '{property_obj.title}'", file=sys.stderr)
            # Save many-to-many relationships (amenities) if method exists
            if hasattr(form, 'save_m2m'):
                form.save_m2m()
            
            # Handle formset separately - don't let it block the main save
            try:
                if image_formset.is_valid():
                    image_formset.save()
            except Exception as e:
                # Continue anyway - main form is saved
                pass
            
            messages.success(request, 'Property updated successfully!')
            print("→ Redirecting to property list...", file=sys.stderr)
            return redirect('properties:property_list')
        else:
            # Form validation failed - show errors and stay on edit page
            print("Form validation failed - staying on edit page", file=sys.stderr)
            if not is_valid:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"{field}: {error}")
            if not image_formset.is_valid():
                for error in image_formset.non_form_errors():
                    messages.error(request, f"Image error: {error}")
            # Don't redirect - stay on edit page to show errors
    else:
        form = form_class(instance=property_obj, owner=request.user)
        image_formset = PropertyImageFormSet(instance=property_obj)
    
    # Determine property type for template
    property_type = property_obj.property_type.name.lower() if property_obj.property_type else ''
    
    context = {
        'form': form,
        'image_formset': image_formset,
        'property': property_obj,
        'property_type': property_type,
        'title': 'Edit Property',
    }
    
    return render(request, 'properties/property_form.html', context)


@login_required
def property_delete(request, pk):
    """View for deleting a property"""
    property_obj = get_object_or_404(Property, pk=pk, owner=request.user)
    
    if request.method == 'POST':
        property_obj.delete()
        messages.success(request, 'Property deleted successfully!')
        return redirect('properties:my_properties')
    
    context = {
        'property': property_obj,
    }
    
    return render(request, 'properties/property_confirm_delete.html', context)


@login_required
def my_properties(request):
    """View for displaying user's properties"""
    properties = Property.objects.filter(owner=request.user).select_related(
        'property_type', 'region'
    ).prefetch_related('images').annotate(
        views_count=Count('views'),
        favorites_count=Count('favorited_by')
    )
    
    # Statistics
    stats = {
        'total': properties.count(),
        'available': properties.filter(status='available').count(),
        'rented': properties.filter(status='rented').count(),
        'under_maintenance': properties.filter(status='under_maintenance').count(),
        'total_views': sum(p.views_count for p in properties),
        'total_favorites': sum(p.favorites_count for p in properties),
    }
    
    context = {
        'properties': properties,
        'stats': stats,
    }
    
    return render(request, 'properties/my_properties.html', context)


@login_required
@require_http_methods(["POST"])
def toggle_favorite(request, pk):
    """AJAX view for toggling property favorite status"""
    property_obj = get_object_or_404(Property, pk=pk)
    
    favorite, created = PropertyFavorite.objects.get_or_create(
        user=request.user,
        property=property_obj
    )
    
    if not created:
        favorite.delete()
        is_favorited = False
    else:
        is_favorited = True
    
    return JsonResponse({
        'is_favorited': is_favorited,
        'favorites_count': property_obj.favorited_by.count()
    })


@login_required
def favorites(request):
    """View for displaying user's favorite properties"""
    favorites = PropertyFavorite.objects.filter(user=request.user).select_related(
        'property__property_type', 'property__region', 'property__owner'
    ).prefetch_related('property__images')
    
    context = {
        'favorites': favorites,
    }
    
    return render(request, 'properties/favorites.html', context)


# Management views for property types, regions, and amenities
@login_required
def manage_regions(request):
    """View for managing regions with pagination and search"""
    regions = Region.objects.all().order_by('name')
    
    # Server-side search
    search_query = request.GET.get('search', '').strip()
    if search_query:
        regions = regions.filter(
            Q(name__icontains=search_query) | Q(description__icontains=search_query)
        )
    
    # Pagination
    page_size = request.GET.get('page_size', '5')  # Default to 5 items per page
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50, 100]:
            page_size = 5
    except (ValueError, TypeError):
        page_size = 5
    
    paginator = Paginator(regions, page_size)
    page_number = request.GET.get('page')
    
    try:
        page_obj = paginator.get_page(page_number)
    except:
        page_obj = paginator.get_page(1)
    
    if request.method == 'POST':
        form = RegionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Region created successfully!')
            return redirect('properties:manage_regions')
    else:
        form = RegionForm()
    
    context = {
        'regions': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'form': form,
        'current_page_size': page_size,
        'search_query': search_query,
    }
    
    return render(request, 'properties/manage_regions.html', context)


@login_required
def manage_districts(request):
    """View for managing districts with pagination and search"""
    districts = District.objects.select_related('region').all().order_by('region__name', 'name')
    
    # Server-side search
    search_query = request.GET.get('search', '').strip()
    if search_query:
        districts = districts.filter(
            Q(name__icontains=search_query) | 
            Q(description__icontains=search_query) |
            Q(region__name__icontains=search_query)
        )
    
    # Filter by region if provided
    region_id = request.GET.get('region')
    if region_id:
        districts = districts.filter(region_id=region_id)
    
    # Pagination
    page_size = request.GET.get('page_size', '5')  # Default to 5 items per page
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50, 100]:
            page_size = 5
    except (ValueError, TypeError):
        page_size = 5
    
    paginator = Paginator(districts, page_size)
    page_number = request.GET.get('page')
    
    try:
        page_obj = paginator.get_page(page_number)
    except:
        page_obj = paginator.get_page(1)
    
    if request.method == 'POST':
        form = DistrictForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'District created successfully!')
            return redirect('properties:manage_districts')
    else:
        form = DistrictForm()
    
    # Get all regions for filter dropdown
    regions = Region.objects.all().order_by('name')
    
    context = {
        'districts': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'form': form,
        'regions': regions,
        'current_page_size': page_size,
        'search_query': search_query,
        'selected_region_id': region_id,
    }
    
    return render(request, 'properties/manage_districts.html', context)


@login_required
def manage_property_types(request):
    """View for managing property types with pagination"""
    property_types = PropertyType.objects.all().order_by('name')
    # Server-side search
    search_query = request.GET.get('search', '').strip()
    if search_query:
        property_types = property_types.filter(
            Q(name__icontains=search_query) | Q(description__icontains=search_query)
        )
    
    # Pagination
    page_size = request.GET.get('page_size', '5')  # Default to 5 items per page
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50, 100]:
            page_size = 5
    except (ValueError, TypeError):
        page_size = 5
    
    paginator = Paginator(property_types, page_size)
    page_number = request.GET.get('page')
    
    try:
        page_obj = paginator.get_page(page_number)
    except:
        page_obj = paginator.get_page(1)
    
    if request.method == 'POST':
        form = PropertyTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Property type created successfully!')
            return redirect('properties:manage_property_types')
    else:
        form = PropertyTypeForm()
    
    context = {
        'property_types': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'form': form,
        'current_page_size': page_size,
        'search_query': search_query,
    }
    
    return render(request, 'properties/manage_property_types.html', context)


@login_required
def manage_amenities(request):
    """View for managing amenities with pagination and search"""
    amenities = Amenity.objects.all().order_by('name')
    
    # Server-side search
    search_query = request.GET.get('search', '').strip()
    if search_query:
        amenities = amenities.filter(
            Q(name__icontains=search_query) | Q(description__icontains=search_query) | Q(icon__icontains=search_query)
        )
    
    # Pagination
    page_size = request.GET.get('page_size', '5')  # Default to 5 items per page
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50, 100]:
            page_size = 5
    except (ValueError, TypeError):
        page_size = 5
    
    paginator = Paginator(amenities, page_size)
    page_number = request.GET.get('page')
    
    try:
        page_obj = paginator.get_page(page_number)
    except:
        page_obj = paginator.get_page(1)
    
    if request.method == 'POST':
        form = AmenityForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Amenity created successfully!')
            return redirect('properties:manage_amenities')
    else:
        form = AmenityForm()
    
    context = {
        'amenities': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'form': form,
        'current_page_size': page_size,
        'search_query': search_query,
    }
    
    return render(request, 'properties/manage_amenities.html', context)


@login_required
def property_dashboard(request):
    """Dashboard view with property statistics - all data from database, no hardcoded values"""
    from django.utils import timezone
    from datetime import timedelta
    from django.db.models import Sum
    
    # MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        # Property owner: only see own properties (excluding unavailable)
        properties_queryset = Property.objects.filter(owner=request.user).exclude(status='unavailable')
        bookings_queryset = Booking.objects.filter(property_obj__owner=request.user).exclude(booking_status='cancelled')
        payments_queryset = Payment.objects.filter(booking__property_obj__owner=request.user).exclude(booking__booking_status='cancelled')
    else:
        # Admin/staff: see all properties (excluding unavailable)
        properties_queryset = Property.objects.exclude(status='unavailable')
        bookings_queryset = Booking.objects.exclude(booking_status='cancelled')
        payments_queryset = Payment.objects.exclude(booking__booking_status='cancelled')
    
    # Overall statistics - all from database
    total_properties = properties_queryset.count()
    available_properties = properties_queryset.filter(status='available').count()
    rented_properties = properties_queryset.filter(status='rented').count()
    maintenance_properties = properties_queryset.filter(status='under_maintenance').count()
    
    # Booking statistics - exclude cancelled bookings
    total_bookings = bookings_queryset.count()
    active_bookings = bookings_queryset.filter(booking_status__in=['confirmed', 'checked_in']).count()
    completed_bookings = bookings_queryset.filter(booking_status='checked_out').count()
    pending_bookings = bookings_queryset.filter(booking_status='pending').count()
    
    # Revenue statistics - from actual payments, exclude cancelled bookings
    total_revenue = payments_queryset.aggregate(total=Sum('amount'))['total'] or 0
    paid_bookings = bookings_queryset.filter(payment_status='paid')
    revenue_from_bookings = sum(booking.total_amount for booking in paid_bookings)
    
    # Current month revenue
    current_month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    monthly_revenue = payments_queryset.filter(
        payment_date__gte=current_month_start
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Property statistics by type - respect multi-tenancy
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        property_types_stats = PropertyType.objects.filter(
            properties__owner=request.user
        ).annotate(
            property_count=Count('properties', filter=Q(properties__owner=request.user) & ~Q(properties__status='unavailable'))
        ).distinct().order_by('-property_count')
    else:
        property_types_stats = PropertyType.objects.annotate(
            property_count=Count('properties', filter=~Q(properties__status='unavailable'))
        ).order_by('-property_count')
    
    # Property statistics by region - respect multi-tenancy
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        regions_stats = Region.objects.filter(
            properties__owner=request.user
        ).annotate(
            property_count=Count('properties', filter=Q(properties__owner=request.user) & ~Q(properties__status='unavailable'))
        ).distinct().order_by('-property_count')
    else:
        regions_stats = Region.objects.annotate(
            property_count=Count('properties', filter=~Q(properties__status='unavailable'))
        ).order_by('-property_count')
    
    # Average rent - from actual property data
    average_rent = properties_queryset.aggregate(
        avg_rent=Avg('rent_amount')
    )['avg_rent'] or 0
    
    # Recent properties - respect multi-tenancy
    recent_properties = properties_queryset.select_related(
        'property_type', 'region', 'owner'
    ).prefetch_related('images').order_by('-created_at')[:5]
    
    # Recent bookings - exclude cancelled
    recent_bookings = bookings_queryset.select_related(
        'property_obj', 'customer', 'property_obj__property_type'
    ).order_by('-created_at')[:5]
    
    # Today's check-ins and check-outs - exclude cancelled
    today = timezone.now().date()
    todays_checkins = bookings_queryset.filter(
        check_in_date=today,
        booking_status__in=['confirmed', 'checked_in']
    ).select_related('property_obj', 'customer').count()
    
    todays_checkouts = bookings_queryset.filter(
        check_out_date=today,
        booking_status__in=['confirmed', 'checked_in']
    ).select_related('property_obj', 'customer').count()
    
    # Upcoming bookings (next 7 days) - exclude cancelled
    week_from_now = today + timedelta(days=7)
    upcoming_bookings = bookings_queryset.filter(
        check_in_date__range=[today + timedelta(days=1), week_from_now],
        booking_status__in=['confirmed', 'pending']
    ).select_related('property_obj', 'customer').count()
    
    context = {
        'total_properties': total_properties,
        'available_properties': available_properties,
        'rented_properties': rented_properties,
        'maintenance_properties': maintenance_properties,
        'total_bookings': total_bookings,
        'active_bookings': active_bookings,
        'completed_bookings': completed_bookings,
        'pending_bookings': pending_bookings,
        'total_revenue': total_revenue,
        'revenue_from_bookings': revenue_from_bookings,
        'monthly_revenue': monthly_revenue,
        'property_types_stats': property_types_stats,
        'regions_stats': regions_stats,
        'average_rent': average_rent,
        'recent_properties': recent_properties,
        'recent_bookings': recent_bookings,
        'todays_checkins': todays_checkins,
        'todays_checkouts': todays_checkouts,
        'upcoming_bookings': upcoming_bookings,
    }
    
    return render(request, 'properties/dashboard.html', context)


# Management System Views

def validate_property_id(property_id):
    """Helper function to validate property ID and handle 'all' case"""
    if property_id == 'all':
        return None
    elif property_id:
        try:
            return int(property_id)
        except (ValueError, TypeError):
            return None
    return None


@login_required
def hotel_dashboard(request):
    """Hotel management dashboard"""
    from datetime import date, datetime
    
    # Get selected property from request parameter or session
    # Priority: URL parameter > Session
    url_property_id = request.GET.get('property_id')
    session_property_id = request.session.get('selected_hotel_property_id')
    
    # Use URL parameter if present, otherwise use session
    selected_property_id = url_property_id or session_property_id
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    # If no hotel is selected, ensure session is clear and redirect to selection page
    if not selected_property_id:
        # Double-check session is clear (in case of stale data)
        if 'selected_hotel_property_id' in request.session:
            del request.session['selected_hotel_property_id']
            request.session.modified = True
        messages.info(request, 'Please select a hotel to view the dashboard.')
        return redirect('properties:hotel_select_property')
    
    # Get hotel properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        # Property owner: only see own hotels
        hotel_properties = Property.objects.filter(property_type__name__iexact='hotel', owner=request.user)
    else:
        # Admin/staff: see all hotels
        hotel_properties = Property.objects.filter(property_type__name__iexact='hotel')
    
    # Filter data based on selected property
    if selected_property_id:
        try:
            # MULTI-TENANCY: Ensure owner can only access their own properties
            if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='hotel', owner=request.user)
            else:
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='hotel')
            # Store selected property in session for persistence
            # Save to session if it came from URL (new selection) or if session doesn't match
            if url_property_id or request.session.get('selected_hotel_property_id') != selected_property_id:
                request.session['selected_hotel_property_id'] = selected_property_id
                request.session.modified = True
            
            # Get stats for selected property - exclude cancelled bookings
            bookings = Booking.objects.filter(property_obj=selected_property).exclude(booking_status='cancelled')
            total_bookings = bookings.count()
            active_bookings = bookings.filter(booking_status__in=['confirmed', 'checked_in']).count()
            revenue = sum(booking.total_amount for booking in bookings.filter(payment_status='paid'))
            
            # Get today's check-ins and check-outs for selected property - exclude cancelled
            today = date.today()
            todays_checkins = bookings.filter(check_in_date=today).exclude(booking_status='cancelled').select_related('customer')
            todays_checkouts = bookings.filter(check_out_date=today).exclude(booking_status='cancelled').select_related('customer')
            
            # Get room status for selected property - count actual Room objects, not property.total_rooms field
            # Sync room status before counting to ensure accuracy
            rooms = Room.objects.filter(property_obj=selected_property)
            # Sync room status from bookings to ensure accuracy
            for room in rooms:
                room.sync_status_from_bookings()
            total_rooms = rooms.count()  # Use actual count of Room objects
            available_rooms = rooms.filter(status='available').count()
            occupied_rooms = rooms.filter(status='occupied').count()
            maintenance_rooms = rooms.filter(status='maintenance').count()
            
            # Get recent payments for selected property - exclude cancelled bookings
            recent_payments = Payment.objects.filter(
                booking__property_obj=selected_property
            ).exclude(booking__booking_status='cancelled').select_related('booking__customer').order_by('-payment_date')[:5]
            
        except Property.DoesNotExist:
            # Property not found - clear invalid session value
            if 'selected_hotel_property_id' in request.session:
                del request.session['selected_hotel_property_id']
                request.session.modified = True
            selected_property = None
            total_rooms = 0
            total_bookings = 0
            active_bookings = 0
            revenue = 0
            todays_checkins = Booking.objects.none()
            todays_checkouts = Booking.objects.none()
            available_rooms = 0
            occupied_rooms = 0
            maintenance_rooms = 0
            recent_payments = Payment.objects.none()
    else:
        selected_property = None
        # Get stats for all hotels (filtered by owner if not admin)
        # Count actual Room objects, not property.total_rooms field
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            total_rooms = Room.objects.filter(property_obj__property_type__name__iexact='hotel', property_obj__owner=request.user).count()
        else:
            total_rooms = Room.objects.filter(property_obj__property_type__name__iexact='hotel').count()
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            bookings = Booking.objects.filter(property_obj__property_type__name__iexact='hotel', property_obj__owner=request.user).exclude(booking_status='cancelled')
        else:
            bookings = Booking.objects.filter(property_obj__property_type__name__iexact='hotel').exclude(booking_status='cancelled')
        total_bookings = bookings.count()
        active_bookings = bookings.filter(booking_status__in=['confirmed', 'checked_in']).count()
        revenue = sum(booking.total_amount for booking in bookings.filter(payment_status='paid'))
        
        # Get today's check-ins and check-outs for all hotels - exclude cancelled
        today = date.today()
        todays_checkins = bookings.filter(check_in_date=today).exclude(booking_status='cancelled').select_related('customer')
        todays_checkouts = bookings.filter(check_out_date=today).exclude(booking_status='cancelled').select_related('customer')
        
        # Get room status for all hotels (filtered by owner if not admin)
        # Sync room status before counting to ensure accuracy
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            rooms = Room.objects.filter(property_obj__property_type__name__iexact='hotel', property_obj__owner=request.user)
        else:
            rooms = Room.objects.filter(property_obj__property_type__name__iexact='hotel')
        # Sync room status from bookings to ensure accuracy
        for room in rooms:
            room.sync_status_from_bookings()
        available_rooms = rooms.filter(status='available').count()
        occupied_rooms = rooms.filter(status='occupied').count()
        maintenance_rooms = rooms.filter(status='maintenance').count()
        
        # Get recent payments for all hotels (filtered by owner if not admin) - exclude cancelled bookings
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            recent_payments = Payment.objects.filter(
                booking__property_obj__property_type__name__iexact='hotel',
                booking__property_obj__owner=request.user
            ).exclude(booking__booking_status='cancelled').select_related('booking__customer').order_by('-payment_date')[:5]
        else:
            recent_payments = Payment.objects.filter(
                booking__property_obj__property_type__name__iexact='hotel'
            ).exclude(booking__booking_status='cancelled').select_related('booking__customer').order_by('-payment_date')[:5]
    
    context = {
        'hotel_properties': hotel_properties,
        'selected_property': selected_property,
        'total_rooms': total_rooms,
        'total_bookings': total_bookings,
        'active_bookings': active_bookings,
        'revenue': revenue,
        'is_single_property_mode': bool(selected_property_id),
        'todays_checkins': todays_checkins,
        'todays_checkouts': todays_checkouts,
        'available_rooms': available_rooms,
        'occupied_rooms': occupied_rooms,
        'maintenance_rooms': maintenance_rooms,
        'recent_payments': recent_payments,
    }
    
    return render(request, 'properties/hotel_dashboard.html', context)


@login_required
def hotel_bookings(request):
    """Hotel bookings management"""
    # Get selected property from request parameter or session
    # Priority: URL parameter > Session
    url_property_id = request.GET.get('property_id')
    session_property_id = request.session.get('selected_hotel_property_id')
    
    # Use URL parameter if present, otherwise use session
    selected_property_id = url_property_id or session_property_id
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    # Get hotel properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        hotel_properties = Property.objects.filter(property_type__name__iexact='hotel', owner=request.user)
    else:
        hotel_properties = Property.objects.filter(property_type__name__iexact='hotel')
    
    # Get search and filter parameters
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    payment_filter = request.GET.get('payment_status', '')
    show_deleted = request.GET.get('show_deleted', 'false').lower() == 'true'  # Filter for soft-deleted bookings
    
    # Get bookings from properties.Booking (web admin bookings)
    # By default, exclude soft-deleted bookings unless show_deleted is True
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        bookings_query = Booking.objects.filter(property_obj__property_type__name__iexact='hotel', property_obj__owner=request.user)
        doc_bookings_query = DocumentBooking.objects.filter(property_ref__property_type__name__iexact='hotel', property_ref__owner=request.user).select_related('property_ref', 'tenant').order_by('-created_at')
    else:
        bookings_query = Booking.objects.filter(property_obj__property_type__name__iexact='hotel')
        doc_bookings_query = DocumentBooking.objects.filter(property_ref__property_type__name__iexact='hotel').select_related('property_ref', 'tenant').order_by('-created_at')
    
    # Filter soft-deleted bookings based on show_deleted parameter
    if not show_deleted:
        bookings_query = bookings_query.filter(is_deleted=False)
    
    bookings_query = bookings_query.select_related('customer', 'property_obj').order_by('-created_at')
    
    # Automatic checkout: Check and update bookings where stay period has ended
    from django.utils import timezone
    today = timezone.now().date()
    auto_checked_out = bookings_query.filter(
        check_out_date__lt=today,
        booking_status__in=['pending', 'confirmed', 'checked_in']
    ).update(
        booking_status='checked_out',
        checked_out_at=timezone.now()
    )
    
    # Filter by selected property if specified
    selected_property = None
    if selected_property_id:
        try:
            # MULTI-TENANCY: Ensure owner can only access their own properties
            if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='hotel', owner=request.user)
            else:
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='hotel')
            # Store selected property in session only if it came from URL parameter (new selection)
            if url_property_id:
                request.session['selected_hotel_property_id'] = selected_property_id
                request.session.modified = True
        except Property.DoesNotExist:
            selected_property = None
    
    # Combine bookings from both models using helper function
    all_bookings = combine_bookings(
        bookings_query,
        doc_bookings_query,
        'hotel',
        selected_property=selected_property,
        search_query=search_query,
        status_filter=status_filter,
        payment_filter=payment_filter,
        user=request.user
    )
    
    # Sync total_amount with calculated_total_amount for consistency
    # This ensures bookings show the correct amount based on room's base_rate
    # NOTE: Venues are EXCLUDED - they use time-based calculation, not date-based
    from decimal import Decimal
    for booking in all_bookings:
        # Only process properties.Booking instances (not DocumentBookingWrapper)
        if hasattr(booking, 'calculated_total_amount') and hasattr(booking, 'total_amount'):
            # Skip venues - they use manually entered total_amount from time-based calculation
            # Hotel bookings view only processes hotel bookings, but add check for safety
            is_venue = booking.property_obj and booking.property_obj.property_type.name.lower() == 'venue'
            if is_venue:
                continue  # Skip venue bookings - do not sync
            
            calculated_total = booking.calculated_total_amount if booking.calculated_total_amount and booking.calculated_total_amount > 0 else Decimal('0')
            stored_total = booking.total_amount or Decimal('0')
            
            # Sync if there's a mismatch (calculated uses room.base_rate, stored might be outdated)
            # This only applies to hotel bookings, NOT venues
            if calculated_total > 0 and abs(stored_total - calculated_total) > Decimal('0.01'):
                booking.total_amount = calculated_total
                booking.save()
    
    # Pagination
    from django.core.paginator import Paginator
    page_size = request.GET.get('page_size', '10')
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50]:
            page_size = 10
    except (ValueError, TypeError):
        page_size = 10
    
    paginator = Paginator(all_bookings, page_size)
    page_number = request.GET.get('page')
    
    try:
        bookings_page = paginator.get_page(page_number)
    except:
        bookings_page = paginator.get_page(1)
    
    context = {
        'hotel_properties': hotel_properties,
        'selected_property': selected_property,
        'bookings': bookings_page,
        'is_single_property_mode': bool(selected_property_id),
        'search_query': search_query,
        'status_filter': status_filter,
        'payment_filter': payment_filter,
        'show_deleted': show_deleted,
        'status_filter': status_filter,
        'payment_filter': payment_filter,
    }
    return render(request, 'properties/hotel_bookings.html', context)


@login_required
def hotel_rooms(request):
    """Hotel room status management"""
    from django.core.paginator import Paginator
    from django.shortcuts import redirect
    
    # Get hotel properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        hotel_properties = Property.objects.filter(property_type__name__iexact='hotel', owner=request.user)
    else:
        hotel_properties = Property.objects.filter(property_type__name__iexact='hotel')
    
    # For admins/staff: Must select a hotel first
    if request.user.is_staff or request.user.is_superuser:
        # Get selected property from request parameter or session
        # Priority: URL parameter > Session
        url_property_id = request.GET.get('property_id')
        session_property_id = request.session.get('selected_hotel_property_id')
        
        # Use URL parameter if present, otherwise use session
        selected_property_id = url_property_id or session_property_id
        
        # Validate selected_property_id - handle 'all' case
        selected_property_id = validate_property_id(selected_property_id)
        
        # If no hotel selected, redirect to selection page
        if not selected_property_id:
            return redirect('properties:hotel_select_property')
        
        # Filter rooms based on selected property
        try:
            selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='hotel')
            rooms_queryset = Room.objects.filter(property_obj=selected_property).select_related('property_obj', 'current_booking')
            
            # Sync room statuses from bookings to ensure accuracy
            for room in rooms_queryset:
                room.sync_status_from_bookings()
            
            # Refresh property status based on room availability
            selected_property.refresh_status_from_activity()
            
            # Store selected property in session only if it came from URL parameter (new selection)
            if url_property_id:
                request.session['selected_hotel_property_id'] = selected_property_id
                request.session.modified = True
        except Property.DoesNotExist:
            selected_property = None
            rooms_queryset = Room.objects.none()
    else:
        # For individual users: Automatically show their hotel's rooms (they only have one)
        selected_property = hotel_properties.first()  # Get their only hotel
        
        if selected_property:
            selected_property_id = selected_property.id
            # Store in session for consistency
            request.session['selected_hotel_property_id'] = selected_property_id
            
            # Get rooms for their hotel
            rooms_queryset = Room.objects.filter(property_obj=selected_property).select_related('property_obj', 'current_booking')
            
            # Sync room statuses from bookings to ensure accuracy
            for room in rooms_queryset:
                room.sync_status_from_bookings()
            
            # Refresh property status based on room availability
            selected_property.refresh_status_from_activity()
        else:
            # No hotel found for this user
            selected_property = None
            selected_property_id = None
            rooms_queryset = Room.objects.none()
    
    # Pagination - default 5 items per page
    paginator = Paginator(rooms_queryset, 5)
    page_number = request.GET.get('page')
    rooms = paginator.get_page(page_number)
    
    context = {
        'hotel_properties': hotel_properties,
        'selected_property': selected_property,
        'rooms': rooms,
        'is_single_property_mode': bool(selected_property_id),
        'user': request.user,  # Pass user to template for admin check
    }
    return render(request, 'properties/hotel_rooms.html', context)


@login_required
def add_room(request):
    """Add a new room to a hotel"""
    print(f"=== ADD ROOM DEBUG ===")
    print(f"Request method: {request.method}")
    print(f"User: {request.user}")
    
    if request.method == 'POST':
        print(f"POST data: {dict(request.POST)}")
        try:
            # Get hotel property
            hotel_id = request.POST.get('hotel_id')
            print(f"Hotel ID from form: {hotel_id}")
            
            if not hotel_id:
                print("ERROR: No hotel_id provided")
                messages.error(request, 'Please select a hotel.')
                return redirect('properties:hotel_rooms')
                
            # MULTI-TENANCY: Ensure owner can only add rooms to their own hotels
            if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
                hotel = Property.objects.get(id=hotel_id, property_type__name__iexact='hotel', owner=request.user)
            else:
                hotel = Property.objects.get(id=hotel_id, property_type__name__iexact='hotel')
            print(f"Found hotel: {hotel.title} (ID: {hotel.id})")
            
            # Get form data
            room_number = request.POST.get('room_number')
            room_type = request.POST.get('room_type')
            floor_number = request.POST.get('floor_number')
            floor_number = int(floor_number) if floor_number else None  # Make floor number optional
            capacity = int(request.POST.get('capacity', 1))
            base_rate = float(request.POST.get('base_rate', 0))
            bed_type = request.POST.get('bed_type', '')
            amenities = request.POST.get('amenities', '')
            
            print(f"Room data: {room_number}, {room_type}, {floor_number}, {capacity}, {base_rate}")
            
            # Validate that base_rate is set and greater than 0
            # Each room MUST have its own price - no default/fallback pricing
            if not base_rate or base_rate <= 0:
                messages.error(request, f'Base Rate is required and must be greater than 0. Each room must have its own price set.')
                return redirect('properties:hotel_rooms')
            
            # Check if room number already exists for this hotel
            if Room.objects.filter(property_obj=hotel, room_number=room_number).exists():
                messages.error(request, f'Room {room_number} already exists in {hotel.title}. Please choose a different room number.')
                return redirect('properties:hotel_rooms')
            
            # Create room
            room = Room.objects.create(
                property_obj=hotel,
                room_number=room_number,
                room_type=room_type,
                floor_number=floor_number,
                capacity=capacity,
                base_rate=base_rate,  # This is the ONLY source of pricing for this room
                bed_type=bed_type,
                amenities=amenities,
                status='available'
            )
            
            print(f"SUCCESS: Created room {room.room_number} for hotel {hotel.title}")
            messages.success(request, f'Room {room.room_number} added successfully to {hotel.title}!')
            
            # Redirect back to hotel rooms page
            return redirect('properties:hotel_rooms')
            
        except Property.DoesNotExist:
            messages.error(request, 'Selected hotel not found.')
        except ValueError as e:
            messages.error(request, f'Invalid data provided: {str(e)}')
        except Exception as e:
            messages.error(request, f'Error creating room: {str(e)}')
    
    return redirect('properties:hotel_rooms')


@login_required
def hotel_customers(request):
    """Hotel customer management"""
    # Get selected property from request parameter or session
    # Priority: URL parameter > Session
    url_property_id = request.GET.get('property_id')
    session_property_id = request.session.get('selected_hotel_property_id')
    
    # Use URL parameter if present, otherwise use session
    selected_property_id = url_property_id or session_property_id
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    # Get hotel properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        hotel_properties = Property.objects.filter(property_type__name__iexact='hotel', owner=request.user)
    else:
        hotel_properties = Property.objects.filter(property_type__name__iexact='hotel')
    
    # Filter customers based on selected property
    if selected_property_id:
        try:
            # MULTI-TENANCY: Ensure owner can only access their own properties
            if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='hotel', owner=request.user)
            else:
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='hotel')
            # Get customers who have bookings for this hotel
            customers = Customer.objects.filter(
                customer_bookings__property_obj=selected_property
            ).distinct().select_related().prefetch_related('customer_bookings')
            # Store selected property in session only if it came from URL parameter (new selection)
            if url_property_id:
                request.session['selected_hotel_property_id'] = selected_property_id
                request.session.modified = True
        except Property.DoesNotExist:
            selected_property = None
            customers = Customer.objects.none()
    else:
        selected_property = None
        # Show all customers who have hotel bookings (filtered by owner if not admin)
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            customers = Customer.objects.filter(
                customer_bookings__property_obj__property_type__name__iexact='hotel',
                customer_bookings__property_obj__owner=request.user
            ).distinct().select_related().prefetch_related('customer_bookings')
        else:
            customers = Customer.objects.filter(
                customer_bookings__property_obj__property_type__name__iexact='hotel'
            ).distinct().select_related().prefetch_related('customer_bookings')
    
    # Add total spent calculation and hotel info for each customer
    for customer in customers:
        customer.total_spent = sum(booking.total_amount for booking in customer.customer_bookings.all())
        customer.total_bookings_count = customer.customer_bookings.count()
        customer.last_booking = customer.customer_bookings.order_by('-check_out_date').first()
        
        # Get hotel information for this customer
        if selected_property_id:
            # If viewing a specific hotel, use that hotel's info
            customer.hotel_name = selected_property.title
            customer.hotel_address = selected_property.address
        else:
            # If viewing all hotels, get the most recent hotel they booked
            recent_booking = customer.customer_bookings.filter(
                property_obj__property_type__name__iexact='hotel'
            ).order_by('-created_at').first()
            if recent_booking:
                customer.hotel_name = recent_booking.property_obj.title
                customer.hotel_address = recent_booking.property_obj.address
            else:
                customer.hotel_name = "No Hotel Bookings"
                customer.hotel_address = ""
    
    # Calculate statistics
    total_customers = customers.count()
    active_customers_count = customers.filter(is_active=True).count()
    repeat_customers_count = customers.filter(customer_bookings__isnull=False).distinct().count()
    vip_customers_count = customers.filter(notes__icontains='VIP').count()
    
    context = {
        'hotel_properties': hotel_properties,
        'selected_property': selected_property,
        'customers': customers,
        'is_single_property_mode': bool(selected_property_id),
        'total_customers': total_customers,
        'active_customers_count': active_customers_count,
        'repeat_customers_count': repeat_customers_count,
        'vip_customers_count': vip_customers_count,
    }
    return render(request, 'properties/hotel_customers.html', context)


@login_required
def hotel_payments(request):
    """Hotel payment management"""
    from django.db.models import Sum, Q
    from django.core.paginator import Paginator
    from collections import defaultdict
    from decimal import Decimal
    
    # Get selected property from request parameter or session
    # Priority: URL parameter > Session
    url_property_id = request.GET.get('property_id')
    session_property_id = request.session.get('selected_hotel_property_id')
    
    # Use URL parameter if present, otherwise use session
    selected_property_id = url_property_id or session_property_id
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    # Get hotel properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        hotel_properties = Property.objects.filter(property_type__name__iexact='hotel', owner=request.user)
    else:
        hotel_properties = Property.objects.filter(property_type__name__iexact='hotel')
    
    # Get bookings for the Record Payment dropdown (filtered by owner if not admin)
    if selected_property_id:
        try:
            # MULTI-TENANCY: Ensure owner can only access their own properties
            if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='hotel', owner=request.user)
            else:
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='hotel')
            bookings = Booking.objects.filter(
                property_obj=selected_property
            ).exclude(booking_status='cancelled').select_related('customer', 'property_obj').order_by('-created_at')
            # Store selected property in session only if it came from URL parameter (new selection)
            if url_property_id:
                request.session['selected_hotel_property_id'] = selected_property_id
                request.session.modified = True
        except Property.DoesNotExist:
            selected_property = None
            bookings = Booking.objects.none()
    else:
        selected_property = None
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            bookings = Booking.objects.filter(
                property_obj__property_type__name__iexact='hotel',
                property_obj__owner=request.user
            ).exclude(booking_status='cancelled').select_related('customer', 'property_obj').order_by('-created_at')
        else:
            bookings = Booking.objects.filter(
                property_obj__property_type__name__iexact='hotel'
            ).exclude(booking_status='cancelled').select_related('customer', 'property_obj').order_by('-created_at')
    
    # Get all hotel bookings (for grouping payments by booking) - exclude cancelled bookings
    if selected_property_id and selected_property:
        hotel_bookings = Booking.objects.filter(
            property_obj=selected_property
        ).exclude(booking_status='cancelled').select_related('customer', 'property_obj')
    else:
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            hotel_bookings = Booking.objects.filter(
                property_obj__property_type__name__iexact='hotel',
                property_obj__owner=request.user
            ).exclude(booking_status='cancelled').select_related('customer', 'property_obj')
        else:
            hotel_bookings = Booking.objects.filter(
                property_obj__property_type__name__iexact='hotel'
            ).exclude(booking_status='cancelled').select_related('customer', 'property_obj')
    
    # Group payments by booking to show booking-level payment information
    bookings_data = {}
    for booking in hotel_bookings:
        # Calculate actual totals from all payments for this booking
        all_payments = Payment.objects.filter(booking=booking).exclude(payment_type='refund')
        all_refunds = Payment.objects.filter(booking=booking, payment_type='refund')
        total_paid = all_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_refunded = all_refunds.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        net_paid = max(Decimal('0'), total_paid - total_refunded)
        
        # Get total amount using ONLY room's base_rate (no property rent_amount fallback)
        # For hotel bookings, room assignment is mandatory, so base_rate comes from room only
        # NOTE: Venues are handled separately - they use time-based calculation, not date-based
        is_venue = booking.property_obj and booking.property_obj.property_type.name.lower() == 'venue'
        stored_total = booking.total_amount or Decimal('0')
        
        if is_venue:
            # For venues: ALWAYS use stored total_amount (manually entered from form)
            # Venues calculate days from start_time/end_time, which is different from hotel/lodge
            # DO NOT sync or calculate for venues - use what was entered
            total_required = stored_total
        else:
            # For hotels/lodges: sync with calculated_total_amount if needed
            calculated_total = booking.calculated_total_amount if booking.calculated_total_amount and booking.calculated_total_amount > 0 else Decimal('0')
            
            # Sync total_amount with calculated_total_amount if there's a mismatch
            # This ensures consistency between bookings and payments views
            # calculated_total_amount uses booking.base_rate which now ONLY uses room.base_rate
            if calculated_total > 0 and abs(stored_total - calculated_total) > Decimal('0.01'):
                booking.total_amount = calculated_total
                booking.save()
            
            # Use calculated_total_amount if available, otherwise use stored total_amount
            # Both should now be based on room.base_rate only (no property rent_amount)
            total_required = calculated_total if calculated_total > 0 else stored_total
        
        remaining = max(Decimal('0'), total_required - net_paid)
        
        # Update booking's paid_amount if inconsistent
        if abs(booking.paid_amount - net_paid) > Decimal('0.01'):
            booking.paid_amount = net_paid
            booking.update_payment_status()
            booking.save()
        
        # Get payment count and last payment date
        payment_count = all_payments.count()
        last_payment = all_payments.order_by('-payment_date').first()
        
        bookings_data[booking.id] = {
            'booking': booking,
            'total_required': total_required,
            'paid_so_far': net_paid,
            'remaining': remaining,
            'payment_count': payment_count,
            'last_payment_date': last_payment.payment_date if last_payment else None,
            'last_payment_method': last_payment.payment_method if last_payment else None,
        }
    
    # Convert to list for display
    all_payments_data = list(bookings_data.values())
    
    # Sort by last payment date or booking creation date
    all_payments_data.sort(key=lambda x: x['last_payment_date'] or x['booking'].created_at, reverse=True)
    
    # Pagination
    page_size = request.GET.get('page_size', '5')  # Default to 5
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50]:
            page_size = 5
    except (ValueError, TypeError):
        page_size = 5
    
    paginator = Paginator(all_payments_data, page_size)
    page_number = request.GET.get('page')
    payments_page = paginator.get_page(page_number)
    
    # Calculate payment statistics
    total_payments = len(all_payments_data)
    total_amount = sum(item['total_required'] for item in all_payments_data)
    paid_bookings = sum(1 for item in all_payments_data if item['booking'].payment_status == 'paid')
    
    context = {
        'hotel_properties': hotel_properties,
        'selected_property': selected_property,
        'payments': payments_page,  # Paginated payment data (booking-level)
        'payments_all': all_payments_data,  # For stats
        'bookings': bookings,
        'total_payments': total_payments,
        'total_amount': total_amount,
        'paid_bookings': paid_bookings,
        'is_single_property_mode': bool(selected_property_id),
    }
    return render(request, 'properties/hotel_payments.html', context)


@login_required
def hotel_reports(request):
    """Hotel reports and analytics - redirect to simplified reports dashboard"""
    return redirect('reports:dashboard')


# Lodge Management Views
@login_required
def lodge_dashboard(request):
    """Lodge management dashboard"""
    # Get selected property from request parameter or session
    # Priority: URL parameter > Session
    url_property_id = request.GET.get('property_id')
    session_property_id = request.session.get('selected_lodge_property_id')
    
    # Use URL parameter if present, otherwise use session
    selected_property_id = url_property_id or session_property_id
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    # If no lodge is selected, ensure session is clear and redirect to selection page
    if not selected_property_id:
        # Double-check session is clear (in case of stale data)
        if 'selected_lodge_property_id' in request.session:
            del request.session['selected_lodge_property_id']
            request.session.modified = True
        messages.info(request, 'Please select a lodge to view the dashboard.')
        return redirect('properties:lodge_select_property')
    
    # Get lodge properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        lodge_properties = Property.objects.filter(property_type__name__iexact='lodge', owner=request.user)
    else:
        lodge_properties = Property.objects.filter(property_type__name__iexact='lodge')
    
    # Filter data based on selected property
    if selected_property_id:
        try:
            # MULTI-TENANCY: Ensure owner can only access their own properties
            if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='lodge', owner=request.user)
            else:
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='lodge')
            # Store selected property in session only if it came from URL parameter (new selection)
            if url_property_id:
                request.session['selected_lodge_property_id'] = selected_property_id
                request.session.modified = True
            
            # Get stats for selected property - exclude cancelled bookings
            total_rooms = selected_property.total_rooms or 0
            bookings = Booking.objects.filter(property_obj=selected_property).exclude(booking_status='cancelled')
            total_bookings = bookings.count()
            active_bookings = bookings.filter(booking_status__in=['confirmed', 'checked_in']).count()
            revenue = sum(booking.total_amount for booking in bookings.filter(payment_status='paid'))
            
            # Calculate room counts from Room model
            # Sync room status before counting to ensure accuracy
            from .models import Room
            property_rooms = Room.objects.filter(property_obj=selected_property, is_active=True)
            # Sync room status from bookings to ensure cancelled bookings don't show rooms as occupied
            for room in property_rooms:
                room.sync_status_from_bookings()
            available_rooms = property_rooms.filter(status='available').count()
            occupied_rooms = property_rooms.filter(status='occupied').count()
            maintenance_rooms = property_rooms.filter(status='maintenance').count()
            
            # Calculate percentages for progress bars
            if total_rooms > 0:
                available_percentage = (available_rooms / total_rooms) * 100
                occupied_percentage = (occupied_rooms / total_rooms) * 100
                maintenance_percentage = (maintenance_rooms / total_rooms) * 100
            else:
                available_percentage = 0
                occupied_percentage = 0
                maintenance_percentage = 0
            
            # Get today's check-ins with pagination - exclude cancelled bookings
            from datetime import date
            today = date.today()
            todays_checkins_query = bookings.filter(
                check_in_date=today
            ).exclude(booking_status='cancelled').select_related('customer', 'property_obj').order_by('-created_at')
        except Property.DoesNotExist:
            selected_property = None
            total_rooms = 0
            total_bookings = 0
            active_bookings = 0
            revenue = 0
            available_rooms = 0
            occupied_rooms = 0
            maintenance_rooms = 0
            available_percentage = 0
            occupied_percentage = 0
            maintenance_percentage = 0
            todays_checkins_query = Booking.objects.none()
    else:
        selected_property = None
        # Get stats for all lodges (filtered by owner if not admin)
        total_rooms = sum(prop.total_rooms or 0 for prop in lodge_properties)
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            bookings = Booking.objects.filter(property_obj__property_type__name__iexact='lodge', property_obj__owner=request.user).exclude(booking_status='cancelled')
        else:
            bookings = Booking.objects.filter(property_obj__property_type__name__iexact='lodge').exclude(booking_status='cancelled')
        total_bookings = bookings.count()
        active_bookings = bookings.filter(booking_status__in=['confirmed', 'checked_in']).count()
        revenue = sum(booking.total_amount for booking in bookings.filter(payment_status='paid'))
        
        # Calculate room counts from Room model for all lodges
        # Sync room status before counting to ensure accuracy
        from .models import Room
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            property_rooms = Room.objects.filter(property_obj__property_type__name__iexact='lodge', property_obj__owner=request.user, is_active=True)
        else:
            property_rooms = Room.objects.filter(property_obj__property_type__name__iexact='lodge', is_active=True)
        # Sync room status from bookings to ensure cancelled bookings don't show rooms as occupied
        for room in property_rooms:
            room.sync_status_from_bookings()
        available_rooms = property_rooms.filter(status='available').count()
        occupied_rooms = property_rooms.filter(status='occupied').count()
        maintenance_rooms = property_rooms.filter(status='maintenance').count()
        
        # Calculate percentages for progress bars
        if total_rooms > 0:
            available_percentage = (available_rooms / total_rooms) * 100
            occupied_percentage = (occupied_rooms / total_rooms) * 100
            maintenance_percentage = (maintenance_rooms / total_rooms) * 100
        else:
            available_percentage = 0
            occupied_percentage = 0
            maintenance_percentage = 0
        
        # Get today's check-ins for all lodges with pagination - exclude cancelled bookings
        from datetime import date
        today = date.today()
        todays_checkins_query = bookings.filter(
            check_in_date=today
        ).exclude(booking_status='cancelled').select_related('customer', 'property_obj').order_by('-created_at')
    
    # Pagination for today's check-ins
    from django.core.paginator import Paginator
    page_size = request.GET.get('page_size', '5')
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25]:
            page_size = 5
    except (ValueError, TypeError):
        page_size = 5
    
    paginator = Paginator(todays_checkins_query, page_size)
    page_number = request.GET.get('page', 1)
    
    try:
        todays_checkins = paginator.get_page(page_number)
    except:
        todays_checkins = paginator.get_page(1)
    
    context = {
        'lodge_properties': lodge_properties,
        'selected_property': selected_property,
        'total_rooms': total_rooms,
        'total_bookings': total_bookings,
        'active_bookings': active_bookings,
        'revenue': revenue,
        'is_single_property_mode': bool(selected_property_id),
        'todays_checkins': todays_checkins,
        'available_rooms': available_rooms,
        'occupied_rooms': occupied_rooms,
        'maintenance_rooms': maintenance_rooms,
        'available_percentage': round(available_percentage, 1),
        'occupied_percentage': round(occupied_percentage, 1),
        'maintenance_percentage': round(maintenance_percentage, 1),
    }
    return render(request, 'properties/lodge_dashboard.html', context)


@login_required
def lodge_bookings(request):
    """Lodge bookings management"""
    # Get selected property from session or request
    selected_property_id = request.session.get('selected_lodge_property_id') or request.GET.get('property_id')
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    # Get lodge properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        lodge_properties = Property.objects.filter(property_type__name__iexact='lodge', owner=request.user)
    else:
        lodge_properties = Property.objects.filter(property_type__name__iexact='lodge')
    
    # Get search and filter parameters
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    payment_filter = request.GET.get('payment_status', '')
    show_deleted = request.GET.get('show_deleted', 'false').lower() == 'true'  # Filter for soft-deleted bookings
    
    # Get bookings from properties.Booking (web admin bookings)
    # By default, exclude soft-deleted bookings unless show_deleted is True
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        bookings_query = Booking.objects.filter(property_obj__property_type__name__iexact='lodge', property_obj__owner=request.user)
        doc_bookings_query = DocumentBooking.objects.filter(property_ref__property_type__name__iexact='lodge', property_ref__owner=request.user).select_related('property_ref', 'tenant').order_by('-created_at')
    else:
        bookings_query = Booking.objects.filter(property_obj__property_type__name__iexact='lodge')
        doc_bookings_query = DocumentBooking.objects.filter(property_ref__property_type__name__iexact='lodge').select_related('property_ref', 'tenant').order_by('-created_at')
    
    # Filter soft-deleted bookings based on show_deleted parameter
    if not show_deleted:
        bookings_query = bookings_query.filter(is_deleted=False)
    
    bookings_query = bookings_query.select_related('customer', 'property_obj').order_by('-created_at')
    
    # Filter by selected property if specified
    selected_property = None
    if selected_property_id:
        try:
            # MULTI-TENANCY: Ensure owner can only access their own properties
            if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='lodge', owner=request.user)
            else:
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='lodge')
            # Store selected property in session
            request.session['selected_lodge_property_id'] = selected_property_id
        except Property.DoesNotExist:
            selected_property = None
    
    # Combine bookings from both models using helper function
    all_bookings = combine_bookings(
        bookings_query,
        doc_bookings_query,
        'lodge',
        selected_property=selected_property,
        search_query=search_query,
        status_filter=status_filter,
        payment_filter=payment_filter,
        user=request.user
    )
    
    # Sync total_amount with calculated_total_amount for consistency
    # This ensures bookings show the correct amount based on room's base_rate
    # NOTE: Venues are EXCLUDED - they use time-based calculation, not date-based
    from decimal import Decimal
    for booking in all_bookings:
        # Only process properties.Booking instances (not DocumentBookingWrapper)
        if hasattr(booking, 'calculated_total_amount') and hasattr(booking, 'total_amount'):
            # Skip venues - they use manually entered total_amount from time-based calculation
            # Lodge bookings view only processes lodge bookings, but add check for safety
            is_venue = booking.property_obj and booking.property_obj.property_type.name.lower() == 'venue'
            if is_venue:
                continue  # Skip venue bookings - do not sync
            
            calculated_total = booking.calculated_total_amount if booking.calculated_total_amount and booking.calculated_total_amount > 0 else Decimal('0')
            stored_total = booking.total_amount or Decimal('0')
            
            # Sync if there's a mismatch (calculated uses room.base_rate, stored might be outdated)
            # This only applies to lodge bookings, NOT venues
            if calculated_total > 0 and abs(stored_total - calculated_total) > Decimal('0.01'):
                booking.total_amount = calculated_total
                booking.save()
    
    # Pagination
    from django.core.paginator import Paginator
    page_size = request.GET.get('page_size', '10')
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50]:
            page_size = 10
    except (ValueError, TypeError):
        page_size = 10
    
    paginator = Paginator(all_bookings, page_size)
    page_number = request.GET.get('page')
    
    try:
        bookings_page = paginator.get_page(page_number)
    except:
        bookings_page = paginator.get_page(1)
    
    context = {
        'lodge_properties': lodge_properties,
        'selected_property': selected_property,
        'bookings': bookings_page,
        'is_single_property_mode': bool(selected_property_id),
        'search_query': search_query,
        'status_filter': status_filter,
        'payment_filter': payment_filter,
        'show_deleted': show_deleted,
        'status_filter': status_filter,
        'payment_filter': payment_filter,
    }
    return render(request, 'properties/lodge_bookings.html', context)


@login_required
def lodge_rooms(request):
    """Lodge room status management"""
    from django.shortcuts import redirect
    from django.urls import reverse
    
    # Get lodge properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        lodge_properties = Property.objects.filter(property_type__name__iexact='lodge', owner=request.user)
    else:
        lodge_properties = Property.objects.filter(property_type__name__iexact='lodge')
    
    # For admins/staff: Must select a lodge first
    if request.user.is_staff or request.user.is_superuser:
        # Get selected property from session or request
        selected_property_id = request.session.get('selected_lodge_property_id') or request.GET.get('property_id')
        
        # Validate selected_property_id - handle 'all' case
        selected_property_id = validate_property_id(selected_property_id)
        
        # If no lodge selected, redirect to selection page
        if not selected_property_id:
            return redirect('properties:lodge_select_property')
        
        # Filter rooms based on selected property
        try:
            selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='lodge')
            # Store selected property in session
            request.session['selected_lodge_property_id'] = selected_property_id
            
            # Get rooms for selected property
            rooms = Room.objects.filter(property_obj=selected_property).select_related('current_booking', 'current_booking__customer')
        except Property.DoesNotExist:
            selected_property = None
            rooms = Room.objects.none()
    else:
        # For individual users: Automatically show their lodge's rooms (they only have one)
        selected_property = lodge_properties.first()  # Get their only lodge
        
        if selected_property:
            selected_property_id = selected_property.id
            # Store in session for consistency
            request.session['selected_lodge_property_id'] = selected_property_id
            
            # Get rooms for their lodge
            rooms = Room.objects.filter(property_obj=selected_property).select_related('current_booking', 'current_booking__customer')
        else:
            # No lodge found for this user
            selected_property = None
            selected_property_id = None
            rooms = Room.objects.none()
    
    # Calculate room statistics
    total_rooms = rooms.count()
    available_rooms = rooms.filter(status='available').count()
    occupied_rooms = rooms.filter(status='occupied').count()
    maintenance_rooms = rooms.filter(status='maintenance').count()
    out_of_order_rooms = rooms.filter(status='out_of_order').count()
    
    # Get current bookings for occupied rooms (only for selected property if available)
    if selected_property:
        current_bookings = Booking.objects.filter(
            property_obj=selected_property,
            booking_status__in=['confirmed', 'checked_in']
        ).select_related('customer', 'property_obj')
    else:
        current_bookings = Booking.objects.none()
    
    # Group rooms by area/type for better organization
    rooms_by_type = {}
    for room in rooms:
        room_type = room.room_type or 'standard'
        if room_type not in rooms_by_type:
            rooms_by_type[room_type] = []
        rooms_by_type[room_type].append(room)
    
    context = {
        'lodge_properties': lodge_properties,
        'selected_property': selected_property,
        'rooms': rooms,
        'rooms_by_type': rooms_by_type,
        'total_rooms': total_rooms,
        'available_rooms': available_rooms,
        'occupied_rooms': occupied_rooms,
        'maintenance_rooms': maintenance_rooms,
        'out_of_order_rooms': out_of_order_rooms,
        'current_bookings': current_bookings,
        'is_single_property_mode': bool(selected_property_id),
        'user': request.user,  # Pass user to template for admin check
    }
    return render(request, 'properties/lodge_rooms.html', context)


@login_required
def lodge_customers(request):
    """Lodge customer management"""
    from django.db.models import Count, Sum, Max
    
    # Get selected property from session or request
    selected_property_id = request.session.get('selected_lodge_property_id') or request.GET.get('property_id')
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    # Get lodge properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        lodge_properties = Property.objects.filter(property_type__name__iexact='lodge', owner=request.user)
    else:
        lodge_properties = Property.objects.filter(property_type__name__iexact='lodge')
    
    # Filter customers based on selected property
    if selected_property_id:
        try:
            # MULTI-TENANCY: Ensure owner can only access their own properties
            if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='lodge', owner=request.user)
            else:
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='lodge')
            # Get customers who have bookings for this lodge
            customers_query = Customer.objects.filter(
                customer_bookings__property_obj=selected_property
            ).distinct().annotate(
                booking_count=Count('customer_bookings'),
                total_spent=Sum('customer_bookings__total_amount'),
                last_visit=Max('customer_bookings__check_in_date')
            ).select_related().order_by('-last_visit')
            # Store selected property in session
            request.session['selected_lodge_property_id'] = selected_property_id
            request.session.modified = True
            # Store selected property in session
            request.session['selected_lodge_property_id'] = selected_property_id
        except Property.DoesNotExist:
            selected_property = None
            customers_query = Customer.objects.none()
    else:
        selected_property = None
        # Show all customers who have lodge bookings (filtered by owner if not admin)
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            customers_query = Customer.objects.filter(
                customer_bookings__property_obj__property_type__name__iexact='lodge',
                customer_bookings__property_obj__owner=request.user
            ).distinct().annotate(
                booking_count=Count('customer_bookings'),
                total_spent=Sum('customer_bookings__total_amount'),
                last_visit=Max('customer_bookings__check_in_date')
            ).order_by('-last_visit')
        else:
            customers_query = Customer.objects.filter(
                customer_bookings__property_obj__property_type__name__iexact='lodge'
            ).distinct().annotate(
                booking_count=Count('customer_bookings'),
                total_spent=Sum('customer_bookings__total_amount'),
                last_visit=Max('customer_bookings__check_in_date')
            ).order_by('-last_visit')
    
    # Pagination
    from django.core.paginator import Paginator
    page_size = request.GET.get('page_size', '10')
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50]:
            page_size = 10
    except (ValueError, TypeError):
        page_size = 10
    
    paginator = Paginator(customers_query, page_size)
    page_number = request.GET.get('page')
    
    try:
        customers_page = paginator.get_page(page_number)
    except:
        customers_page = paginator.get_page(1)
    
    # Calculate statistics
    total_customers = paginator.count
    active_customers = customers_query.filter(is_active=True).count()
    
    context = {
        'lodge_properties': lodge_properties,
        'selected_property': selected_property,
        'customers': customers_page,
        'total_customers': total_customers,
        'active_customers': active_customers,
        'is_single_property_mode': bool(selected_property_id),
    }
    return render(request, 'properties/lodge_customers.html', context)


@login_required
def lodge_payments(request):
    """Lodge payment management - similar to hotel_payments"""
    from django.db.models import Sum, Q
    from django.core.paginator import Paginator
    from collections import defaultdict
    from decimal import Decimal
    
    # Get selected property from request parameter or session
    # Priority: URL parameter > Session
    url_property_id = request.GET.get('property_id')
    session_property_id = request.session.get('selected_lodge_property_id')
    
    # Use URL parameter if present, otherwise use session
    selected_property_id = url_property_id or session_property_id
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    # Get lodge properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        lodge_properties = Property.objects.filter(property_type__name__iexact='lodge', owner=request.user)
    else:
        lodge_properties = Property.objects.filter(property_type__name__iexact='lodge')
    
    # Get bookings for the Record Payment dropdown (filtered by owner if not admin)
    if selected_property_id:
        try:
            # MULTI-TENANCY: Ensure owner can only access their own properties
            if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='lodge', owner=request.user)
            else:
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='lodge')
            bookings = Booking.objects.filter(
                property_obj=selected_property
            ).exclude(booking_status='cancelled').select_related('customer', 'property_obj').order_by('-created_at')
            # Store selected property in session only if it came from URL parameter (new selection)
            if url_property_id:
                request.session['selected_lodge_property_id'] = selected_property_id
                request.session.modified = True
        except Property.DoesNotExist:
            selected_property = None
            bookings = Booking.objects.none()
    else:
        selected_property = None
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            bookings = Booking.objects.filter(
                property_obj__property_type__name__iexact='lodge',
                property_obj__owner=request.user
            ).exclude(booking_status='cancelled').select_related('customer', 'property_obj').order_by('-created_at')
        else:
            bookings = Booking.objects.filter(
                property_obj__property_type__name__iexact='lodge'
            ).exclude(booking_status='cancelled').select_related('customer', 'property_obj').order_by('-created_at')
    
    # Get all lodge bookings (for grouping payments by booking) - exclude cancelled bookings
    if selected_property_id and selected_property:
        lodge_bookings = Booking.objects.filter(
            property_obj=selected_property
        ).exclude(booking_status='cancelled').select_related('customer', 'property_obj')
    else:
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            lodge_bookings = Booking.objects.filter(
                property_obj__property_type__name__iexact='lodge',
                property_obj__owner=request.user
            ).exclude(booking_status='cancelled').select_related('customer', 'property_obj')
        else:
            lodge_bookings = Booking.objects.filter(
                property_obj__property_type__name__iexact='lodge'
            ).exclude(booking_status='cancelled').select_related('customer', 'property_obj')
    
    # Group payments by booking to show booking-level payment information
    bookings_data = {}
    for booking in lodge_bookings:
        # Calculate actual totals from all payments for this booking
        all_payments = Payment.objects.filter(booking=booking).exclude(payment_type='refund')
        all_refunds = Payment.objects.filter(booking=booking, payment_type='refund')
        total_paid = all_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_refunded = all_refunds.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        net_paid = max(Decimal('0'), total_paid - total_refunded)
        
        # Get total amount using ONLY room's base_rate (no property rent_amount fallback)
        # For lodge bookings, room assignment is mandatory, so base_rate comes from room only
        # NOTE: Venues are handled separately - they use time-based calculation, not date-based
        is_venue = booking.property_obj and booking.property_obj.property_type.name.lower() == 'venue'
        stored_total = booking.total_amount or Decimal('0')
        
        if is_venue:
            # For venues: ALWAYS use stored total_amount (manually entered from form)
            # Venues calculate days from start_time/end_time, which is different from lodge
            # DO NOT sync or calculate for venues - use what was entered
            total_required = stored_total
        else:
            # For lodges: sync with calculated_total_amount if needed
            calculated_total = booking.calculated_total_amount if booking.calculated_total_amount and booking.calculated_total_amount > 0 else Decimal('0')
            
            # Sync total_amount with calculated_total_amount if there's a mismatch
            # This ensures consistency between bookings and payments views
            # calculated_total_amount uses booking.base_rate which now ONLY uses room.base_rate
            if calculated_total > 0 and abs(stored_total - calculated_total) > Decimal('0.01'):
                booking.total_amount = calculated_total
                booking.save()
            
            # Use calculated_total_amount if available, otherwise use stored total_amount
            # Both should now be based on room.base_rate only (no property rent_amount)
            total_required = calculated_total if calculated_total > 0 else stored_total
        
        remaining = max(Decimal('0'), total_required - net_paid)
        
        # Update booking's paid_amount if inconsistent
        if abs(booking.paid_amount - net_paid) > Decimal('0.01'):
            booking.paid_amount = net_paid
            booking.update_payment_status()
            booking.save()
        
        # Get payment count and last payment date
        payment_count = all_payments.count()
        last_payment = all_payments.order_by('-payment_date').first()
        
        bookings_data[booking.id] = {
            'booking': booking,
            'total_required': total_required,
            'paid_so_far': net_paid,
            'remaining': remaining,
            'payment_count': payment_count,
            'last_payment_date': last_payment.payment_date if last_payment else None,
            'last_payment_method': last_payment.payment_method if last_payment else None,
        }
    
    # Convert to list for display
    all_payments_data = list(bookings_data.values())
    
    # Sort by last payment date or booking creation date
    all_payments_data.sort(key=lambda x: x['last_payment_date'] or x['booking'].created_at, reverse=True)
    
    # Pagination
    page_size = request.GET.get('page_size', '5')  # Default to 5
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50]:
            page_size = 5
    except (ValueError, TypeError):
        page_size = 5
    
    paginator = Paginator(all_payments_data, page_size)
    page_number = request.GET.get('page')
    payments_page = paginator.get_page(page_number)
    
    # Calculate payment statistics
    total_payments = len(all_payments_data)
    total_amount = sum(item['total_required'] for item in all_payments_data)
    paid_bookings = sum(1 for item in all_payments_data if item['booking'].payment_status == 'paid')
    
    context = {
        'lodge_properties': lodge_properties,
        'selected_property': selected_property,
        'payments': payments_page,  # Paginated payment data (booking-level)
        'payments_all': all_payments_data,  # For stats
        'bookings': bookings,
        'total_payments': total_payments,
        'total_amount': total_amount,
        'paid_bookings': paid_bookings,
        'is_single_property_mode': bool(selected_property_id),
    }
    return render(request, 'properties/lodge_payments.html', context)


@login_required
def lodge_reports(request):
    """Lodge reports and analytics"""
    from django.utils import timezone
    from datetime import datetime, timedelta
    from django.db.models import Sum, Count, Avg
    
    # Get selected property from session or request
    selected_property_id = request.session.get('selected_lodge_property_id') or request.GET.get('property_id')
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    # Get lodge properties
    lodge_properties = Property.objects.filter(property_type__name__iexact='lodge')
    
    # Filter data based on selected property
    if selected_property_id:
        try:
            selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='lodge')
            # Store selected property in session
            request.session['selected_lodge_property_id'] = selected_property_id
            
            # Get bookings and payments for selected property
            bookings = Booking.objects.filter(property_obj=selected_property)
            payments = Payment.objects.filter(booking__property_obj=selected_property)
            rooms = Room.objects.filter(property_obj=selected_property)
            customers = Customer.objects.filter(customer_bookings__property_obj=selected_property).distinct()
        except Property.DoesNotExist:
            selected_property = None
            bookings = Booking.objects.none()
            payments = Payment.objects.none()
            rooms = Room.objects.none()
            customers = Customer.objects.none()
    else:
        selected_property = None
        # Get data for all lodges (filtered by owner if not admin)
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            bookings = Booking.objects.filter(property_obj__property_type__name__iexact='lodge', property_obj__owner=request.user)
            payments = Payment.objects.filter(booking__property_obj__property_type__name__iexact='lodge', booking__property_obj__owner=request.user)
            rooms = Room.objects.filter(property_obj__property_type__name__iexact='lodge', property_obj__owner=request.user)
            customers = Customer.objects.filter(customer_bookings__property_obj__property_type__name__iexact='lodge', customer_bookings__property_obj__owner=request.user).distinct()
        else:
            bookings = Booking.objects.filter(property_obj__property_type__name__iexact='lodge')
            payments = Payment.objects.filter(booking__property_obj__property_type__name__iexact='lodge')
            rooms = Room.objects.filter(property_obj__property_type__name__iexact='lodge')
            customers = Customer.objects.filter(customer_bookings__property_obj__property_type__name__iexact='lodge').distinct()
    
    # Calculate key metrics
    total_rooms = rooms.count()
    available_rooms = rooms.filter(status='available').count()
    occupied_rooms = rooms.filter(status='occupied').count()
    
    # Calculate occupancy rate
    if total_rooms > 0:
        occupancy_rate = round((occupied_rooms / total_rooms) * 100, 1)
    else:
        occupancy_rate = 0
    
    # Revenue calculations
    total_revenue = payments.aggregate(total=Sum('amount'))['total'] or 0
    today_revenue = payments.filter(payment_date__date=timezone.now().date()).aggregate(total=Sum('amount'))['total'] or 0
    
    # Monthly revenue
    current_month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    monthly_revenue = payments.filter(payment_date__gte=current_month_start).aggregate(total=Sum('amount'))['total'] or 0
    
    # Booking statistics
    total_bookings = bookings.count()
    active_bookings = bookings.filter(booking_status__in=['confirmed', 'checked_in']).count()
    
    # Customer statistics
    total_customers = customers.count()
    
    # Activity statistics (based on room types and booking patterns)
    room_type_stats = list(rooms.values('room_type').annotate(
        count=Count('id'),
        avg_rate=Avg('base_rate')
    ).order_by('-count'))
    
    # Calculate revenue for each room type
    for stat in room_type_stats:
        stat['revenue'] = stat['count'] * (stat['avg_rate'] or 0)
    
    # Payment method statistics
    payment_method_stats = list(payments.values('payment_method').annotate(
        count=Count('id'),
        total_amount=Sum('amount')
    ).order_by('-total_amount'))
    
    # Calculate percentage for each payment method
    for stat in payment_method_stats:
        if total_revenue > 0:
            stat['percentage'] = round((stat['total_amount'] / total_revenue) * 100, 1)
        else:
            stat['percentage'] = 0
    
    # Top customers by spending
    top_customers = customers.annotate(
        total_spent=Sum('customer_bookings__booking_payments__amount'),
        booking_count=Count('customer_bookings')
    ).filter(total_spent__gt=0).order_by('-total_spent')[:5]
    
    # Recent bookings for activity analysis
    recent_bookings = bookings.select_related('customer', 'property_obj').order_by('-created_at')[:10]
    
    # Revenue trend data (last 7 days)
    revenue_trend = []
    labels = []
    for i in range(7):
        date = timezone.now().date() - timedelta(days=i)
        daily_revenue = payments.filter(payment_date__date=date).aggregate(total=Sum('amount'))['total'] or 0
        revenue_trend.insert(0, float(daily_revenue))
        labels.insert(0, date.strftime('%a'))
    
    context = {
        'lodge_properties': lodge_properties,
        'selected_property': selected_property,
        'total_rooms': total_rooms,
        'available_rooms': available_rooms,
        'occupied_rooms': occupied_rooms,
        'occupancy_rate': occupancy_rate,
        'total_revenue': total_revenue,
        'today_revenue': today_revenue,
        'monthly_revenue': monthly_revenue,
        'total_bookings': total_bookings,
        'active_bookings': active_bookings,
        'total_customers': total_customers,
        'room_type_stats': room_type_stats,
        'payment_method_stats': payment_method_stats,
        'top_customers': top_customers,
        'recent_bookings': recent_bookings,
        'revenue_trend': revenue_trend,
        'revenue_labels': labels,
        'is_single_property_mode': bool(selected_property_id),
    }
    return render(request, 'properties/lodge_reports.html', context)


@login_required
def lodge_reports_export(request):
    """Export lodge reports in various formats"""
    from django.http import HttpResponse
    import csv
    from datetime import datetime, timedelta
    
    # Get parameters
    report_type = request.GET.get('type', 'monthly')
    date_range = request.GET.get('range', 'this_month')
    properties = request.GET.get('properties', 'all')
    format_type = request.GET.get('format', 'excel')
    
    # Calculate date range (same as house_reports_export)
    now = datetime.now()
    if date_range == 'this_month':
        start_date = now.replace(day=1)
        end_date = now
    elif date_range == 'last_month':
        last_month = now.replace(day=1) - timedelta(days=1)
        start_date = last_month.replace(day=1)
        end_date = last_month
    elif date_range == 'this_quarter':
        quarter_start = now.month - (now.month - 1) % 3
        start_date = now.replace(month=quarter_start, day=1)
        end_date = now
    elif date_range == 'last_quarter':
        quarter_start = now.month - (now.month - 1) % 3
        last_quarter_end = now.replace(month=quarter_start, day=1) - timedelta(days=1)
        quarter_start = last_quarter_end.month - (last_quarter_end.month - 1) % 3
        start_date = last_quarter_end.replace(month=quarter_start, day=1)
        end_date = last_quarter_end
    elif date_range == 'this_year':
        start_date = now.replace(month=1, day=1)
        end_date = now
    elif date_range == 'last_year':
        start_date = now.replace(year=now.year-1, month=1, day=1)
        end_date = now.replace(year=now.year-1, month=12, day=31)
    elif date_range.startswith('custom_'):
        try:
            date_parts = date_range.split('_')
            start_date = datetime.strptime(date_parts[1], '%Y-%m-%d')
            end_date = datetime.strptime(date_parts[2], '%Y-%m-%d')
        except (ValueError, IndexError):
            start_date = now.replace(day=1)
            end_date = now
    else:
        start_date = now.replace(day=1)
        end_date = now
    
    # Get lodge properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        lodge_properties = Property.objects.filter(property_type__name__iexact='lodge', owner=request.user)
    else:
        lodge_properties = Property.objects.filter(property_type__name__iexact='lodge')
    
    if properties != 'all':
        try:
            property_ids = [int(p) for p in properties.split(',')]
            lodge_properties = lodge_properties.filter(id__in=property_ids)
        except (ValueError, TypeError):
            pass
    
    # Get bookings and payments
    bookings = Booking.objects.filter(
        property_obj__in=lodge_properties,
        created_at__gte=start_date,
        created_at__lte=end_date
    )
    payments = Payment.objects.filter(
        booking__property_obj__in=lodge_properties,
        payment_date__gte=start_date,
        payment_date__lte=end_date
    )
    
    # Generate Excel file
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_type.title()} Report"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write report header
        ws['A1'] = 'Lodge Property Management Report'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f'Report Type: {report_type.title()}'
        ws['B2'] = f'Date Range: {date_range}'
        ws['A3'] = f'Generated At: {now.strftime("%Y-%m-%d %H:%M:%S")}'
        ws['B3'] = f'Period: {start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}'
        
        # Add summary data
        row = 5
        ws['A5'] = 'Property'
        ws['B5'] = 'Bookings'
        ws['C5'] = 'Revenue (Tsh)'
        ws['D5'] = 'Customers'
        
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        row += 1
        total_bookings = 0
        total_revenue = 0
        total_customers = 0
        
        for prop in lodge_properties:
            prop_bookings = bookings.filter(property_obj=prop)
            prop_payments = payments.filter(booking__property_obj=prop)
            prop_revenue = sum(p.amount for p in prop_payments)
            customers = Customer.objects.filter(customer_bookings__property_obj=prop).distinct().count()
            
            ws.cell(row=row, column=1, value=prop.title)
            ws.cell(row=row, column=2, value=prop_bookings.count())
            ws.cell(row=row, column=3, value=prop_revenue)
            ws.cell(row=row, column=4, value=customers)
            
            # Accumulate totals
            total_bookings += prop_bookings.count()
            total_revenue += prop_revenue
            total_customers += customers
            
            row += 1
        
        # Add total row
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=2, value=total_bookings)
        ws.cell(row=row, column=3, value=total_revenue)
        ws.cell(row=row, column=4, value=total_customers)
        
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = total_font
            cell.fill = total_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="lodge_{report_type}_report_{date_range}.xlsx"'
        
    except ImportError:
        # Excel format is required - raise error if openpyxl is not available
        from django.http import JsonResponse
        return JsonResponse({
            'error': 'Excel generation requires openpyxl library. Please install it: pip install openpyxl'
        }, status=500)
    
    return response


# Venue Management Views
@login_required
def venue_dashboard(request):
    """Venue management dashboard"""
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    # Get selected property from session or request
    selected_property_id = request.session.get('selected_venue_property_id') or request.GET.get('property_id')
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    # Get venue properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        venue_properties = Property.objects.filter(property_type__name__iexact='venue', owner=request.user)
    else:
        venue_properties = Property.objects.filter(property_type__name__iexact='venue')
    
    # Filter data based on selected property
    if selected_property_id:
        try:
            # MULTI-TENANCY: Ensure owner can only access their own properties
            if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='venue', owner=request.user)
            else:
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='venue')
            # Store selected property in session
            request.session['selected_venue_property_id'] = selected_property_id
            
            # Get stats for selected property - exclude cancelled bookings
            capacity = selected_property.capacity or 0
            bookings = Booking.objects.filter(property_obj=selected_property).exclude(booking_status='cancelled')
            total_bookings = bookings.count()
            active_bookings = bookings.filter(booking_status__in=['confirmed', 'checked_in']).count()
            revenue = sum(booking.total_amount for booking in bookings.filter(payment_status='paid'))
            
            # Get today's events - exclude cancelled bookings
            today = timezone.now().date()
            todays_events = bookings.filter(check_in_date=today).exclude(booking_status='cancelled').select_related('customer')
            
            # Add pagination for today's events
            from django.core.paginator import Paginator
            todays_events_paginator = Paginator(todays_events, 5)  # Show 5 events per page
            todays_events_page = request.GET.get('todays_page')
            todays_events = todays_events_paginator.get_page(todays_events_page)
            
            # Get upcoming events (next 7 days) - exclude cancelled bookings
            week_from_now = today + timedelta(days=7)
            upcoming_events = bookings.filter(
                check_in_date__range=[today + timedelta(days=1), week_from_now]
            ).exclude(booking_status='cancelled').select_related('customer').order_by('check_in_date')[:5]
            
        except Property.DoesNotExist:
            selected_property = None
            capacity = 0
            total_bookings = 0
            active_bookings = 0
            revenue = 0
            todays_events = Booking.objects.none()
            upcoming_events = Booking.objects.none()
    else:
        selected_property = None
        # Get stats for all venues (filtered by owner if not admin)
        capacity = sum(prop.capacity or 0 for prop in venue_properties)
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            bookings = Booking.objects.filter(property_obj__property_type__name__iexact='venue', property_obj__owner=request.user).exclude(booking_status='cancelled')
        else:
            bookings = Booking.objects.filter(property_obj__property_type__name__iexact='venue').exclude(booking_status='cancelled')
        total_bookings = bookings.count()
        active_bookings = bookings.filter(booking_status__in=['confirmed', 'checked_in']).count()
        revenue = sum(booking.total_amount for booking in bookings.filter(payment_status='paid'))
        
        # Get today's events for all venues - exclude cancelled bookings
        today = timezone.now().date()
        todays_events = bookings.filter(check_in_date=today).exclude(booking_status='cancelled').select_related('customer', 'property_obj')
        
        # Add pagination for today's events
        from django.core.paginator import Paginator
        todays_events_paginator = Paginator(todays_events, 5)  # Show 5 events per page
        todays_events_page = request.GET.get('todays_page')
        todays_events = todays_events_paginator.get_page(todays_events_page)
        
        # Get upcoming events (next 7 days) - exclude cancelled bookings
        week_from_now = today + timedelta(days=7)
        upcoming_events = bookings.filter(
            check_in_date__range=[today + timedelta(days=1), week_from_now]
        ).exclude(booking_status='cancelled').select_related('customer', 'property_obj').order_by('check_in_date')[:5]
    
    # Calculate additional metrics - exclude cancelled bookings from occupancy calculation
    occupancy_rate = 0
    if capacity > 0:
        total_guests = sum(booking.number_of_guests or 0 for booking in bookings.filter(booking_status__in=['confirmed', 'checked_in']).exclude(booking_status='cancelled'))
        occupancy_rate = min(100, (total_guests / capacity) * 100)
    
    context = {
        'venue_properties': venue_properties,
        'selected_property': selected_property,
        'capacity': capacity,
        'total_bookings': total_bookings,
        'active_bookings': active_bookings,
        'revenue': revenue,
        'occupancy_rate': occupancy_rate,
        'todays_events': todays_events,
        'upcoming_events': upcoming_events,
        'is_single_property_mode': bool(selected_property_id),
    }
    return render(request, 'properties/venue_dashboard.html', context)


@login_required
def venue_bookings(request):
    """Venue bookings management"""
    auto_complete_venue_bookings(request.user)
    # Get selected property from session or request
    selected_property_id = request.session.get('selected_venue_property_id') or request.GET.get('property_id')
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    # Get venue properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        venue_properties = Property.objects.filter(property_type__name__iexact='venue', owner=request.user)
    else:
        venue_properties = Property.objects.filter(property_type__name__iexact='venue')
    
    # Get search and filter parameters
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    payment_filter = request.GET.get('payment_status', '')
    show_deleted = request.GET.get('show_deleted', 'false').lower() == 'true'  # Filter for soft-deleted bookings
    
    # Get bookings from properties.Booking (web admin bookings)
    # By default, exclude soft-deleted bookings unless show_deleted is True
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        bookings_query = Booking.objects.filter(property_obj__property_type__name__iexact='venue', property_obj__owner=request.user)
        doc_bookings_query = DocumentBooking.objects.filter(property_ref__property_type__name__iexact='venue', property_ref__owner=request.user).select_related('property_ref', 'tenant').order_by('-created_at')
    else:
        bookings_query = Booking.objects.filter(property_obj__property_type__name__iexact='venue')
        doc_bookings_query = DocumentBooking.objects.filter(property_ref__property_type__name__iexact='venue').select_related('property_ref', 'tenant').order_by('-created_at')
    
    # Filter soft-deleted bookings based on show_deleted parameter
    if not show_deleted:
        bookings_query = bookings_query.filter(is_deleted=False)
    
    bookings_query = bookings_query.select_related('customer', 'property_obj').order_by('-created_at')
    
    # Filter by selected property if specified
    selected_property = None
    if selected_property_id:
        try:
            # MULTI-TENANCY: Ensure owner can only access their own properties
            if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='venue', owner=request.user)
            else:
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='venue')
            # Store selected property in session
            request.session['selected_venue_property_id'] = selected_property_id
        except Property.DoesNotExist:
            selected_property = None
    
    # Combine bookings from both models using helper function
    all_bookings = combine_bookings(
        bookings_query,
        doc_bookings_query,
        'venue',
        selected_property=selected_property,
        search_query=search_query,
        status_filter=status_filter,
        payment_filter=payment_filter,
        user=request.user
    )
    
    context = {
        'venue_properties': venue_properties,
        'selected_property': selected_property,
        'bookings': all_bookings,
        'is_single_property_mode': bool(selected_property_id),
        'search_query': search_query,
        'status_filter': status_filter,
        'payment_filter': payment_filter,
        'show_deleted': show_deleted,
    }
    return render(request, 'properties/venue_bookings.html', context)


@login_required
def venue_availability(request):
    """Venue availability management"""
    auto_complete_venue_bookings(request.user)
    from datetime import timedelta

    today = timezone.now().date()
    default_end = today + timedelta(days=6)

    context = {
        'initial_start_date': today.strftime('%Y-%m-%d'),
        'initial_end_date': default_end.strftime('%Y-%m-%d'),
    }
    return render(request, 'properties/venue_availability.html', context)


@login_required
def venue_customers(request):
    """Venue customer management"""
    from django.core.paginator import Paginator
    
    # Get selected property from session or request
    selected_property_id = request.session.get('selected_venue_property_id') or request.GET.get('property_id')
    
    # Get venue properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        venue_properties = Property.objects.filter(property_type__name__iexact='venue', owner=request.user)
    else:
        venue_properties = Property.objects.filter(property_type__name__iexact='venue')
    
    # Filter customers based on selected property
    if selected_property_id:
        try:
            # MULTI-TENANCY: Ensure owner can only access their own properties
            if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='venue', owner=request.user)
            else:
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='venue')
            # Get customers who have bookings for this venue
            customers = Customer.objects.filter(
                customer_bookings__property_obj=selected_property
            ).distinct().select_related().prefetch_related('customer_bookings')
            # Store selected property in session
            request.session['selected_venue_property_id'] = selected_property_id
        except Property.DoesNotExist:
            selected_property = None
            customers = Customer.objects.none()
    else:
        selected_property = None
        # Show all customers who have venue bookings (filtered by owner if not admin)
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            customers = Customer.objects.filter(
                customer_bookings__property_obj__property_type__name__iexact='venue',
                customer_bookings__property_obj__owner=request.user
            ).distinct().select_related().prefetch_related('customer_bookings')
        else:
            customers = Customer.objects.filter(
                customer_bookings__property_obj__property_type__name__iexact='venue'
            ).distinct().select_related().prefetch_related('customer_bookings')
    
    # Add pagination
    paginator = Paginator(customers, 10)  # Show 10 customers per page
    page_number = request.GET.get('page')
    customers = paginator.get_page(page_number)
    
    # Add latest booking date and total spent for each customer
    for customer in customers:
        # Get latest booking date (filtered by owner if not admin)
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            latest_booking = customer.customer_bookings.filter(
                property_obj__property_type__name__iexact='venue',
                property_obj__owner=request.user
            ).order_by('-check_out_date').first()
        else:
            latest_booking = customer.customer_bookings.filter(
                property_obj__property_type__name__iexact='venue'
            ).order_by('-check_out_date').first()
        customer.latest_booking_date = latest_booking.check_out_date if latest_booking else None
        
        # Calculate total spent on venue bookings (filtered by owner if not admin)
        from django.db.models import Sum
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            total_spent = customer.customer_bookings.filter(
                property_obj__property_type__name__iexact='venue',
                property_obj__owner=request.user
            ).aggregate(total=Sum('total_amount'))['total']
        else:
            total_spent = customer.customer_bookings.filter(
                property_obj__property_type__name__iexact='venue'
            ).aggregate(total=Sum('total_amount'))['total']
        customer.total_spent = total_spent if total_spent else 0
    
    # Calculate statistics (filtered by owner if not admin)
    total_customers = paginator.count
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        active_customers_count = Customer.objects.filter(
            customer_bookings__property_obj__property_type__name__iexact='venue',
            customer_bookings__property_obj__owner=request.user,
            is_active=True
        ).distinct().count()
        repeat_customers_count = Customer.objects.filter(
            customer_bookings__property_obj__property_type__name__iexact='venue',
            customer_bookings__property_obj__owner=request.user
        ).annotate(booking_count=Count('customer_bookings')).filter(booking_count__gt=1).count()
        vip_customers_count = Customer.objects.filter(
            customer_bookings__property_obj__property_type__name__iexact='venue',
            customer_bookings__property_obj__owner=request.user,
            notes__icontains='VIP'
        ).distinct().count()
    else:
        active_customers_count = Customer.objects.filter(
            customer_bookings__property_obj__property_type__name__iexact='venue',
            is_active=True
        ).distinct().count()
        repeat_customers_count = Customer.objects.filter(
            customer_bookings__property_obj__property_type__name__iexact='venue'
        ).annotate(booking_count=Count('customer_bookings')).filter(booking_count__gt=1).count()
        vip_customers_count = Customer.objects.filter(
            customer_bookings__property_obj__property_type__name__iexact='venue',
            notes__icontains='VIP'
        ).distinct().count()
    
    context = {
        'venue_properties': venue_properties,
        'selected_property': selected_property,
        'customers': customers,
        'is_single_property_mode': bool(selected_property_id),
        'total_customers': total_customers,
        'active_customers_count': active_customers_count,
        'repeat_customers_count': repeat_customers_count,
        'vip_customers_count': vip_customers_count,
    }
    return render(request, 'properties/venue_customers.html', context)


@login_required
def venue_payments(request):
    """Venue payment management - matches hotel/lodge pattern"""
    from django.db.models import Sum, Q
    from django.core.paginator import Paginator
    from collections import defaultdict
    from decimal import Decimal
    
    # Get selected property from request parameter or session
    url_property_id = request.GET.get('property_id')
    session_property_id = request.session.get('selected_venue_property_id')
    
    # Use URL parameter if present, otherwise use session
    selected_property_id = url_property_id or session_property_id
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    # Get venue properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        venue_properties = Property.objects.filter(property_type__name__iexact='venue', owner=request.user)
    else:
        venue_properties = Property.objects.filter(property_type__name__iexact='venue')
    
    # Get bookings for the Record Payment dropdown (filtered by owner if not admin)
    if selected_property_id:
        try:
            # MULTI-TENANCY: Ensure owner can only access their own properties
            if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='venue', owner=request.user)
            else:
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='venue')
            bookings = Booking.objects.filter(
                property_obj=selected_property
            ).exclude(booking_status='cancelled').select_related('customer', 'property_obj').order_by('-created_at')
            # Store selected property in session only if it came from URL parameter (new selection)
            if url_property_id:
                request.session['selected_venue_property_id'] = selected_property_id
                request.session.modified = True
        except Property.DoesNotExist:
            selected_property = None
            bookings = Booking.objects.none()
    else:
        selected_property = None
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            bookings = Booking.objects.filter(
                property_obj__property_type__name__iexact='venue',
                property_obj__owner=request.user
            ).exclude(booking_status='cancelled').select_related('customer', 'property_obj').order_by('-created_at')
        else:
            bookings = Booking.objects.filter(
                property_obj__property_type__name__iexact='venue'
            ).exclude(booking_status='cancelled').select_related('customer', 'property_obj').order_by('-created_at')
    
    # Get all venue bookings (for grouping payments by booking) - exclude cancelled bookings
    if selected_property_id and selected_property:
        venue_bookings = Booking.objects.filter(
            property_obj=selected_property
        ).exclude(booking_status='cancelled').select_related('customer', 'property_obj')
    else:
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            venue_bookings = Booking.objects.filter(
                property_obj__property_type__name__iexact='venue',
                property_obj__owner=request.user
            ).exclude(booking_status='cancelled').select_related('customer', 'property_obj')
        else:
            venue_bookings = Booking.objects.filter(
                property_obj__property_type__name__iexact='venue'
            ).exclude(booking_status='cancelled').select_related('customer', 'property_obj')
    
    # Group payments by booking to show booking-level payment information
    bookings_data = {}
    for booking in venue_bookings:
        # Calculate actual totals from all payments for this booking
        all_payments = Payment.objects.filter(booking=booking).exclude(payment_type='refund')
        all_refunds = Payment.objects.filter(booking=booking, payment_type='refund')
        total_paid = all_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        total_refunded = all_refunds.aggregate(total=Sum('amount'))['total'] or Decimal('0')
        net_paid = max(Decimal('0'), total_paid - total_refunded)
        
        # Get total amount - handle venues separately (venues use time-based calculation, not date-based)
        is_venue = booking.property_obj and booking.property_obj.property_type.name.lower() == 'venue'
        stored_total = booking.total_amount or Decimal('0')
        
        if is_venue:
            # For venues: ALWAYS use stored total_amount (manually entered from form)
            # Venues calculate days from start_time/end_time, which is different from hotel/lodge
            # DO NOT sync or calculate for venues - use what was entered
            total_required = stored_total
        else:
            # For hotel/lodge: sync with calculated_total_amount if needed
            calculated_total = booking.calculated_total_amount if booking.calculated_total_amount and booking.calculated_total_amount > 0 else Decimal('0')
            
            # Sync total_amount with calculated_total_amount if there's a mismatch
            # Only sync if calculated_total is valid (> 0) and stored_total is 0 or significantly different
            if calculated_total > 0:
                # If stored_total is 0 or empty, use calculated_total
                if stored_total == 0:
                    booking.total_amount = calculated_total
                    booking.save()
                # If there's a significant mismatch (more than 1% difference), sync
                elif abs(stored_total - calculated_total) > Decimal('0.01') and abs(stored_total - calculated_total) / max(stored_total, calculated_total) > Decimal('0.01'):
                    # Only sync if calculated is reasonable (not zero and not way off)
                    booking.total_amount = calculated_total
                    booking.save()
            
            # Use calculated_total_amount if available, otherwise use stored total_amount
            total_required = calculated_total if calculated_total > 0 else stored_total
        
        remaining = max(Decimal('0'), total_required - net_paid)
        
        # Update booking's paid_amount if inconsistent
        if abs(booking.paid_amount - net_paid) > Decimal('0.01'):
            booking.paid_amount = net_paid
            booking.update_payment_status()
            booking.save()
        
        # Get payment count and last payment date
        payment_count = all_payments.count()
        last_payment = all_payments.order_by('-payment_date').first()
        
        bookings_data[booking.id] = {
            'booking': booking,
            'total_required': total_required,
            'paid_so_far': net_paid,
            'remaining': remaining,
            'payment_count': payment_count,
            'last_payment_date': last_payment.payment_date if last_payment else None,
            'last_payment_method': last_payment.payment_method if last_payment else None,
        }
    
    # Convert to list for display
    all_payments_data = list(bookings_data.values())
    
    # Sort by last payment date or booking creation date
    all_payments_data.sort(key=lambda x: x['last_payment_date'] or x['booking'].created_at, reverse=True)
    
    # Pagination
    page_size = request.GET.get('page_size', '5')  # Default to 5
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50]:
            page_size = 5
    except (ValueError, TypeError):
        page_size = 5
    
    paginator = Paginator(all_payments_data, page_size)
    page_number = request.GET.get('page')
    payments_page = paginator.get_page(page_number)
    
    # Calculate payment statistics
    total_payments = len(all_payments_data)
    total_amount = sum(item['total_required'] for item in all_payments_data)
    paid_bookings = sum(1 for item in all_payments_data if item['booking'].payment_status == 'paid')
    
    context = {
        'venue_properties': venue_properties,
        'selected_property': selected_property,
        'payments': payments_page,  # Paginated payment data (booking-level)
        'payments_all': all_payments_data,  # For stats
        'bookings': bookings,
        'total_payments': total_payments,
        'total_amount': total_amount,
        'paid_bookings': paid_bookings,
        'is_single_property_mode': bool(selected_property_id),
    }
    return render(request, 'properties/venue_payments.html', context)


@login_required
def venue_reports(request):
    """Venue reports and analytics"""
    from django.db.models import Sum, Count
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    # Get venue properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        venue_properties = Property.objects.filter(property_type__name__iexact='venue', owner=request.user)
    else:
        venue_properties = Property.objects.filter(property_type__name__iexact='venue')
    
    # Calculate statistics
    total_venues = venue_properties.count()
    
    # Calculate monthly revenue (current month, filtered by owner if not admin)
    current_month = timezone.now().replace(day=1)
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        monthly_revenue = Payment.objects.filter(
            booking__property_obj__property_type__name__iexact='venue',
            booking__property_obj__owner=request.user,
            payment_date__gte=current_month,
            status='active'
        ).aggregate(total=Sum('amount'))['total'] or 0
    else:
        monthly_revenue = Payment.objects.filter(
            booking__property_obj__property_type__name__iexact='venue',
            payment_date__gte=current_month,
            status='active'
        ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Calculate total events (filtered by owner if not admin)
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        total_events = Booking.objects.filter(
            property_obj__property_type__name__iexact='venue',
            property_obj__owner=request.user
        ).count()
    else:
        total_events = Booking.objects.filter(
            property_obj__property_type__name__iexact='venue'
        ).count()
    
    # Calculate active customers (filtered by owner if not admin)
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        active_customers = Customer.objects.filter(
            customer_bookings__property_obj__property_type__name__iexact='venue',
            customer_bookings__property_obj__owner=request.user,
            is_active=True
        ).distinct().count()
    else:
        active_customers = Customer.objects.filter(
            customer_bookings__property_obj__property_type__name__iexact='venue',
            is_active=True
        ).distinct().count()
    
    context = {
        'venue_properties': venue_properties,
        'total_venues': total_venues,
        'monthly_revenue': monthly_revenue,
        'total_events': total_events,
        'active_customers': active_customers,
    }
    return render(request, 'properties/venue_reports.html', context)


@login_required
def venue_reports_export(request):
    """Export venue reports in various formats"""
    from django.http import HttpResponse
    import csv
    from datetime import datetime, timedelta
    
    # Get parameters
    report_type = request.GET.get('type', 'monthly')
    date_range = request.GET.get('range', 'this_month')
    venues = request.GET.get('venues', 'all')
    format_type = request.GET.get('format', 'excel')
    
    # Calculate date range (same as house_reports_export)
    now = datetime.now()
    if date_range == 'this_month':
        start_date = now.replace(day=1)
        end_date = now
    elif date_range == 'last_month':
        last_month = now.replace(day=1) - timedelta(days=1)
        start_date = last_month.replace(day=1)
        end_date = last_month
    elif date_range == 'this_quarter':
        quarter_start = now.month - (now.month - 1) % 3
        start_date = now.replace(month=quarter_start, day=1)
        end_date = now
    elif date_range == 'last_quarter':
        quarter_start = now.month - (now.month - 1) % 3
        last_quarter_end = now.replace(month=quarter_start, day=1) - timedelta(days=1)
        quarter_start = last_quarter_end.month - (last_quarter_end.month - 1) % 3
        start_date = last_quarter_end.replace(month=quarter_start, day=1)
        end_date = last_quarter_end
    elif date_range == 'this_year':
        start_date = now.replace(month=1, day=1)
        end_date = now
    elif date_range == 'last_year':
        start_date = now.replace(year=now.year-1, month=1, day=1)
        end_date = now.replace(year=now.year-1, month=12, day=31)
    elif date_range.startswith('custom_'):
        try:
            date_parts = date_range.split('_')
            start_date = datetime.strptime(date_parts[1], '%Y-%m-%d')
            end_date = datetime.strptime(date_parts[2], '%Y-%m-%d')
        except (ValueError, IndexError):
            start_date = now.replace(day=1)
            end_date = now
    else:
        start_date = now.replace(day=1)
        end_date = now
    
    # Get venue properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        venue_properties = Property.objects.filter(property_type__name__iexact='venue', owner=request.user)
    else:
        venue_properties = Property.objects.filter(property_type__name__iexact='venue')
    if venues != 'all':
        try:
            venue_ids = [int(v) for v in venues.split(',')]
            venue_properties = venue_properties.filter(id__in=venue_ids)
        except (ValueError, TypeError):
            pass
    
    # Get bookings and payments
    bookings = Booking.objects.filter(
        property_obj__in=venue_properties,
        created_at__gte=start_date,
        created_at__lte=end_date
    )
    payments = Payment.objects.filter(
        booking__property_obj__in=venue_properties,
        payment_date__gte=start_date,
        payment_date__lte=end_date
    )
    
    # Generate Excel file
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_type.title()} Report"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write report header
        ws['A1'] = 'Venue Property Management Report'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f'Report Type: {report_type.title()}'
        ws['B2'] = f'Date Range: {date_range}'
        ws['A3'] = f'Generated At: {now.strftime("%Y-%m-%d %H:%M:%S")}'
        ws['B3'] = f'Period: {start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}'
        
        # Add summary data
        row = 5
        ws['A5'] = 'Venue'
        ws['B5'] = 'Events'
        ws['C5'] = 'Revenue (Tsh)'
        ws['D5'] = 'Customers'
        
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        row += 1
        total_events = 0
        total_revenue = 0
        total_customers = 0
        
        for prop in venue_properties:
            prop_bookings = bookings.filter(property_obj=prop)
            prop_payments = payments.filter(booking__property_obj=prop)
            prop_revenue = sum(p.amount for p in prop_payments)
            customers = Customer.objects.filter(customer_bookings__property_obj=prop).distinct().count()
            
            ws.cell(row=row, column=1, value=prop.title)
            ws.cell(row=row, column=2, value=prop_bookings.count())
            ws.cell(row=row, column=3, value=prop_revenue)
            ws.cell(row=row, column=4, value=customers)
            
            # Accumulate totals
            total_events += prop_bookings.count()
            total_revenue += prop_revenue
            total_customers += customers
            
            row += 1
        
        # Add total row
        total_font = Font(bold=True, size=11)
        total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=2, value=total_events)
        ws.cell(row=row, column=3, value=total_revenue)
        ws.cell(row=row, column=4, value=total_customers)
        
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.font = total_font
            cell.fill = total_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="venue_{report_type}_report_{date_range}.xlsx"'
        
    except ImportError:
        # Excel format is required - raise error if openpyxl is not available
        from django.http import JsonResponse
        return JsonResponse({
            'error': 'Excel generation requires openpyxl library. Please install it: pip install openpyxl'
        }, status=500)
    
    return response


# House Management Views
@login_required
def house_dashboard(request):
    """House management dashboard"""
    # Get selected property from session or request
    selected_property_id = request.session.get('selected_house_property_id') or request.GET.get('property_id')
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    # Get house properties with pagination
    house_properties_query = Property.objects.filter(property_type__name__iexact='house').select_related('property_type', 'region', 'owner').prefetch_related('bookings')
    
    # If a property is selected, filter to only show that property
    if selected_property_id:
        house_properties_query = house_properties_query.filter(id=selected_property_id)
    
    # Apply search filter if provided (only if no specific property is selected)
    search_query = request.GET.get('search', '').strip()
    if search_query and not selected_property_id:
        house_properties_query = house_properties_query.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(address__icontains=search_query)
        )
    
    # Apply status filter if provided (only if no specific property is selected)
    status_filter = request.GET.get('status', '')
    if status_filter and not selected_property_id:
        if status_filter == 'occupied':
            # Filter properties that have active bookings (confirmed or checked_in)
            house_properties_query = house_properties_query.filter(
                property_bookings__booking_status__in=['confirmed', 'checked_in']
            ).distinct()
        elif status_filter == 'available':
            # Filter properties that don't have active bookings
            house_properties_query = house_properties_query.exclude(
                property_bookings__booking_status__in=['confirmed', 'checked_in']
            ).distinct()
    
    # Pagination
    from django.core.paginator import Paginator
    page_size = request.GET.get('page_size', '5')
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50]:
            page_size = 5
    except (ValueError, TypeError):
        page_size = 5
    
    paginator = Paginator(house_properties_query, page_size)
    page_number = request.GET.get('page')
    
    try:
        house_properties_page = paginator.get_page(page_number)
    except:
        house_properties_page = paginator.get_page(1)
    
    # Get all house properties for stats (not paginated)
    house_properties = Property.objects.filter(property_type__name__iexact='house')
    
    # Filter data based on selected property
    if selected_property_id:
        try:
            selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='house')
            # Store selected property in session
            request.session['selected_house_property_id'] = selected_property_id
            
            # Get stats for selected property
            bedrooms = selected_property.bedrooms or 0
            bookings = Booking.objects.filter(property_obj=selected_property)
            total_bookings = bookings.count()
            active_bookings = bookings.filter(booking_status__in=['confirmed', 'checked_in']).count()
            revenue = sum(booking.total_amount for booking in bookings.filter(payment_status='paid'))
            # When a property is selected, total_houses should be 1
            total_houses = 1
        except Property.DoesNotExist:
            selected_property = None
            bedrooms = 0
            total_bookings = 0
            active_bookings = 0
            revenue = 0
            total_houses = house_properties.count()
    else:
        selected_property = None
        # Get stats for all houses
        bedrooms = sum(prop.bedrooms or 0 for prop in house_properties)
        bookings = Booking.objects.filter(property_obj__property_type__name__iexact='house')
        total_bookings = bookings.count()
        active_bookings = bookings.filter(booking_status__in=['confirmed', 'checked_in']).count()
        revenue = sum(booking.total_amount for booking in bookings.filter(payment_status='paid'))
        # Calculate additional statistics for all houses
        total_houses = house_properties.count()
    
    # Calculate percentages for progress bars
    occupied_percentage = 0
    available_percentage = 0
    if total_houses > 0:
        occupied_percentage = round((active_bookings / total_houses) * 100)
        available_percentage = round(((total_houses - active_bookings) / total_houses) * 100)
    
    # Get recent bookings for display
    recent_bookings = bookings.order_by('-created_at')[:5]
    
    # Get upcoming due dates (if using rent system)
    from datetime import date, timedelta
    today = date.today()
    upcoming_due_dates = []
    
    # Try to get upcoming rent due dates if rent system is available
    try:
        from rent.models import RentInvoice
        if selected_property:
            # Filter by selected property
            upcoming_invoices = RentInvoice.objects.filter(
                lease__property_ref=selected_property,
                due_date__gte=today,
                due_date__lte=today + timedelta(days=7),
                status__in=['draft', 'sent']
            ).select_related('lease__property_ref', 'tenant')[:5]
        else:
            # Filter for all houses
            upcoming_invoices = RentInvoice.objects.filter(
                lease__property_ref__property_type__name__iexact='house',
                due_date__gte=today,
                due_date__lte=today + timedelta(days=7),
                status__in=['draft', 'sent']
            ).select_related('lease__property_ref', 'tenant')[:5]
        upcoming_due_dates = upcoming_invoices
    except ImportError:
        # Rent system not available, use empty list
        upcoming_due_dates = []
    
    # Get overdue invoices
    overdue_invoices = []
    try:
        if selected_property:
            # Filter by selected property
            overdue_invoices = RentInvoice.objects.filter(
                lease__property_ref=selected_property,
                due_date__lt=today,
                status__in=['sent', 'overdue']
            ).select_related('lease__property_ref', 'tenant')[:5]
        else:
            # Filter for all houses
            overdue_invoices = RentInvoice.objects.filter(
                lease__property_ref__property_type__name__iexact='house',
                due_date__lt=today,
                status__in=['sent', 'overdue']
            ).select_related('lease__property_ref', 'tenant')[:5]
    except ImportError:
        overdue_invoices = []
    
    context = {
        'house_properties': house_properties_page,
        'house_properties_all': house_properties,  # For stats
        'selected_property': selected_property,
        'total_houses': total_houses,
        'bedrooms': bedrooms,
        'total_bookings': total_bookings,
        'active_bookings': active_bookings,
        'revenue': revenue,
        'recent_bookings': recent_bookings,
        'upcoming_due_dates': upcoming_due_dates,
        'overdue_invoices': overdue_invoices,
        'occupied_percentage': occupied_percentage,
        'available_percentage': available_percentage,
        'is_single_property_mode': bool(selected_property_id),
        'search_query': search_query,
        'status_filter': status_filter,
    }
    return render(request, 'properties/house_dashboard.html', context)


@login_required
def house_bookings(request):
    """House bookings management"""
    
    # Get selected property from URL parameter only (don't use session for house bookings)
    selected_property_id = request.GET.get('property_id')
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    # Get house properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        house_properties = Property.objects.filter(property_type__name__iexact='house', owner=request.user)
    else:
        house_properties = Property.objects.filter(property_type__name__iexact='house')
    
    # Get bookings from properties.Booking (web admin bookings)
    # Get bookings from documents.Booking (mobile app bookings) for house properties
    # Get leases from documents.Lease (mobile app leases) for house properties
    from documents.models import Lease
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        bookings_query = Booking.objects.filter(property_obj__property_type__name__iexact='house', property_obj__owner=request.user).select_related('customer', 'property_obj').order_by('-created_at')
        doc_bookings_query = DocumentBooking.objects.filter(property_ref__property_type__name__iexact='house', property_ref__owner=request.user).select_related('property_ref', 'tenant').order_by('-created_at')
        leases_query = Lease.objects.filter(property_ref__property_type__name__iexact='house', property_ref__owner=request.user).select_related('property_ref', 'tenant').order_by('-created_at')
    else:
        bookings_query = Booking.objects.filter(property_obj__property_type__name__iexact='house').select_related('customer', 'property_obj').order_by('-created_at')
        doc_bookings_query = DocumentBooking.objects.filter(property_ref__property_type__name__iexact='house').select_related('property_ref', 'tenant').order_by('-created_at')
        leases_query = Lease.objects.filter(property_ref__property_type__name__iexact='house').select_related('property_ref', 'tenant').order_by('-created_at')
    
    # Get search and filter parameters
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '')
    payment_filter = request.GET.get('payment_status', '')
    
    # Filter by selected property if specified
    selected_property = None
    if selected_property_id:
        try:
            # MULTI-TENANCY: Ensure owner can only access their own properties
            if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='house', owner=request.user)
            else:
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='house')
        except Property.DoesNotExist:
            selected_property = None
    
    # Filter leases by selected property and search
    if selected_property:
        leases_query = leases_query.filter(property_ref=selected_property)
    if search_query:
        leases_query = leases_query.filter(
            Q(tenant__first_name__icontains=search_query) |
            Q(tenant__last_name__icontains=search_query) |
            Q(tenant__email__icontains=search_query) |
            Q(tenant__username__icontains=search_query) |
            Q(property_ref__title__icontains=search_query) |
            Q(property_ref__address__icontains=search_query)
        )
    if status_filter:
        # Map booking status to lease status
        lease_status_map = {
            'pending': 'pending',
            'confirmed': 'active',
            'active': 'active',
            'cancelled': 'terminated',
        }
        lease_status = lease_status_map.get(status_filter, status_filter)
        leases_query = leases_query.filter(status=lease_status)
    
    # Combine bookings from both models using helper function
    all_bookings = combine_bookings(
        bookings_query,
        doc_bookings_query,
        'house',
        selected_property=selected_property,
        search_query=search_query,
        status_filter=status_filter,
        payment_filter=payment_filter,
        user=request.user
    )
    
    # Add wrapped leases to the bookings list
    wrapped_leases = [LeaseWrapper(lease) for lease in leases_query]
    all_bookings = sorted(
        list(all_bookings) + wrapped_leases,
        key=lambda x: x.created_at,
        reverse=True
    )
    
    # Pagination
    from django.core.paginator import Paginator
    page_size = request.GET.get('page_size', '5')
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50]:
            page_size = 5
    except (ValueError, TypeError):
        page_size = 5
    
    paginator = Paginator(all_bookings, page_size)
    page_number = request.GET.get('page')
    
    try:
        bookings_page = paginator.get_page(page_number)
    except:
        bookings_page = paginator.get_page(1)
    
    # Get all bookings for stats (use same filtered queries as main display)
    from documents.models import Lease
    if selected_property:
        bookings_all_prop = Booking.objects.filter(property_obj=selected_property)
        bookings_all_doc = DocumentBooking.objects.filter(property_ref=selected_property)
        leases_all = Lease.objects.filter(property_ref=selected_property)
    else:
        # Use the same filtering logic as the main queries (respects multi-tenancy)
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            bookings_all_prop = Booking.objects.filter(property_obj__property_type__name__iexact='house', property_obj__owner=request.user)
            bookings_all_doc = DocumentBooking.objects.filter(property_ref__property_type__name__iexact='house', property_ref__owner=request.user)
            leases_all = Lease.objects.filter(property_ref__property_type__name__iexact='house', property_ref__owner=request.user)
        else:
            bookings_all_prop = Booking.objects.filter(property_obj__property_type__name__iexact='house')
            bookings_all_doc = DocumentBooking.objects.filter(property_ref__property_type__name__iexact='house')
            leases_all = Lease.objects.filter(property_ref__property_type__name__iexact='house')
    
    # Calculate statistics (combine all three types: bookings, doc_bookings, and leases)
    total_bookings = bookings_all_prop.count() + bookings_all_doc.count() + leases_all.count()
    active_bookings = (
        bookings_all_prop.filter(booking_status__in=['confirmed', 'checked_in']).count() +
        bookings_all_doc.filter(status='confirmed').count() +
        leases_all.filter(status='active').count()
    )
    pending_bookings = (
        bookings_all_prop.filter(booking_status='pending').count() +
        bookings_all_doc.filter(status='pending').count() +
        leases_all.filter(status='pending').count()
    )
    total_revenue = (
        sum(booking.total_amount for booking in bookings_all_prop.filter(payment_status='paid')) +
        sum(booking.total_amount for booking in bookings_all_doc.filter(status='confirmed')) +
        sum(lease.rent_amount for lease in leases_all.filter(status='active'))
    )
    
    # Create bookings_all list for stats (wrapped document bookings and leases)
    bookings_all = list(bookings_all_prop) + [DocumentBookingWrapper(db) for db in bookings_all_doc] + [LeaseWrapper(lease) for lease in leases_all]
    
    context = {
        'house_properties': house_properties,
        'selected_property': selected_property,
        'bookings': bookings_page,
        'bookings_all': bookings_all,  # For stats
        'total_bookings': total_bookings,
        'active_bookings': active_bookings,
        'pending_bookings': pending_bookings,
        'total_revenue': total_revenue,
        'is_single_property_mode': bool(selected_property_id),
        'search_query': search_query,
        'status_filter': status_filter,
        'payment_filter': payment_filter,
    }
    return render(request, 'properties/house_bookings.html', context)


@login_required
def house_tenants(request):
    """House tenant management"""
    # Get selected property from session or request
    selected_property_id = request.session.get('selected_house_property_id') or request.GET.get('property_id')
    
    # Get search and filter parameters
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '').strip()
    tenant_type_filter = request.GET.get('tenant_type', '').strip()
    
    # Get house properties
    house_properties = Property.objects.filter(property_type__name__iexact='house')
    
    # Filter tenants based on selected property
    if selected_property_id:
        try:
            selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='house')
            # Get customers who have bookings for this house (tenants)
            tenants = Customer.objects.filter(
                customer_bookings__property_obj=selected_property
            ).distinct().select_related().prefetch_related('customer_bookings')
            # Store selected property in session
            request.session['selected_house_property_id'] = selected_property_id
        except Property.DoesNotExist:
            selected_property = None
            tenants = Customer.objects.none()
    else:
        selected_property = None
        # Show all customers who have house bookings (tenants)
        tenants = Customer.objects.filter(
            customer_bookings__property_obj__property_type__name__iexact='house'
        ).distinct().select_related().prefetch_related('customer_bookings')
    
    # Apply search filter
    if search_query:
        tenants = tenants.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )
    
    # Apply status filter (based on booking status)
    if status_filter:
        if status_filter == 'active':
            tenants = tenants.filter(customer_bookings__booking_status__in=['confirmed', 'checked_in'])
        elif status_filter == 'inactive':
            tenants = tenants.filter(customer_bookings__booking_status__in=['checked_out', 'cancelled'])
        elif status_filter == 'vip':
            # For now, we'll use a simple check - you can enhance this based on your VIP logic
            tenants = tenants.filter(customer_bookings__total_amount__gte=1000000)  # High rent = VIP
    
    # Apply tenant type filter (you can enhance this based on your tenant type logic)
    if tenant_type_filter:
        # This is a placeholder - you can implement tenant type logic based on your requirements
        pass
    
    # Order tenants
    tenants = tenants.order_by('-created_at')
    
    # Pagination
    from django.core.paginator import Paginator
    page_size = request.GET.get('page_size', '5')  # Default to 5
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50]:
            page_size = 5
    except (ValueError, TypeError):
        page_size = 5
    
    paginator = Paginator(tenants, page_size)
    page_number = request.GET.get('page')
    tenants_page = paginator.get_page(page_number)
    
    context = {
        'house_properties': house_properties,
        'selected_property': selected_property,
        'tenants': tenants_page,  # Paginated tenants
        'tenants_all': tenants,  # For stats
        'is_single_property_mode': bool(selected_property_id),
        'search_query': search_query,
        'status_filter': status_filter,
        'tenant_type_filter': tenant_type_filter,
    }
    return render(request, 'properties/house_tenants.html', context)


@login_required
def house_payments(request):
    """House payment management"""
    from .models import Payment
    
    # Get selected property from session or request
    selected_property_id = request.session.get('selected_house_property_id') or request.GET.get('property_id')
    
    # Get search and filter parameters
    search_query = request.GET.get('search', '').strip()
    status_filter = request.GET.get('status', '').strip()
    payment_type_filter = request.GET.get('payment_type', '').strip()
    
    # Get house properties
    house_properties = Property.objects.filter(property_type__name__iexact='house')
    
    # Filter payments based on selected property
    if selected_property_id:
        try:
            selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='house')
        except Property.DoesNotExist:
            selected_property = None
        
        if selected_property:
            # Get booking payments (properties.models.Payment - has booking, recorded_by, NO tenant)
            booking_payments = Payment.objects.filter(
                booking__property_obj=selected_property
            ).select_related('booking', 'booking__customer', 'recorded_by')
            
            # Get visit payments (payments.models.Payment - has tenant, recorded_by, NO booking)
            from .models import PropertyVisitPayment
            from payments.models import Payment as UnifiedPayment
            visit_payment_ids = PropertyVisitPayment.objects.filter(
                property=selected_property
            ).exclude(payment__isnull=True).values_list('payment_id', flat=True)
            visit_payments = UnifiedPayment.objects.filter(
                id__in=visit_payment_ids
            ).select_related('tenant', 'recorded_by').prefetch_related('property_visit_payments', 'property_visit_payments__property')
            
            # Apply filters to each queryset separately before combining
            # Search filter
            if search_query:
                from django.db.models import Q
                booking_payments = booking_payments.filter(
                    Q(booking__customer__first_name__icontains=search_query) |
                    Q(booking__customer__last_name__icontains=search_query) |
                    Q(booking__customer__email__icontains=search_query) |
                    Q(booking__customer__phone__icontains=search_query) |
                    Q(booking__property_obj__title__icontains=search_query) |
                    Q(payment_method__icontains=search_query) |
                    Q(notes__icontains=search_query)
                )
                visit_payments = visit_payments.filter(
                    Q(tenant__first_name__icontains=search_query) |
                    Q(tenant__last_name__icontains=search_query) |
                    Q(tenant__email__icontains=search_query) |
                    Q(property_visit_payments__property__title__icontains=search_query) |
                    Q(payment_method__icontains=search_query) |
                    Q(notes__icontains=search_query)
                )
            
            # Status filter
            if status_filter:
                if status_filter == 'paid':
                    booking_payments = booking_payments.filter(status='active')  # properties.Payment uses 'active' for paid
                    visit_payments = visit_payments.filter(status='completed')  # payments.Payment uses 'completed'
                elif status_filter == 'pending':
                    booking_payments = booking_payments.filter(status='active')  # Adjust as needed
                    visit_payments = visit_payments.filter(status='pending')
            
            # Payment type filter
            if payment_type_filter == 'visit':
                booking_payments = Payment.objects.none()  # Only show visit payments
            elif payment_type_filter and payment_type_filter != 'visit':
                visit_payments = UnifiedPayment.objects.none()  # Only show booking payments
            
            # Store selected property in session
            request.session['selected_house_property_id'] = selected_property_id
        else:
            booking_payments = Payment.objects.none()
            visit_payments = UnifiedPayment.objects.none()
    else:
        selected_property = None
        # Show all payments for house bookings (properties.models.Payment)
        booking_payments = Payment.objects.filter(
            booking__property_obj__property_type__name__iexact='house'
        ).select_related('booking', 'booking__customer', 'recorded_by')
        
        # Get visit payments for all house properties (payments.models.Payment)
        from .models import PropertyVisitPayment
        from payments.models import Payment as UnifiedPayment
        visit_payment_ids = PropertyVisitPayment.objects.filter(
            property__property_type__name__iexact='house'
        ).exclude(payment__isnull=True).values_list('payment_id', flat=True)
        visit_payments = UnifiedPayment.objects.filter(
            id__in=visit_payment_ids
        ).select_related('tenant', 'recorded_by').prefetch_related('property_visit_payments', 'property_visit_payments__property')
        
        # Apply filters to each queryset separately before combining
        # Search filter
        if search_query:
            from django.db.models import Q
            booking_payments = booking_payments.filter(
                Q(booking__customer__first_name__icontains=search_query) |
                Q(booking__customer__last_name__icontains=search_query) |
                Q(booking__customer__email__icontains=search_query) |
                Q(booking__customer__phone__icontains=search_query) |
                Q(booking__property_obj__title__icontains=search_query) |
                Q(payment_method__icontains=search_query) |
                Q(notes__icontains=search_query)
            )
            visit_payments = visit_payments.filter(
                Q(tenant__first_name__icontains=search_query) |
                Q(tenant__last_name__icontains=search_query) |
                Q(tenant__email__icontains=search_query) |
                Q(property_visit_payments__property__title__icontains=search_query) |
                Q(payment_method__icontains=search_query) |
                Q(notes__icontains=search_query)
            )
        
        # Status filter
        if status_filter:
            if status_filter == 'paid':
                booking_payments = booking_payments.filter(status='active')  # properties.Payment uses 'active' for paid
                visit_payments = visit_payments.filter(status='completed')  # payments.Payment uses 'completed'
            elif status_filter == 'pending':
                booking_payments = booking_payments.filter(status='active')  # Adjust as needed
                visit_payments = visit_payments.filter(status='pending')
        
        # Payment type filter
        if payment_type_filter == 'visit':
            booking_payments = Payment.objects.none()  # Only show visit payments
        elif payment_type_filter and payment_type_filter != 'visit':
            visit_payments = UnifiedPayment.objects.none()  # Only show booking payments
    
    # Group payments by booking to show booking-level payment information (after filtering)
    from collections import defaultdict
    from decimal import Decimal
    
    # Group booking payments by booking
    bookings_data = {}
    for payment in booking_payments:
        booking = payment.booking
        if booking.id not in bookings_data:
            # Calculate actual totals from all payments for this booking
            from django.db.models import Sum
            all_payments = Payment.objects.filter(booking=booking).exclude(payment_type='refund')
            all_refunds = Payment.objects.filter(booking=booking, payment_type='refund')
            total_paid = all_payments.aggregate(total=Sum('amount'))['total'] or Decimal('0')
            total_refunded = all_refunds.aggregate(total=Sum('amount'))['total'] or Decimal('0')
            net_paid = max(Decimal('0'), total_paid - total_refunded)
            
            # Get total amount (prefer calculated_total_amount)
            total_required = booking.calculated_total_amount if booking.calculated_total_amount and booking.calculated_total_amount > 0 else (booking.total_amount or Decimal('0'))
            remaining = max(Decimal('0'), total_required - net_paid)
            
            # Update booking's paid_amount if inconsistent
            if abs(booking.paid_amount - net_paid) > Decimal('0.01'):
                booking.paid_amount = net_paid
                booking.update_payment_status()
                booking.save()
            
            bookings_data[booking.id] = {
                'booking': booking,
                'total_required': total_required,
                'paid_so_far': net_paid,
                'remaining': remaining,
                'payment_count': 0,
                'last_payment_date': None,
                'last_payment_method': None,
            }
        
        bookings_data[booking.id]['payment_count'] += 1
        if not bookings_data[booking.id]['last_payment_date'] or payment.payment_date > bookings_data[booking.id]['last_payment_date']:
            bookings_data[booking.id]['last_payment_date'] = payment.payment_date
            bookings_data[booking.id]['last_payment_method'] = payment.payment_method
    
    # Convert to list for display
    bookings_list = list(bookings_data.values())
    
    # Add visit payments as separate entries (they don't have bookings)
    visit_payments_list = []
    for payment in visit_payments:
        visit_payment = payment.property_visit_payments.first() if payment.property_visit_payments.exists() else None
        if visit_payment:
            visit_payments_list.append({
                'booking': None,
                'visit_payment': payment,
                'visit_payment_obj': visit_payment,
                'total_required': visit_payment.amount,
                'paid_so_far': payment.amount if payment.status == 'completed' else Decimal('0'),
                'remaining': visit_payment.amount - (payment.amount if payment.status == 'completed' else Decimal('0')),
                'payment_count': 1,
                'last_payment_date': visit_payment.paid_at if visit_payment.paid_at else None,
                'last_payment_method': payment.payment_method,
            })
    
    # Apply status filter at booking level
    if status_filter:
        if status_filter == 'paid':
            bookings_list = [item for item in bookings_list if item['booking'].payment_status == 'paid']
            visit_payments_list = [item for item in visit_payments_list if item.get('visit_payment_obj') and item['visit_payment_obj'].status == 'completed']
        elif status_filter == 'partial':
            bookings_list = [item for item in bookings_list if item['booking'].payment_status == 'partial']
        elif status_filter == 'pending':
            bookings_list = [item for item in bookings_list if item['booking'].payment_status == 'pending']
            visit_payments_list = [item for item in visit_payments_list if item.get('visit_payment_obj') and item['visit_payment_obj'].status == 'pending']
    
    # Combine and sort
    all_payments_data = bookings_list + visit_payments_list
    all_payments_data.sort(key=lambda x: x['last_payment_date'] or (x['booking'].created_at if x['booking'] else x['visit_payment'].created_at), reverse=True)
    
    # Pagination
    from django.core.paginator import Paginator
    page_size = request.GET.get('page_size', '5')  # Default to 5
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50]:
            page_size = 5
    except (ValueError, TypeError):
        page_size = 5
    
    paginator = Paginator(all_payments_data, page_size)
    page_number = request.GET.get('page')
    payments_page = paginator.get_page(page_number)
    
    # Calculate payment statistics
    total_payments = len(all_payments_data)
    total_amount = sum(item['total_required'] for item in all_payments_data)
    paid_bookings = sum(1 for item in all_payments_data if (
        item['booking'] and item['booking'].payment_status == 'paid' if item['booking'] else
        (item.get('visit_payment') and item['visit_payment'].status == 'completed')
    ))
    
    # Get tenants (customers with house bookings) for the payment form
    tenants = Customer.objects.filter(
        customer_bookings__property_obj__property_type__name__iexact='house'
    ).distinct().select_related().prefetch_related('customer_bookings__property_obj')
    
    # Get house bookings for the payment form - use the exact same logic as house_bookings view
    # This ensures consistency between house bookings page and house payments form
    house_bookings = Booking.objects.filter(
        property_obj__property_type__name__iexact='house'
    ).select_related('customer', 'property_obj').order_by('-created_at')
    
    # Get payment choices from the model
    payment_method_choices = Payment.PAYMENT_METHOD_CHOICES
    payment_type_choices = Payment.PAYMENT_TYPE_CHOICES
    payment_status_choices = Payment.PAYMENT_STATUS_CHOICES
    
    context = {
        'house_properties': house_properties,
        'selected_property': selected_property,
        'payments': payments_page,  # Paginated payment data (booking-level)
        'payments_all': all_payments_data,  # For stats
        'total_payments': total_payments,
        'total_amount': total_amount,
        'paid_bookings': paid_bookings,
        'is_single_property_mode': bool(selected_property_id),
        'search_query': search_query,
        'status_filter': status_filter,
        'payment_type_filter': payment_type_filter,
        'tenants': tenants,  # For tenant dropdown
        'house_bookings': house_bookings,  # For booking dropdown
        'payment_method_choices': payment_method_choices,  # For payment method dropdown
        'payment_type_choices': payment_type_choices,  # For payment type dropdown
        'payment_status_choices': payment_status_choices,  # For status dropdown
    }
    return render(request, 'properties/house_payments.html', context)


@login_required
def test_bookings(request):
    """Test view to check booking queries"""
    from .models import Booking
    
    # Test the exact same query as house_bookings
    bookings_query = Booking.objects.filter(property_obj__property_type__name__iexact='house').select_related('customer', 'property_obj').order_by('-created_at')
    
    # Get all bookings for comparison
    all_bookings = Booking.objects.all().select_related('customer', 'property_obj')
    
    context = {
        'house_bookings': bookings_query,
        'all_bookings': all_bookings,
        'house_count': bookings_query.count(),
        'all_count': all_bookings.count(),
    }
    return render(request, 'properties/test_bookings.html', context)

@login_required
def house_reports(request):
    """House reports and analytics"""
    # Get selected property from session or request
    selected_property_id = request.session.get('selected_house_property_id') or request.GET.get('property_id')
    
    # Get house properties
    house_properties = Property.objects.filter(property_type__name__iexact='house')
    
    # Calculate statistics
    total_houses = house_properties.count()
    
    # Get active bookings (tenants)
    active_bookings = Booking.objects.filter(
        property_obj__property_type__name__iexact='house',
        booking_status__in=['confirmed', 'checked_in']
    )
    active_tenants = active_bookings.count()
    
    # Calculate occupancy rate
    occupancy_rate = 0
    if total_houses > 0:
        occupancy_rate = (active_tenants / total_houses) * 100
    
    # Calculate monthly revenue
    from datetime import datetime, timedelta
    current_month = datetime.now().replace(day=1)
    monthly_revenue = Payment.objects.filter(
        booking__property_obj__property_type__name__iexact='house',
        payment_date__gte=current_month,
        status='completed'
    ).aggregate(total=models.Sum('amount'))['total'] or 0
    
    # Calculate average rent amount
    avg_rent_amount = house_properties.aggregate(avg=models.Avg('rent_amount'))['avg'] or 0
    
    # Get maintenance requests count (placeholder - you can implement this based on your maintenance system)
    maintenance_requests = 0  # Placeholder
    
    # Calculate completion rate (placeholder)
    completion_rate = 85.0  # Placeholder
    
    # Calculate payment collection rate
    total_invoices = Payment.objects.filter(
        booking__property_obj__property_type__name__iexact='house'
    ).count()
    paid_invoices = Payment.objects.filter(
        booking__property_obj__property_type__name__iexact='house',
        status='completed'
    ).count()
    payment_collection_rate = 0
    if total_invoices > 0:
        payment_collection_rate = (paid_invoices / total_invoices) * 100
    
    # Create recent activities (placeholder data)
    recent_activities = [
        {
            'type': 'payment',
            'title': 'Payment Received',
            'date': datetime.now() - timedelta(hours=2),
            'description': 'Monthly rent payment received'
        },
        {
            'type': 'tenant',
            'title': 'New Tenant',
            'date': datetime.now() - timedelta(days=1),
            'description': 'New tenant moved in'
        },
        {
            'type': 'maintenance',
            'title': 'Maintenance Request',
            'date': datetime.now() - timedelta(days=2),
            'description': 'Plumbing issue reported'
        },
        {
            'type': 'property',
            'title': 'Lease Renewal',
            'date': datetime.now() - timedelta(days=3),
            'description': 'Lease renewed for another year'
        }
    ]
    
    context = {
        'house_properties': house_properties,
        'selected_property_id': selected_property_id,
        'total_houses': total_houses,
        'active_tenants': active_tenants,
        'occupancy_rate': occupancy_rate,
        'monthly_revenue': monthly_revenue,
        'avg_rent_amount': avg_rent_amount,
        'maintenance_requests': maintenance_requests,
        'completion_rate': completion_rate,
        'payment_collection_rate': payment_collection_rate,
        'recent_activities': recent_activities,
        'is_single_property_mode': bool(selected_property_id),
    }
    return render(request, 'properties/house_reports.html', context)


@login_required
def house_reports_export(request):
    """Export house reports in various formats"""
    from django.http import HttpResponse, JsonResponse
    from django.template.loader import render_to_string
    import csv
    import json
    from datetime import datetime, timedelta
    from io import StringIO
    
    # Get parameters
    report_type = request.GET.get('type', 'monthly')
    date_range = request.GET.get('range', 'this_month')
    properties = request.GET.get('properties', 'all')
    format_type = request.GET.get('format', 'excel')
    
    # Calculate date range
    now = datetime.now()
    if date_range == 'this_month':
        start_date = now.replace(day=1)
        end_date = now
    elif date_range == 'last_month':
        last_month = now.replace(day=1) - timedelta(days=1)
        start_date = last_month.replace(day=1)
        end_date = last_month
    elif date_range == 'this_quarter':
        quarter_start = now.month - (now.month - 1) % 3
        start_date = now.replace(month=quarter_start, day=1)
        end_date = now
    elif date_range == 'last_quarter':
        quarter_start = now.month - (now.month - 1) % 3
        last_quarter_end = now.replace(month=quarter_start, day=1) - timedelta(days=1)
        quarter_start = last_quarter_end.month - (last_quarter_end.month - 1) % 3
        start_date = last_quarter_end.replace(month=quarter_start, day=1)
        end_date = last_quarter_end
    elif date_range == 'this_year':
        start_date = now.replace(month=1, day=1)
        end_date = now
    elif date_range == 'last_year':
        start_date = now.replace(year=now.year-1, month=1, day=1)
        end_date = now.replace(year=now.year-1, month=12, day=31)
    elif date_range.startswith('custom_'):
        # Handle custom date range: custom_2024-01-01_2024-12-31
        try:
            date_parts = date_range.split('_')
            start_date = datetime.strptime(date_parts[1], '%Y-%m-%d')
            end_date = datetime.strptime(date_parts[2], '%Y-%m-%d')
        except (ValueError, IndexError):
            start_date = now.replace(day=1)
            end_date = now
    else:
        start_date = now.replace(day=1)
        end_date = now
    
    # Get house properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        house_properties = Property.objects.filter(property_type__name__iexact='house', owner=request.user)
    else:
        house_properties = Property.objects.filter(property_type__name__iexact='house')
    
    if properties == 'available':
        # Filter for available properties (no active bookings)
        house_properties = house_properties.exclude(
            property_bookings__booking_status__in=['confirmed', 'checked_in']
        ).distinct()
    elif properties == 'occupied':
        # Filter for occupied properties (with active bookings)
        house_properties = house_properties.filter(
            property_bookings__booking_status__in=['confirmed', 'checked_in']
        ).distinct()
    elif properties != 'all':
        try:
            property_ids = [int(p) for p in properties.split(',')]
            house_properties = house_properties.filter(id__in=property_ids)
        except (ValueError, TypeError):
            pass
    
    # Generate report data based on type
    report_data = {
        'report_type': report_type,
        'date_range': date_range,
        'start_date': start_date,
        'end_date': end_date,
        'properties': house_properties,
        'generated_at': now,
    }
    
    if report_type == 'monthly':
        # Monthly financial report
        payments = Payment.objects.filter(
            booking__property_obj__in=house_properties,
            payment_date__gte=start_date,
            payment_date__lte=end_date,
            status='completed'
        )
        report_data.update({
            'total_revenue': sum(p.amount for p in payments),
            'total_payments': payments.count(),
            'payments': payments,
        })
        
    elif report_type == 'quarterly':
        # Quarterly performance report
        bookings = Booking.objects.filter(
            property_obj__in=house_properties,
            created_at__gte=start_date,
            created_at__lte=end_date
        )
        report_data.update({
            'total_bookings': bookings.count(),
            'bookings': bookings,
            'occupancy_rate': (bookings.filter(booking_status__in=['confirmed', 'checked_in']).count() / max(house_properties.count(), 1)) * 100,
        })
        
    elif report_type == 'tenant':
        # Tenant report
        tenants = Customer.objects.filter(
            customer_bookings__property_obj__in=house_properties
        ).distinct()
        report_data.update({
            'total_tenants': tenants.count(),
            'tenants': tenants,
        })
        
    elif report_type == 'maintenance':
        # Maintenance report (placeholder data)
        report_data.update({
            'maintenance_requests': 0,  # Placeholder
            'completion_rate': 85.0,   # Placeholder
        })
        
    elif report_type == 'financial':
        # Comprehensive financial report
        payments = Payment.objects.filter(
            booking__property_obj__in=house_properties,
            payment_date__gte=start_date,
            payment_date__lte=end_date
        )
        report_data.update({
            'total_revenue': sum(p.amount for p in payments.filter(status='completed')),
            'pending_amount': sum(p.amount for p in payments.filter(status='pending')),
            'overdue_amount': sum(p.amount for p in payments.filter(status='overdue')),
            'payments': payments,
        })
    
    # Always generate Excel format (format_type parameter is ignored - Excel is required)
    # Generate Excel response using openpyxl
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_type.title()} Report"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write report header
        ws['A1'] = 'House Property Management Report'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f'Report Type: {report_type.title()}'
        ws['B2'] = f'Date Range: {date_range}'
        ws['A3'] = f'Generated At: {now.strftime("%Y-%m-%d %H:%M:%S")}'
        ws['B3'] = f'Period: {start_date.strftime("%Y-%m-%d")} to {end_date.strftime("%Y-%m-%d")}'
        
        # Add data based on report type
        row = 5
        if report_type == 'monthly':
            # Headers
            headers = ['Property', 'Address', 'Revenue (Tsh)', 'Payments Count', 'Average Amount']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Data
            row += 1
            total_revenue = 0
            total_payment_count = 0
            
            for prop in house_properties:
                prop_payments = payments.filter(booking__property_obj=prop)
                prop_revenue = sum(p.amount for p in prop_payments)
                payment_count = prop_payments.count()
                avg_amount = prop_revenue / max(payment_count, 1)
                
                ws.cell(row=row, column=1, value=prop.title)
                ws.cell(row=row, column=2, value=prop.address)
                ws.cell(row=row, column=3, value=prop_revenue)
                ws.cell(row=row, column=4, value=payment_count)
                ws.cell(row=row, column=5, value=avg_amount)
                
                # Accumulate totals
                total_revenue += prop_revenue
                total_payment_count += payment_count
                
                row += 1
            
            # Add total row for monthly report
            total_font = Font(bold=True, size=11)
            total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            avg_total = total_revenue / max(total_payment_count, 1)
            
            ws.cell(row=row, column=1, value="TOTAL")
            ws.cell(row=row, column=2, value="")
            ws.cell(row=row, column=3, value=total_revenue)
            ws.cell(row=row, column=4, value=total_payment_count)
            ws.cell(row=row, column=5, value=avg_total)
            
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.font = total_font
                cell.fill = total_fill
                
        elif report_type == 'tenant':
            # Headers
            headers = ['Tenant Name', 'Email', 'Phone', 'Property', 'Move-in Date', 'Status']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Data
            row += 1
            for tenant in tenants:
                latest_booking = tenant.customer_bookings.filter(property_obj__in=house_properties).first()
                ws.cell(row=row, column=1, value=tenant.full_name)
                ws.cell(row=row, column=2, value=tenant.email)
                ws.cell(row=row, column=3, value=tenant.phone)
                ws.cell(row=row, column=4, value=latest_booking.property_obj.title if latest_booking else 'N/A')
                ws.cell(row=row, column=5, value=latest_booking.check_in_date.strftime('%Y-%m-%d') if latest_booking else 'N/A')
                ws.cell(row=row, column=6, value=latest_booking.booking_status if latest_booking else 'N/A')
                row += 1
                
        elif report_type == 'quarterly':
            # Headers
            headers = ['Property', 'Bookings Count', 'Occupancy Rate (%)', 'Total Revenue (Tsh)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Data
            row += 1
            total_bookings = 0
            total_revenue = 0
            
            for prop in house_properties:
                prop_bookings = bookings.filter(property_obj=prop)
                occupancy_rate = (prop_bookings.filter(booking_status__in=['confirmed', 'checked_in']).count() / max(house_properties.count(), 1)) * 100
                prop_revenue = sum(p.amount for p in payments.filter(booking__property_obj=prop))
                
                ws.cell(row=row, column=1, value=prop.title)
                ws.cell(row=row, column=2, value=prop_bookings.count())
                ws.cell(row=row, column=3, value=occupancy_rate)
                ws.cell(row=row, column=4, value=prop_revenue)
                
                # Accumulate totals
                total_bookings += prop_bookings.count()
                total_revenue += prop_revenue
                
                row += 1
            
            # Add total row for quarterly report
            total_font = Font(bold=True, size=11)
            total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
            ws.cell(row=row, column=1, value="TOTAL")
            ws.cell(row=row, column=2, value=total_bookings)
            ws.cell(row=row, column=3, value="")
            ws.cell(row=row, column=4, value=total_revenue)
            
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.font = total_font
                cell.fill = total_fill
                
        else:
            # Generic report
            ws.cell(row=row, column=1, value='Report Type')
            ws.cell(row=row, column=2, value=report_type.title())
            row += 1
            ws.cell(row=row, column=1, value='Date Range')
            ws.cell(row=row, column=2, value=date_range)
            row += 1
            ws.cell(row=row, column=1, value='Properties Count')
            ws.cell(row=row, column=2, value=house_properties.count())
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="house_{report_type}_report_{date_range}.xlsx"'
        
    except ImportError:
        # Excel format is required - raise error if openpyxl is not available
        from django.http import JsonResponse
        return JsonResponse({
            'error': 'Excel generation requires openpyxl library. Please install it: pip install openpyxl'
        }, status=500)
    
    return response


# Property Selection Views

@login_required
def hotel_select_property(request):
    """Select a hotel property for management"""
    if request.method == 'POST':
        property_id = request.POST.get('property_id')
        if property_id and property_id != 'all':
            # Validate property exists before saving
            try:
                property_id_int = int(property_id)
                # Verify property exists and is a hotel
                Property.objects.get(id=property_id_int, property_type__name__iexact='hotel')
                set_property_selection(request, property_id, 'hotel')
                messages.success(request, 'Hotel property selected successfully!')
                # Redirect with property_id in URL query parameter to ensure it's immediately available
                from django.urls import reverse
                return redirect(f"{reverse('properties:hotel_dashboard')}?property_id={property_id_int}")
            except (ValueError, Property.DoesNotExist):
                messages.error(request, 'Selected hotel property not found.')
                return redirect('properties:hotel_dashboard')
        elif property_id == 'all':
            # Clear selection for "all"
            clear_property_selection(request, 'hotel')
            messages.info(request, 'Showing all hotels.')
        return redirect('properties:hotel_dashboard')
    
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Get all hotel properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        hotel_properties = Property.objects.filter(property_type__name__iexact='hotel', owner=request.user)
    else:
        hotel_properties = Property.objects.filter(property_type__name__iexact='hotel')
    
    # Get total count before filtering
    total_count = hotel_properties.count()
    
    # Apply search filter
    search_query = request.GET.get('search', '').strip()
    if search_query:
        hotel_properties = hotel_properties.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(region__name__icontains=search_query)
        )
    
    # Apply status filter
    status_filter = request.GET.get('status', '')
    if status_filter:
        hotel_properties = hotel_properties.filter(status=status_filter)
    
    # Annotate with actual room count from Room model
    from properties.models import Room
    from django.db.models import Count
    hotel_properties = hotel_properties.select_related('property_type', 'region', 'owner').prefetch_related('images').annotate(
        actual_room_count=Count('property_rooms', distinct=True)
    ).order_by('title')
    
    # Pagination
    page_size = request.GET.get('page_size', '5')
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50]:
            page_size = 5
    except (ValueError, TypeError):
        page_size = 5
    
    paginator = Paginator(hotel_properties, page_size)
    page_number = request.GET.get('page', 1)
    
    try:
        properties_page = paginator.get_page(page_number)
    except:
        properties_page = paginator.get_page(1)
    
    context = {
        'properties': properties_page,
        'property_type': 'hotel',
        'management_type': 'Hotel Management',
        'search_query': search_query,
        'status_filter': status_filter,
        'total_count': total_count,
    }
    
    # Check if this is an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'properties/partials/property_selection_table.html', context)
    
    return render(request, 'properties/property_selection.html', context)


@login_required
def hotel_clear_selection(request):
    """Clear hotel property selection"""
    # Clear from session using utility function (now uses correct session key format)
    clear_property_selection(request, 'hotel')
    messages.info(request, 'Hotel property selection cleared. Please select a hotel to continue.')
    # Redirect to selection page (dashboard will redirect here anyway if no hotel selected)
    return redirect('properties:hotel_select_property')


@login_required
def lodge_select_property(request):
    """Select a lodge property for management"""
    if request.method == 'POST':
        property_id = request.POST.get('property_id')
        if property_id and property_id != 'all':
            # Validate property exists before saving
            try:
                property_id_int = int(property_id)
                # Verify property exists and is a lodge
                Property.objects.get(id=property_id_int, property_type__name__iexact='lodge')
                set_property_selection(request, property_id, 'lodge')
                messages.success(request, 'Lodge property selected successfully!')
                # Redirect with property_id in URL query parameter to ensure it's immediately available
                from django.urls import reverse
                return redirect(f"{reverse('properties:lodge_dashboard')}?property_id={property_id_int}")
            except (ValueError, Property.DoesNotExist):
                messages.error(request, 'Selected lodge property not found.')
                return redirect('properties:lodge_dashboard')
        elif property_id == 'all':
            # Clear selection for "all"
            clear_property_selection(request, 'lodge')
            messages.info(request, 'Showing all lodges.')
        return redirect('properties:lodge_dashboard')
    
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Get all lodge properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        lodge_properties = Property.objects.filter(property_type__name__iexact='lodge', owner=request.user)
    else:
        lodge_properties = Property.objects.filter(property_type__name__iexact='lodge')
    
    # Get total count before filtering
    total_count = lodge_properties.count()
    
    # Apply search filter
    search_query = request.GET.get('search', '').strip()
    if search_query:
        lodge_properties = lodge_properties.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(region__name__icontains=search_query)
        )
    
    # Apply status filter
    status_filter = request.GET.get('status', '')
    if status_filter:
        lodge_properties = lodge_properties.filter(status=status_filter)
    
    # Annotate with actual room count from Room model
    from properties.models import Room
    from django.db.models import Count
    lodge_properties = lodge_properties.select_related('property_type', 'region', 'owner').prefetch_related('images').annotate(
        actual_room_count=Count('property_rooms', distinct=True)
    ).order_by('title')
    
    # Pagination
    page_size = request.GET.get('page_size', '5')
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50]:
            page_size = 5
    except (ValueError, TypeError):
        page_size = 5
    
    paginator = Paginator(lodge_properties, page_size)
    page_number = request.GET.get('page', 1)
    
    try:
        properties_page = paginator.get_page(page_number)
    except:
        properties_page = paginator.get_page(1)
    
    # Check if this is an AJAX request (for table updates)
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'properties/partials/property_selection_table.html', {
            'properties': properties_page,
            'property_type': 'lodge',
            'total_count': total_count,
        })
    
    context = {
        'properties': properties_page,
        'property_type': 'lodge',
        'management_type': 'Lodge Management',
        'search_query': search_query,
        'status_filter': status_filter,
        'total_count': total_count,
    }
    return render(request, 'properties/property_selection.html', context)


@login_required
def lodge_clear_selection(request):
    """Clear lodge property selection"""
    # Clear from session using utility function (now uses correct session key format)
    clear_property_selection(request, 'lodge')
    messages.info(request, 'Lodge property selection cleared. Please select a lodge to continue.')
    # Redirect to selection page (dashboard will redirect here anyway if no lodge selected)
    return redirect('properties:lodge_select_property')


@login_required
def venue_select_property(request):
    """Select a venue property for management"""
    if request.method == 'POST':
        property_id = request.POST.get('property_id')
        if property_id:
            set_property_selection(request, property_id, 'venue')
            messages.success(request, 'Venue property selected successfully!')
        return redirect('properties:venue_dashboard')
    
    # Get all venue properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        venue_properties = Property.objects.filter(property_type__name__iexact='venue', owner=request.user)
    else:
        venue_properties = Property.objects.filter(property_type__name__iexact='venue')
    
    context = {
        'properties': venue_properties,
        'property_type': 'venue',
        'management_type': 'Venue Management',
    }
    return render(request, 'properties/property_selection.html', context)


@login_required
def venue_clear_selection(request):
    """Clear venue property selection"""
    clear_property_selection(request, 'venue')
    messages.info(request, 'Venue property selection cleared.')
    return redirect('properties:venue_dashboard')


@login_required
def house_select_property(request):
    """Select a house property for management"""
    if request.method == 'POST':
        property_id = request.POST.get('property_id')
        if property_id and property_id != 'all':
            # Validate property exists before saving
            try:
                property_id_int = int(property_id)
                # Verify property exists and is a house
                Property.objects.get(id=property_id_int, property_type__name__iexact='house')
                set_property_selection(request, property_id, 'house')
                messages.success(request, 'House property selected successfully!')
                # Redirect with property_id in URL query parameter to ensure it's immediately available
                from django.urls import reverse
                return redirect(f"{reverse('properties:house_dashboard')}?property_id={property_id_int}")
            except (ValueError, Property.DoesNotExist):
                messages.error(request, 'Selected house property not found.')
                return redirect('properties:house_dashboard')
        elif property_id == 'all':
            # Clear selection for "all"
            clear_property_selection(request, 'house')
            messages.info(request, 'Showing all houses.')
        return redirect('properties:house_dashboard')
    
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Get all house properties - MULTI-TENANCY: Filter by owner for data isolation
    if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
        house_properties = Property.objects.filter(property_type__name__iexact='house', owner=request.user)
    else:
        house_properties = Property.objects.filter(property_type__name__iexact='house')
    
    # Get total count before filtering
    total_count = house_properties.count()
    
    # Apply search filter
    search_query = request.GET.get('search', '').strip()
    if search_query:
        house_properties = house_properties.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(region__name__icontains=search_query)
        )
    
    # Apply status filter
    status_filter = request.GET.get('status', '')
    if status_filter:
        house_properties = house_properties.filter(status=status_filter)
    
    # Annotate with actual room count from Room model
    from properties.models import Room
    from django.db.models import Count
    house_properties = house_properties.select_related('property_type', 'region', 'owner').prefetch_related('images').annotate(
        actual_room_count=Count('property_rooms', distinct=True)
    ).order_by('title')
    
    # Pagination - default 5 per page (consistent with hotel and lodge)
    page_size = request.GET.get('page_size', '5')
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50]:
            page_size = 5
    except (ValueError, TypeError):
        page_size = 5
    
    paginator = Paginator(house_properties, page_size)
    page_number = request.GET.get('page', 1)
    
    try:
        properties_page = paginator.get_page(page_number)
    except:
        properties_page = paginator.get_page(1)
    
    context = {
        'properties': properties_page,
        'property_type': 'house',
        'management_type': 'House Management',
        'search_query': search_query,
        'status_filter': status_filter,
        'total_count': total_count,
    }
    
    # Check if this is an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'properties/partials/property_selection_table.html', context)
    
    return render(request, 'properties/property_selection.html', context)


@login_required
def house_clear_selection(request):
    """Clear house property selection"""
    clear_property_selection(request, 'house')
    messages.info(request, 'House property selection cleared.')
    return redirect('properties:house_dashboard')


# Booking Creation Views

@login_required
def create_lodge_booking(request):
    """Create a new lodge booking"""
    if request.method == 'POST':
        # Get selected property from session
        selected_property_id = request.session.get('selected_lodge_property_id')
        
        # Validate selected_property_id - handle 'all' case
        selected_property_id = validate_property_id(selected_property_id)
        
        if not selected_property_id:
            messages.error(request, 'Please select a lodge first.')
            return redirect('properties:lodge_select_property')
        
        try:
            selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='lodge')
        except Property.DoesNotExist:
            messages.error(request, 'Selected lodge not found.')
            return redirect('properties:lodge_select_property')
        
        # Check property availability
        from datetime import datetime
        try:
            check_in = datetime.strptime(request.POST.get('check_in_date'), '%Y-%m-%d').date()
            check_out = datetime.strptime(request.POST.get('check_out_date'), '%Y-%m-%d').date()
            
            if not selected_property.is_available_for_booking(check_in, check_out):
                messages.error(request, 'Property is not available for the selected dates. Please choose different dates.')
                return redirect('properties:create_lodge_booking')
        except (ValueError, TypeError):
            messages.error(request, 'Invalid date format.')
            return redirect('properties:create_lodge_booking')
        
        # Create customer
        customer, created = Customer.objects.get_or_create(
            email=request.POST.get('email'),
            defaults={
                'first_name': request.POST.get('first_name'),
                'last_name': request.POST.get('last_name'),
                'phone': request.POST.get('phone'),
            }
        )
        
        # Create booking - use parsed date objects (check_in, check_out) instead of strings
        booking = Booking.objects.create(
            property_obj=selected_property,
            customer=customer,
            booking_reference=f"LDG-{Booking.objects.count() + 1:06d}",
            check_in_date=check_in,  # Use parsed date object, not string
            check_out_date=check_out,  # Use parsed date object, not string
            number_of_guests=int(request.POST.get('number_of_guests', 1)),
            room_type=request.POST.get('room_type'),
            total_amount=float(request.POST.get('total_amount', 0)),
            created_by=request.user,
        )
        
        # Assign room - REQUIRED for lodge bookings
        room_number = request.POST.get('room_number')
        if not room_number:
            messages.error(request, 'Room selection is required for lodge bookings.')
            return redirect('properties:create_lodge_booking')
        
        try:
            room = Room.objects.get(property_obj=selected_property, room_number=room_number)
            booking.room_number = room_number
            booking.room_type = room.room_type
            booking.save()
            
            # Link room to booking
            room.current_booking = booking
            room.status = 'occupied'
            room.save()
        except Room.DoesNotExist:
            messages.error(request, f'Room {room_number} not found.')
            booking.delete()  # Delete the booking if room doesn't exist
            return redirect('properties:create_lodge_booking')
        
        messages.success(request, f'Booking {booking.booking_reference} created successfully!')
        return redirect('properties:lodge_bookings')
    
    # Get selected property
    selected_property_id = request.session.get('selected_lodge_property_id')
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    if not selected_property_id:
        messages.error(request, 'Please select a lodge first.')
        return redirect('properties:lodge_select_property')
    
    try:
        selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='lodge')
    except Property.DoesNotExist:
        messages.error(request, 'Selected lodge not found.')
        return redirect('properties:lodge_select_property')
    
    # Get available rooms
    available_rooms = Room.objects.filter(
        property_obj=selected_property,
        status='available'
    ).order_by('room_number')
    
    context = {
        'selected_property': selected_property,
        'available_rooms': available_rooms,
    }
    return render(request, 'properties/create_lodge_booking.html', context)


@login_required
def create_venue_booking(request):
    """Create a new venue booking"""
    if request.method == 'POST':
        # Get selected property from session
        selected_property_id = request.session.get('selected_venue_property_id')
        
        # Validate selected_property_id - handle 'all' case
        selected_property_id = validate_property_id(selected_property_id)
        
        if not selected_property_id:
            messages.error(request, 'Please select a venue first.')
            return redirect('properties:venue_select_property')
        
        try:
            selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='venue')
        except Property.DoesNotExist:
            messages.error(request, 'Selected venue not found.')
            return redirect('properties:venue_select_property')
        
        # Check property availability
        from datetime import datetime
        try:
            check_in = datetime.strptime(request.POST.get('check_in_date'), '%Y-%m-%d').date()
            check_out = datetime.strptime(request.POST.get('check_out_date'), '%Y-%m-%d').date()
            
            if not selected_property.is_available_for_booking(check_in, check_out):
                messages.error(request, 'Property is not available for the selected dates. Please choose different dates.')
                return redirect('properties:create_venue_booking')
        except (ValueError, TypeError):
            messages.error(request, 'Invalid date format.')
            return redirect('properties:create_venue_booking')
        
        # Create customer
        customer, created = Customer.objects.get_or_create(
            email=request.POST.get('email'),
            defaults={
                'first_name': request.POST.get('first_name'),
                'last_name': request.POST.get('last_name'),
                'phone': request.POST.get('phone'),
            }
        )
        
        # Create booking - use parsed date objects (check_in, check_out) instead of strings
        booking = Booking.objects.create(
            property_obj=selected_property,
            customer=customer,
            booking_reference=f"VEN-{Booking.objects.count() + 1:06d}",
            check_in_date=check_in,  # Use parsed date object, not string
            check_out_date=check_out,  # Use parsed date object, not string
            number_of_guests=int(request.POST.get('number_of_guests', 1)),
            total_amount=float(request.POST.get('total_amount', 0)),
            created_by=request.user,
        )
        
        messages.success(request, f'Venue booking {booking.booking_reference} created successfully!')
        return redirect('properties:venue_bookings')
    
    # Get selected property
    selected_property_id = request.session.get('selected_venue_property_id')
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    if not selected_property_id:
        messages.error(request, 'Please select a venue first.')
        return redirect('properties:venue_select_property')
    
    try:
        selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='venue')
    except Property.DoesNotExist:
        messages.error(request, 'Selected venue not found.')
        return redirect('properties:venue_select_property')
    
    context = {
        'selected_property': selected_property,
    }
    return render(request, 'properties/create_venue_booking.html', context)


@login_required
def create_house_booking(request):
    """Create a new house booking (rental)"""
    if request.method == 'POST':
        # Get selected property from session
        selected_property_id = request.session.get('selected_house_property_id')
        
        # Validate selected_property_id - handle 'all' case
        selected_property_id = validate_property_id(selected_property_id)
        
        if not selected_property_id:
            messages.error(request, 'Please select a house first.')
            return redirect('properties:house_select_property')
        
        try:
            selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='house')
        except Property.DoesNotExist:
            messages.error(request, 'Selected house not found.')
            return redirect('properties:house_select_property')
        
        # Check property availability
        from datetime import datetime
        try:
            check_in = datetime.strptime(request.POST.get('check_in_date'), '%Y-%m-%d').date()
            check_out = datetime.strptime(request.POST.get('check_out_date'), '%Y-%m-%d').date()
            
            if not selected_property.is_available_for_booking(check_in, check_out):
                messages.error(request, 'Property is not available for the selected dates. Please choose different dates.')
                return redirect('properties:create_house_booking')
        except (ValueError, TypeError):
            messages.error(request, 'Invalid date format.')
            return redirect('properties:create_house_booking')
        
        # Create customer (tenant)
        customer, created = Customer.objects.get_or_create(
            email=request.POST.get('email'),
            defaults={
                'first_name': request.POST.get('first_name'),
                'last_name': request.POST.get('last_name'),
                'phone': request.POST.get('phone'),
            }
        )
        
        # Create booking (rental) - calculate total based on months
        # Use parsed date objects (check_in, check_out) instead of strings
        booking = Booking.objects.create(
            property_obj=selected_property,
            customer=customer,
            booking_reference=f"HSE-{Booking.objects.count() + 1:06d}",
            check_in_date=check_in,  # Use parsed date object, not string
            check_out_date=check_out,  # Use parsed date object, not string
            number_of_guests=int(request.POST.get('number_of_guests', 1)),
            total_amount=0,  # Will be calculated automatically
            created_by=request.user,
        )
        
        # Calculate and update total amount based on months
        booking.calculate_and_update_total()
        
        messages.success(request, f'House rental {booking.booking_reference} created successfully!')
        return redirect('properties:house_bookings')
    
    # Get selected property
    selected_property_id = request.session.get('selected_house_property_id')
    
    # Validate selected_property_id - handle 'all' case
    selected_property_id = validate_property_id(selected_property_id)
    
    if not selected_property_id:
        messages.error(request, 'Please select a house first.')
        return redirect('properties:house_select_property')
    
    try:
        selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='house')
    except Property.DoesNotExist:
        messages.error(request, 'Selected house not found.')
        return redirect('properties:house_select_property')
    
    context = {
        'selected_property': selected_property,
    }
    return render(request, 'properties/create_house_booking.html', context)


# Room Management Views

@login_required
def add_hotel_room(request):
    """Add a new room to selected hotel"""
    if request.method == 'POST':
        # Get selected property from session
        selected_property_id = request.session.get('selected_hotel_property_id')
        
        if not selected_property_id:
            messages.error(request, 'Please select a hotel first.')
            return redirect('properties:hotel_select_property')
        
        try:
            selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='hotel')
        except Property.DoesNotExist:
            messages.error(request, 'Selected hotel not found.')
            return redirect('properties:hotel_select_property')
        
        # Validate base_rate is set and greater than 0
        # Each room MUST have its own price - no default/fallback pricing
        base_rate = float(request.POST.get('base_rate', 0))
        if not base_rate or base_rate <= 0:
            messages.error(request, 'Base Rate is required and must be greater than 0. Each room must have its own price set.')
            return redirect('properties:hotel_rooms')
        
        # Handle floor_number - can be empty string, need to convert properly
        floor_number_str = request.POST.get('floor_number', '').strip()
        floor_number = int(floor_number_str) if floor_number_str and floor_number_str.isdigit() else None
        
        # Create room
        room = Room.objects.create(
            property_obj=selected_property,
            room_number=request.POST.get('room_number'),
            room_type=request.POST.get('room_type'),
            floor_number=floor_number,
            capacity=int(request.POST.get('capacity', 1)),
            bed_type=request.POST.get('bed_type'),
            amenities=request.POST.get('amenities'),
            base_rate=base_rate,  # This is the ONLY source of pricing for this room
        )
        
        messages.success(request, f'Room {room.room_number} added successfully!')
        return redirect('properties:hotel_rooms')
    
    # Get selected property
    selected_property_id = request.session.get('selected_hotel_property_id')
    if not selected_property_id:
        messages.error(request, 'Please select a hotel first.')
        return redirect('properties:hotel_select_property')
    
    try:
        selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='hotel')
    except Property.DoesNotExist:
        messages.error(request, 'Selected hotel not found.')
        return redirect('properties:hotel_select_property')
    
    context = {
        'selected_property': selected_property,
    }
    return render(request, 'properties/add_hotel_room.html', context)


@login_required
def add_lodge_room(request):
    """Add a new room to selected lodge"""
    if request.method == 'POST':
        # Get selected property from session
        selected_property_id = request.session.get('selected_lodge_property_id')
        
        if not selected_property_id:
            messages.error(request, 'Please select a lodge first.')
            return redirect('properties:lodge_select_property')
        
        try:
            selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='lodge')
        except Property.DoesNotExist:
            messages.error(request, 'Selected lodge not found.')
            return redirect('properties:lodge_select_property')
        
        # Validate base_rate is set and greater than 0
        # Each room MUST have its own price - no default/fallback pricing
        base_rate = float(request.POST.get('base_rate', 0))
        if not base_rate or base_rate <= 0:
            messages.error(request, 'Base Rate is required and must be greater than 0. Each room must have its own price set.')
            return redirect('properties:lodge_rooms')
        
        # Handle floor_number - can be empty string, need to convert properly
        floor_number_str = request.POST.get('floor_number', '').strip()
        floor_number = int(floor_number_str) if floor_number_str and floor_number_str.isdigit() else None
        
        # Create room
        room = Room.objects.create(
            property_obj=selected_property,
            room_number=request.POST.get('room_number'),
            room_type=request.POST.get('room_type'),
            floor_number=floor_number,
            capacity=int(request.POST.get('capacity', 1)),
            bed_type=request.POST.get('bed_type'),
            amenities=request.POST.get('amenities'),
            base_rate=base_rate,  # This is the ONLY source of pricing for this room
        )
        
        messages.success(request, f'Room {room.room_number} added successfully!')
        return redirect('properties:lodge_rooms')
    
    # Get selected property
    selected_property_id = request.session.get('selected_lodge_property_id')
    if not selected_property_id:
        messages.error(request, 'Please select a lodge first.')
        return redirect('properties:lodge_select_property')
    
    try:
        selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact='lodge')
    except Property.DoesNotExist:
        messages.error(request, 'Selected lodge not found.')
        return redirect('properties:lodge_select_property')
    
    context = {
        'selected_property': selected_property,
    }
    return render(request, 'properties/add_lodge_room.html', context)


# Customer Management Views

@login_required
def create_customer(request):
    """Create a new customer"""
    if request.method == 'POST':
        # Create customer
        customer = Customer.objects.create(
            first_name=request.POST.get('first_name'),
            last_name=request.POST.get('last_name'),
            email=request.POST.get('email'),
            phone=request.POST.get('phone'),
            gender=request.POST.get('gender'),
            date_of_birth=request.POST.get('date_of_birth') or None,
            address=request.POST.get('address'),
            city=request.POST.get('city'),
            country=request.POST.get('country'),
            postal_code=request.POST.get('postal_code'),
            id_type=request.POST.get('id_type'),
            id_number=request.POST.get('id_number'),
            emergency_contact_name=request.POST.get('emergency_contact_name'),
            emergency_contact_phone=request.POST.get('emergency_contact_phone'),
            notes=request.POST.get('notes'),
        )
        
        messages.success(request, f'Customer {customer.full_name} created successfully!')
        
        # Redirect back to the referring management page
        referer = request.META.get('HTTP_REFERER', '/')
        return redirect(referer)
    
    context = {}
    return render(request, 'properties/create_customer.html', context)


@login_required
def customer_detail(request, pk):
    """View customer details and booking history"""
    customer = get_object_or_404(Customer, pk=pk)
    
    # Get all bookings for this customer
    bookings = Booking.objects.filter(customer=customer).select_related('property_obj').order_by('-created_at')
    
    # Get payment history
    payments = Payment.objects.filter(booking__customer=customer).select_related('booking').order_by('-payment_date')
    
    context = {
        'customer': customer,
        'bookings': bookings,
        'payments': payments,
    }
    return render(request, 'properties/customer_detail.html', context)


@login_required
def edit_customer(request, pk):
    """Edit customer information"""
    customer = get_object_or_404(Customer, pk=pk)
    
    if request.method == 'POST':
        errors = []
        
        # Get and validate form data
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        
        # Validate required fields
        if not first_name:
            errors.append('First name is required.')
        else:
            customer.first_name = first_name
            
        if not last_name:
            errors.append('Last name is required.')
        else:
            customer.last_name = last_name
        
        if not email:
            errors.append('Email is required.')
        else:
            # Check email uniqueness (excluding current customer)
            existing_customer = Customer.objects.filter(email__iexact=email).exclude(id=customer.id).first()
            if existing_customer:
                errors.append(f'Email {email} is already in use by another customer ({existing_customer.full_name}).')
            else:
                customer.email = email
        
        if not phone:
            errors.append('Phone number is required.')
        else:
            # Check phone uniqueness (excluding current customer)
            existing_customer = Customer.objects.filter(phone=phone).exclude(id=customer.id).first()
            if existing_customer:
                errors.append(f'Phone number {phone} is already in use by another customer ({existing_customer.full_name}).')
            else:
                customer.phone = phone
        
        # Update other fields
        customer.gender = request.POST.get('gender') or None
        date_of_birth = request.POST.get('date_of_birth')
        customer.date_of_birth = date_of_birth if date_of_birth else None
        customer.address = request.POST.get('address', '').strip() or None
        customer.city = request.POST.get('city', '').strip() or None
        customer.country = request.POST.get('country', '').strip() or None
        customer.postal_code = request.POST.get('postal_code', '').strip() or None
        customer.id_type = request.POST.get('id_type', '').strip() or None
        customer.id_number = request.POST.get('id_number', '').strip() or None
        customer.emergency_contact_name = request.POST.get('emergency_contact_name', '').strip() or None
        customer.emergency_contact_phone = request.POST.get('emergency_contact_phone', '').strip() or None
        customer.notes = request.POST.get('notes', '').strip() or None
        
        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                customer.save()
                messages.success(request, f'Customer {customer.full_name} updated successfully!')
                return redirect('properties:customer_detail', pk=customer.pk)
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error saving customer {pk}: {str(e)}")
                messages.error(request, f'Error saving customer: {str(e)}')
    
    context = {
        'customer': customer,
    }
    return render(request, 'properties/edit_customer.html', context)


# AJAX Modal Views for Customer Management
@login_required
def customer_profile_modal(request, pk):
    """AJAX view for customer profile modal"""
    customer = get_object_or_404(Customer, pk=pk)
    
    # Get property type filter from request
    property_type = request.GET.get('property_type', 'venue').lower()
    
    # Filter bookings by property type
    bookings = customer.customer_bookings.filter(
        property_obj__property_type__name__iexact=property_type
    )
    total_spent = sum(booking.total_amount for booking in bookings)
    last_booking = bookings.order_by('-check_out_date').first()
    
    context = {
        'customer': customer,
        'bookings': bookings,
        'total_spent': total_spent,
        'last_booking': last_booking,
        'property_type': property_type,
    }
    return render(request, 'properties/modals/customer_profile.html', context)


@login_required
def customer_edit_modal(request, pk):
    """AJAX view for customer edit modal"""
    customer = get_object_or_404(Customer, pk=pk)
    
    if request.method == 'POST':
        errors = []
        
        # Get and validate form data
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        
        # Validate required fields
        if not first_name:
            errors.append('First name is required.')
        else:
            customer.first_name = first_name
            
        if not last_name:
            errors.append('Last name is required.')
        else:
            customer.last_name = last_name
        
        if not email:
            errors.append('Email is required.')
        else:
            # Check email uniqueness (excluding current customer)
            existing_customer = Customer.objects.filter(email__iexact=email).exclude(id=customer.id).first()
            if existing_customer:
                errors.append(f'Email {email} is already in use by another customer ({existing_customer.full_name}).')
            else:
                customer.email = email
        
        if not phone:
            errors.append('Phone number is required.')
        else:
            # Check phone uniqueness (excluding current customer)
            existing_customer = Customer.objects.filter(phone=phone).exclude(id=customer.id).first()
            if existing_customer:
                errors.append(f'Phone number {phone} is already in use by another customer ({existing_customer.full_name}).')
            else:
                customer.phone = phone
        
        # Update other fields
        customer.gender = request.POST.get('gender') or None
        date_of_birth = request.POST.get('date_of_birth')
        customer.date_of_birth = date_of_birth if date_of_birth else None
        customer.address = request.POST.get('address', '').strip() or None
        customer.city = request.POST.get('city', '').strip() or None
        customer.country = request.POST.get('country', '').strip() or None
        customer.postal_code = request.POST.get('postal_code', '').strip() or None
        customer.id_type = request.POST.get('id_type', '').strip() or None
        customer.id_number = request.POST.get('id_number', '').strip() or None
        customer.emergency_contact_name = request.POST.get('emergency_contact_name', '').strip() or None
        customer.emergency_contact_phone = request.POST.get('emergency_contact_phone', '').strip() or None
        customer.notes = request.POST.get('notes', '').strip() or None
        
        if errors:
            return JsonResponse({'success': False, 'errors': errors, 'message': 'Validation failed. Please correct the errors.'}, status=400)
        
        try:
            customer.save()
            return JsonResponse({'success': True, 'message': 'Customer updated successfully!'})
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error saving customer {pk}: {str(e)}")
            return JsonResponse({'success': False, 'message': f'Error saving customer: {str(e)}'}, status=500)
    
    context = {
        'customer': customer,
    }
    return render(request, 'properties/modals/customer_edit.html', context)


@login_required
def customer_booking_history_modal(request, pk):
    """AJAX view for customer booking history modal"""
    customer = get_object_or_404(Customer, pk=pk)
    
    # Get property type filter from request
    property_type = request.GET.get('property_type', 'venue').lower()
    
    # Filter bookings by property type
    bookings = customer.customer_bookings.filter(
        property_obj__property_type__name__iexact=property_type
    ).select_related('property_obj').order_by('-created_at')
    
    context = {
        'customer': customer,
        'bookings': bookings,
        'property_type': property_type,
    }
    return render(request, 'properties/modals/customer_booking_history.html', context)


@login_required
def customer_new_booking_modal(request, pk):
    """AJAX view for customer new booking modal"""
    from datetime import datetime
    customer = get_object_or_404(Customer, pk=pk)
    
    # Check referer to determine property type
    referer = request.META.get('HTTP_REFERER', '')
    property_type = 'hotel'  # default
    
    if 'lodge/customers' in referer:
        property_type = 'lodge'
    elif 'venue/customers' in referer:
        property_type = 'venue'
    
    properties = Property.objects.filter(property_type__name__iexact=property_type)
    
    if request.method == 'POST':
        # Create new booking logic here
        property_id = request.POST.get('property')
        check_in_date = request.POST.get('check_in_date')
        check_out_date = request.POST.get('check_out_date')
        number_of_guests = request.POST.get('number_of_guests', 1)
        special_requests = request.POST.get('special_requests', '')
        
        try:
            property_obj = Property.objects.get(id=property_id)
            
            # Calculate duration and total amount
            check_in = datetime.strptime(check_in_date, '%Y-%m-%d').date()
            check_out = datetime.strptime(check_out_date, '%Y-%m-%d').date()
            duration_days = (check_out - check_in).days
            daily_rate = property_obj.rent_amount
            calculated_total = daily_rate * duration_days
            
            booking = Booking.objects.create(
                property_obj=property_obj,
                customer=customer,
                booking_reference=f"BK{customer.id}{property_obj.id}{timezone.now().strftime('%Y%m%d%H%M%S')}",
                check_in_date=check_in,  # Use parsed date object, not string
                check_out_date=check_out,  # Use parsed date object, not string
                number_of_guests=number_of_guests,
                total_amount=calculated_total,
                special_requests=special_requests,
                created_by=request.user,
            )
            return JsonResponse({'success': True, 'message': f'Booking created successfully! Total: Tsh {calculated_total:,.0f} for {duration_days} days'})
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error creating booking: {str(e)}'})
    
    context = {
        'customer': customer,
        'hotel_properties': properties,  # Keep backward compatibility
        'properties': properties,
        'property_type': property_type,
    }
    return render(request, 'properties/modals/customer_new_booking.html', context)


@login_required
def customer_vip_status_modal(request, pk):
    """AJAX view for customer VIP status modal"""
    customer = get_object_or_404(Customer, pk=pk)
    
    if request.method == 'POST':
        # Toggle VIP status (assuming we add a VIP field to Customer model)
        # For now, we'll use a simple approach with notes field
        if 'VIP' in customer.notes:
            customer.notes = customer.notes.replace('VIP', '').strip()
            message = 'VIP status removed'
        else:
            customer.notes = f"{customer.notes} VIP".strip()
            message = 'VIP status added'
        customer.save()
        
        return JsonResponse({'success': True, 'message': message})
    
    context = {
        'customer': customer,
        'is_vip': 'VIP' in customer.notes if customer.notes else False,
    }
    return render(request, 'properties/modals/customer_vip_status.html', context)


# Payment Management Views

@login_required
def create_payment(request, booking_id=None, payment_id=None):
    """Create/record a payment for a booking or visit payment - uniform payment form"""
    from decimal import Decimal
    from payments.models import Payment as UnifiedPayment, PaymentProvider, PaymentTransaction
    from payments.gateway_service import PaymentGatewayService
    from .models import PropertyVisitPayment
    
    # Determine if this is a booking payment or visit payment
    is_visit_payment = payment_id is not None
    booking = None
    visit_payment = None
    property_obj = None
    unified_payment = None
    
    if is_visit_payment:
        # Handle visit payment
        unified_payment = get_object_or_404(UnifiedPayment, pk=payment_id)
        visit_payment = PropertyVisitPayment.objects.filter(payment=unified_payment).first()
        if not visit_payment:
            messages.error(request, 'This payment is not associated with a property visit.')
            return redirect('properties:house_payments')
        property_obj = visit_payment.property
    else:
        # Handle booking payment - booking_id should be provided
        if not booking_id:
            messages.error(request, 'Booking ID is required.')
            return redirect('properties:dashboard')
        booking = get_object_or_404(Booking, pk=booking_id)
    
    if request.method == 'POST':
        payment_type = request.POST.get('payment_type', '')
        # For partial payments, use partial_amount if provided, otherwise use amount
        if payment_type == 'partial' and request.POST.get('partial_amount'):
            partial_amount_str = request.POST.get('partial_amount', '0').strip()
            if partial_amount_str:
                payment_amount = Decimal(partial_amount_str)
            else:
                payment_amount = Decimal(request.POST.get('amount', 0))
        else:
            payment_amount = Decimal(request.POST.get('amount', 0))
        payment_method = request.POST.get('payment_method', 'cash')
        
        # Debug: Log payment method
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Payment method received: {payment_method}")
        logger.info(f"Payment method type: {type(payment_method)}")
        logger.info(f"POST data: {dict(request.POST)}")
        
        # Validate payment amount
        if payment_amount <= 0:
            messages.error(request, 'Payment amount must be greater than zero.')
            if is_visit_payment:
                return redirect('properties:create_visit_payment', payment_id=unified_payment.id)
            else:
                return redirect('properties:create_payment', booking_id=booking.pk)
        
        # Handle visit payment recording
        if is_visit_payment:
            from django.utils import timezone
            
            # Validate amount doesn't exceed visit payment amount
            if payment_amount > visit_payment.amount:
                messages.error(request, f'Payment amount (Tsh{payment_amount:,.0f}) exceeds visit payment amount (Tsh{visit_payment.amount:,.0f}).')
                return redirect('properties:create_visit_payment', payment_id=unified_payment.id)
            
            # Parse paid date
            paid_date = request.POST.get('paid_date')
            if paid_date:
                from django.utils.dateparse import parse_date
                paid_date_obj = parse_date(paid_date) or timezone.now().date()
            else:
                paid_date_obj = timezone.now().date()
            
            reference_number = request.POST.get('reference_number', '').strip()
            transaction_id = request.POST.get('transaction_id', '').strip()
            notes = request.POST.get('notes', '').strip()
            
            # Auto-generate reference number if not provided
            if not reference_number:
                from datetime import datetime
                timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                random_suffix = str(property_obj.id).zfill(3) + str(unified_payment.id).zfill(3)
                reference_number = f'VISIT-{property_obj.id}-{timestamp}{random_suffix[-6:]}'
            
            # Visit payments always go through AZAMPAY (even cash)
            # Get or create AZAM Pay provider
            provider, _ = PaymentProvider.objects.get_or_create(
                name='AZAM Pay',
                defaults={'description': 'AZAM Pay Payment Gateway'}
            )
            
            # Get mobile money provider if mobile_money payment
            mobile_money_provider = None
            if payment_method == 'mobile_money':
                mobile_money_provider = request.POST.get('mobile_money_provider', '').strip()
                if not mobile_money_provider:
                    messages.error(request, 'Mobile Money Provider is required for mobile money payments. Please select your provider (AIRTEL, TIGO, MPESA, or HALOPESA).')
                    return redirect('properties:create_visit_payment', payment_id=unified_payment.id)
            
            # Get customer user - try to find user by email, otherwise use request.user
            from django.contrib.auth import get_user_model
            User = get_user_model()
            try:
                customer_user = User.objects.get(email=visit_payment.user.email)
            except User.DoesNotExist:
                customer_user = visit_payment.user
            
            # Update unified payment with provider info
            unified_payment.amount = payment_amount
            unified_payment.payment_method = payment_method
            unified_payment.provider = provider
            unified_payment.mobile_money_provider = mobile_money_provider
            unified_payment.tenant = customer_user
            unified_payment.status = 'pending'  # Visit payments go through AZAMPAY
            unified_payment.notes = notes or unified_payment.notes
            unified_payment.recorded_by = request.user
            unified_payment.save()
            
            # Initiate AZAMpay payment for visit payments
            from django.conf import settings
            callback_url = getattr(settings, 'AZAM_PAY_WEBHOOK_URL', None)
            if not callback_url:
                base_domain = getattr(settings, 'BASE_URL', 'https://portal.maishaapp.co.tz')
                callback_url = f"{base_domain}/api/v1/payments/webhook/azam-pay/"
            
            gateway_result = PaymentGatewayService.initiate_payment(
                payment=unified_payment,
                provider_name='azam pay',
                callback_url=callback_url,
                payment_method='mobile_money' if payment_method == 'mobile_money' else 'mobile_money'
            )
            
            if gateway_result.get('success'):
                # Create payment transaction with the actual payload sent to AZAM Pay
                # This includes accountNumber (phone number) used for the payment
                request_payload = gateway_result.get('request_payload', {})
                if not request_payload:
                    # Fallback: create minimal payload if not provided
                    request_payload = {'visit_payment_id': visit_payment.id}
                
                PaymentTransaction.objects.create(
                    payment=unified_payment,
                    provider=provider,
                    gateway_transaction_id=gateway_result.get('transaction_id'),
                    azam_reference=gateway_result.get('reference'),
                    status='initiated',
                    request_payload=request_payload  # Store actual payload sent to AZAM Pay
                )
                
                transaction_id = gateway_result.get('transaction_id')
                payment_message = gateway_result.get('message', 'Payment initiated successfully')
                
                payment_link = gateway_result.get('payment_link')
                
                if payment_link:
                    # Bank/Online payment - redirect to payment page
                    request.session['payment_link'] = payment_link
                    request.session['payment_id'] = unified_payment.id
                    messages.info(request, 'Payment initiated. Please complete payment on AZAM Pay.')
                    return redirect(payment_link)
                else:
                    # MNO payment - no redirect, payment initiated on mobile money network
                    messages.success(request, f'Payment initiated successfully! Transaction ID: {transaction_id}. The customer will receive a payment prompt on their phone.')
                    return redirect('properties:house_payments')
            else:
                # Payment initiation failed
                error_msg = gateway_result.get('error', 'Failed to initiate payment')
                messages.error(request, f'Failed to initiate payment: {error_msg}')
                return redirect('properties:create_visit_payment', payment_id=unified_payment.id)
        
        # Continue with booking payment logic below
        
        # Recalculate everything from actual payments to ensure accuracy
        from django.db.models import Sum
        from decimal import Decimal
        
        # Get actual total paid from all Payment records (exclude refunds)
        actual_paid = Payment.objects.filter(
            booking=booking
        ).exclude(
            payment_type='refund'
        ).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0')
        
        # Get total refunded
        total_refunded = Payment.objects.filter(
            booking=booking,
            payment_type='refund'
        ).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0')
        
        # Net paid amount = payments - refunds
        net_paid_amount = actual_paid - total_refunded
        if net_paid_amount < 0:
            net_paid_amount = Decimal('0')
        
        # Get the correct total amount
        calculated_total = booking.calculated_total_amount or 0
        stored_total = booking.total_amount or 0
        
        if calculated_total and calculated_total > 0:
            total_amount = calculated_total
        else:
            total_amount = stored_total
        
        # Update booking's paid_amount to match actual
        if booking.paid_amount != net_paid_amount:
            booking.paid_amount = net_paid_amount
            booking.save()
        
        # Calculate remaining
        remaining_amount = max(Decimal('0'), Decimal(str(total_amount)) - net_paid_amount)
        
        # Check if booking is already fully paid
        if net_paid_amount >= total_amount and total_amount > 0:
            messages.error(request, f'Booking {booking.booking_reference} is already fully paid. Total: Tsh{total_amount:,.0f}, Paid: Tsh{net_paid_amount:,.0f}, Remaining: Tsh{remaining_amount:,.0f}.')
            return redirect('properties:create_payment', booking_id=booking.pk)
        
        # Validate payment doesn't exceed remaining
        if payment_amount > remaining_amount:
            messages.error(request, f'Payment amount (Tsh{payment_amount:,.0f}) exceeds remaining balance (Tsh{remaining_amount:,.0f}). Maximum allowed: Tsh{remaining_amount:,.0f}')
            return redirect('properties:create_payment', booking_id=booking.pk)
        
        # Check if payment exceeds calculated total
        calculated_total = booking.calculated_total_amount
        if payment_amount > calculated_total:
            messages.warning(request, f'Payment amount (Tsh {payment_amount:,.0f}) exceeds calculated total (Tsh {calculated_total:,.0f}).')
        
        # Handle online payment (AZAMpay)
        if payment_method in ['online', 'mobile_money']:
            # Get or create AZAM Pay provider
            provider, _ = PaymentProvider.objects.get_or_create(
                name='AZAM Pay',
                defaults={'description': 'AZAM Pay Payment Gateway'}
            )
            
            # Smart Logic: Use logged-in user as tenant
            # Phone number selection happens in gateway service based on user role:
            # - Admin/Staff: Uses customer phone (from booking) so customer receives payment prompt
            # - Customer: Uses their own profile phone
            customer_user = request.user
            
            # Get mobile money provider if mobile_money payment
            mobile_money_provider = None
            if payment_method == 'mobile_money':
                mobile_money_provider = request.POST.get('mobile_money_provider', '').strip()
                if not mobile_money_provider:
                    messages.error(request, 'Mobile Money Provider is required for mobile money payments. Please select your provider (AIRTEL, TIGO, MPESA, or HALOPESA).')
                    return redirect('properties:create_payment', booking_id=booking.pk)
            
            # Create unified payment record
            unified_payment = UnifiedPayment.objects.create(
                booking=booking,
                tenant=customer_user,
                provider=provider,
                amount=payment_amount,
                payment_method='online' if payment_method == 'online' else 'mobile_money',
                mobile_money_provider=mobile_money_provider,
                status='pending',
                notes=request.POST.get('notes', f'Payment for booking {booking.booking_reference}'),
                recorded_by=request.user,
            )
            
            # Initiate AZAMpay payment
            # Use configured webhook URL (production) instead of localhost
            from django.conf import settings
            callback_url = getattr(settings, 'AZAM_PAY_WEBHOOK_URL', None)
            if not callback_url:
                # Fallback to BASE_URL if webhook URL not configured
                base_domain = getattr(settings, 'BASE_URL', 'https://portal.maishaapp.co.tz')
                callback_url = f"{base_domain}/api/v1/payments/webhook/azam-pay/"
            
            gateway_result = PaymentGatewayService.initiate_payment(
                payment=unified_payment,
                provider_name='azam pay',
                callback_url=callback_url,
                payment_method='mobile_money' if payment_method == 'mobile_money' else 'mobile_money'
            )
            
            if gateway_result.get('success'):
                # Create payment transaction with the actual payload sent to AZAM Pay
                # This includes accountNumber (phone number) used for the payment
                request_payload = gateway_result.get('request_payload', {})
                if not request_payload:
                    # Fallback: create minimal payload if not provided
                    request_payload = {'booking_id': booking.id}
                
                PaymentTransaction.objects.create(
                    payment=unified_payment,
                    provider=provider,
                    gateway_transaction_id=gateway_result.get('transaction_id'),
                    azam_reference=gateway_result.get('reference'),
                    status='initiated',
                    request_payload=request_payload  # Store actual payload sent to AZAM Pay
                )
                
                transaction_id = gateway_result.get('transaction_id')
                payment_message = gateway_result.get('message', 'Payment initiated successfully')
                
                # For MNO (Mobile Money) payments, there's no redirect URL
                # Payment is processed directly on the mobile money network
                payment_link = gateway_result.get('payment_link')
                
                if payment_link:
                    # Bank/Online payment - redirect to payment page
                    request.session['payment_link'] = payment_link
                    request.session['payment_id'] = unified_payment.id
                    messages.info(request, 'Payment initiated. Please complete payment on AZAM Pay.')
                    return redirect(payment_link)
                else:
                    # MNO payment - no redirect, payment initiated on mobile money network
                    messages.success(request, f'Payment initiated successfully! Transaction ID: {transaction_id}. The customer will receive a payment prompt on their phone.')
                    # Redirect back to booking or payment list
                    return redirect('properties:create_payment', booking_id=booking.pk)
            else:
                # Payment initiation failed
                error_msg = gateway_result.get('error', 'Failed to initiate payment')
                messages.error(request, f'Failed to initiate payment: {error_msg}')
                unified_payment.delete()  # Clean up failed payment
                return redirect('properties:create_payment', booking_id=booking.pk)
        
        else:
            # Cash payment - handle differently (no AZAMPAY, requires receipt)
            logger.info(f"Handling cash/offline payment. Payment method: '{payment_method}'")
            if payment_method and payment_method.strip() == 'cash':
                # Validate receipt is provided for cash payments
                if 'receipt' not in request.FILES or not request.FILES.get('receipt'):
                    messages.error(request, 'Receipt/proof of payment is required for cash payments. Please upload a receipt.')
                    return redirect('properties:create_payment', booking_id=booking.pk)
            
            # Offline payment (cash, card, etc.) - use UnifiedPayment model
            payment_date = request.POST.get('payment_date')
            if not payment_date:
                payment_date = timezone.now()
            else:
                from django.utils.dateparse import parse_datetime
                payment_date = parse_datetime(payment_date) or timezone.now()
            
            # Generate transaction reference if not provided
            transaction_ref = request.POST.get('transaction_reference', '').strip()
            if not transaction_ref:
                # Generate auto transaction reference: PAY-BK009017-001
                booking_ref = booking.booking_reference
                last_payment = Payment.objects.filter(
                    booking=booking
                ).exclude(
                    transaction_reference__isnull=True
                ).exclude(
                    transaction_reference=''
                ).order_by('-id').first()
                
                if last_payment and last_payment.transaction_reference and last_payment.transaction_reference.startswith('PAY-'):
                    try:
                        # Extract number from format: PAY-BK009017-001
                        parts = last_payment.transaction_reference.split('-')
                        if len(parts) >= 3:
                            last_num_str = parts[-1]
                            last_num = int(last_num_str)
                            next_num = last_num + 1
                        else:
                            next_num = 1
                    except (ValueError, IndexError):
                        next_num = 1
                else:
                    next_num = 1
                
                transaction_ref = f'PAY-{booking_ref}-{next_num:03d}'
            
            # Recalculate actual paid before creating new payment
            from django.db.models import Sum
            current_paid = Payment.objects.filter(
                booking=booking
            ).exclude(
                payment_type='refund'
            ).aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0')
            
            total_refunded = Payment.objects.filter(
                booking=booking,
                payment_type='refund'
            ).aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0')
            
            current_net_paid = current_paid - total_refunded
            if current_net_paid < 0:
                current_net_paid = Decimal('0')
            
            # Get correct total amount (prefer calculated_total_amount)
            calculated_total = booking.calculated_total_amount or 0
            stored_total = booking.total_amount or 0
            total_amount = calculated_total if calculated_total > 0 else stored_total
            
            # Validate payment doesn't exceed remaining
            remaining = max(Decimal('0'), Decimal(str(total_amount)) - current_net_paid)
            if payment_amount > remaining:
                messages.error(request, f'Payment amount (Tsh{payment_amount:,.0f}) exceeds remaining balance (Tsh{remaining:,.0f}). Maximum allowed: Tsh{remaining:,.0f}')
                return redirect('properties:create_payment', booking_id=booking.pk)
            
            # Smart Logic: Use logged-in user as tenant
            # Phone number selection happens in gateway service based on user role:
            # - Admin/Staff: Uses customer phone (from booking) so customer receives payment prompt
            # - Customer: Uses their own profile phone
            customer_user = request.user
            
            # Create unified payment record for cash payments (bypasses AZAMPAY)
            unified_payment = UnifiedPayment.objects.create(
                booking=booking,
                tenant=customer_user,
                amount=payment_amount,
                payment_method=payment_method,
                paid_date=payment_date.date() if hasattr(payment_date, 'date') else payment_date,
                status='completed',  # Cash payments are completed immediately (no AZAMPAY)
                transaction_ref=transaction_ref,
                reference_number=transaction_ref,
                notes=request.POST.get('notes', f'Payment for booking {booking.booking_reference}'),
                recorded_by=request.user,
            )
            
            # Handle receipt upload for cash payments
            if payment_method == 'cash' and 'receipt' in request.FILES:
                receipt_file = request.FILES['receipt']
                unified_payment.receipt = receipt_file
                unified_payment.save()
            
            # Also create old Payment record for backward compatibility
            payment = Payment.objects.create(
                booking=booking,
                amount=payment_amount,
                payment_method=payment_method,
                payment_type=payment_type,
                transaction_reference=transaction_ref,
                payment_date=payment_date,
                notes=request.POST.get('notes'),
                recorded_by=request.user,
            )
            
            # Recalculate paid_amount from all payments (not just add)
            new_paid = Payment.objects.filter(
                booking=booking
            ).exclude(
                payment_type='refund'
            ).aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0')
            
            new_refunded = Payment.objects.filter(
                booking=booking,
                payment_type='refund'
            ).aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0')
            
            booking.paid_amount = max(Decimal('0'), new_paid - new_refunded)
            booking.update_payment_status()
            
            # Check if stay is over and update booking status
            if booking.is_stay_over and booking.booking_status != 'checked_out':
                booking.booking_status = 'checked_out'
                booking.checked_out_at = timezone.now()
                booking.save()
                messages.info(request, 'Stay period has ended. Booking status updated to Checked Out.')
            
            if payment_method == 'cash':
                messages.success(request, f'Cash payment of Tsh {payment.amount:,.0f} recorded successfully! Receipt uploaded.')
            else:
                messages.success(request, f'Payment of Tsh {payment.amount:,.0f} recorded successfully!')
            return redirect('properties:booking_detail', pk=booking.pk)
    
    # Handle GET request - show form
    # For visit payments, show simpler form
    if is_visit_payment:
        context = {
            'payment': unified_payment,
            'visit_payment': visit_payment,
            'property': property_obj,
            'amount': visit_payment.amount,
            'is_visit_payment': True,
        }
        return render(request, 'properties/create_payment.html', context)
    
    # For booking payments, continue with existing logic
    # Recalculate everything from scratch to ensure consistency
    from django.db.models import Sum, Q
    from decimal import Decimal
    
    # Get actual total paid from all Payment records (exclude refunds)
    actual_paid = Payment.objects.filter(
        booking=booking
    ).exclude(
        payment_type='refund'
    ).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')
    
    # Get total refunded
    total_refunded = Payment.objects.filter(
        booking=booking,
        payment_type='refund'
    ).aggregate(
        total=Sum('amount')
    )['total'] or Decimal('0')
    
    # Net paid amount = payments - refunds
    net_paid_amount = actual_paid - total_refunded
    
    # Ensure net_paid_amount is not negative
    if net_paid_amount < 0:
        net_paid_amount = Decimal('0')
    
    # Get the correct total amount (prefer calculated_total_amount if available)
    calculated_total = booking.calculated_total_amount or 0
    stored_total = booking.total_amount or 0
    
    # Use calculated_total if it exists and is valid, otherwise use stored_total
    if calculated_total and calculated_total > 0:
        total_amount = calculated_total
    else:
        total_amount = stored_total
    
    # If total_amount is still 0 or invalid, try to calculate it
    if not total_amount or total_amount <= 0:
        try:
            total_amount = booking.calculated_total_amount
            if total_amount and total_amount > 0:
                booking.total_amount = total_amount
                booking.save()
        except Exception as calc_error:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Could not calculate total for booking {booking.id}: {str(calc_error)}")
            total_amount = booking.total_amount or 0
    
    # Update booking's paid_amount to match actual payments (fix inconsistency)
    if abs(booking.paid_amount - net_paid_amount) > Decimal('0.01'):  # Allow small rounding differences
        booking.paid_amount = net_paid_amount
        booking.save()
    
    # Calculate remaining amount (cannot be negative)
    remaining_amount = max(Decimal('0'), Decimal(str(total_amount)) - net_paid_amount)
    
    # Update payment status based on actual amounts
    booking.update_payment_status()
    
    # Note: Overpayments are not allowed - validation prevents paid > total
    # If this condition is true, it indicates a data inconsistency that should be investigated
    overpayment = 0
    if net_paid_amount > total_amount and total_amount > 0:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Booking {booking.id} ({booking.booking_reference}): Data inconsistency detected - Paid ({net_paid_amount}) exceeds total ({total_amount}). This should not happen with proper validation.")
        overpayment = float(net_paid_amount - total_amount)
    
    context = {
        'booking': booking,
        'total_amount': float(total_amount),
        'paid_amount': float(net_paid_amount),
        'remaining_amount': float(remaining_amount),
        'calculated_total': float(total_amount),
        'half_amount': float(total_amount / 2) if total_amount > 0 else 0,
        'overpayment': overpayment,
        'daily_rate': booking.daily_rate,
        'duration_days': booking.duration_days,
        'days_remaining': booking.days_remaining,
        'is_stay_over': booking.is_stay_over,
    }
    return render(request, 'properties/create_payment.html', context)


@login_required
def booking_detail(request, pk):
    """View booking details and payment history"""
    booking = get_object_or_404(Booking.objects.select_related('customer', 'property_obj', 'property_obj__property_type'), pk=pk)
    
    # Automatic checkout: If stay period has ended and booking is not already checked out, update status
    if booking.is_stay_over and booking.booking_status not in ['checked_out', 'cancelled']:
        booking.booking_status = 'checked_out'
        if not booking.checked_out_at:
            from django.utils import timezone
            booking.checked_out_at = timezone.now()
        booking.save()
    
    # Get payment history for this booking
    payments = Payment.objects.filter(booking=booking).order_by('-payment_date')
    
    context = {
        'booking': booking,
        'payments': payments,
    }
    return render(request, 'properties/booking_detail.html', context)


# API Views for Payment Modals
@login_required
def api_payment_details(request, payment_id):
    """API endpoint to get payment details for modals"""
    try:
        payment = Payment.objects.select_related(
            'booking', 'booking__customer', 'booking__property_obj', 'booking__property_obj__property_type', 'recorded_by'
        ).get(id=payment_id)
        
        # Check if this should return HTML for modal or JSON
        # If it's a modal request (not explicitly asking for JSON), render template
        if request.META.get('HTTP_ACCEPT', '').find('json') == -1:
            context = {
                'payment': payment,
                'payment_type': 'booking',
            }
            return render(request, 'properties/modals/payment_view_details.html', context)
        
        # Return JSON data
        data = {
            'id': payment.id,
            'amount': float(payment.amount) if payment.amount else 0,
            'payment_method': payment.payment_method or '',
            'payment_method_display': payment.get_payment_method_display() if payment.payment_method else '',
            'payment_type': payment.payment_type or '',
            'payment_type_display': payment.get_payment_type_display() if payment.payment_type else '',
            'payment_date': payment.payment_date.strftime('%Y-%m-%d %H:%M') if payment.payment_date else '',
            'transaction_reference': payment.transaction_reference or '',
            'notes': payment.notes or '',
            'created_at': payment.created_at.strftime('%Y-%m-%d %H:%M') if payment.created_at else '',
            'recorded_by': payment.recorded_by.get_full_name() or payment.recorded_by.username if payment.recorded_by else 'N/A',
        }
        
        # Add booking data if available
        if payment.booking:
            data['booking'] = {
                'id': payment.booking.id,
                'booking_reference': payment.booking.booking_reference or '',
                'room_number': payment.booking.room_number or '',
                'check_in_date': payment.booking.check_in_date.strftime('%Y-%m-%d') if payment.booking.check_in_date else '',
                'check_out_date': payment.booking.check_out_date.strftime('%Y-%m-%d') if payment.booking.check_out_date else '',
                'total_amount': float(payment.booking.total_amount) if payment.booking.total_amount else 0,
                'paid_amount': float(payment.booking.paid_amount) if payment.booking.paid_amount else 0,
                'customer': {
                    'id': payment.booking.customer.id if payment.booking.customer else None,
                    'full_name': payment.booking.customer.full_name if payment.booking.customer else 'N/A',
                    'email': payment.booking.customer.email if payment.booking.customer else 'N/A',
                    'phone': payment.booking.customer.phone if payment.booking.customer else 'N/A',
                } if payment.booking.customer else {}
            }
        
        return JsonResponse(data)
    except Payment.DoesNotExist:
        return JsonResponse({'error': 'Payment not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading payment details for payment {payment_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error loading payment details: {str(e)}'}, status=500)


@login_required
def api_create_booking(request):
    """API endpoint to create booking from modal form - supports hotel, lodge, and venue"""
    if request.method == 'POST':
        try:
            # CRITICAL: Determine property type from HTTP_REFERER or venue parameter
            referer = request.META.get('HTTP_REFERER', '')
            property_type = None
            
            # Check if venue is explicitly provided in POST data (for venue bookings)
            venue_id = request.POST.get('venue', '').strip()
            if venue_id:
                property_type = 'venue'
            # Check referer URL to determine if this is a hotel or lodge request
            elif '/lodge/' in referer:
                property_type = 'lodge'
            elif '/hotel/' in referer:
                property_type = 'hotel'
            elif '/venue/' in referer:
                property_type = 'venue'
            
            # If we cannot determine from referer or POST, return error
            if not property_type:
                return JsonResponse({
                    'success': False,
                    'error': 'Unable to determine property type. Please access this from a hotel, lodge, or venue bookings page.'
                }, status=400)
            
            # Get the correct property ID based on property type
            if property_type == 'lodge':
                selected_property_id = request.session.get('selected_lodge_property_id')
                if not selected_property_id:
                    return JsonResponse({
                        'success': False,
                        'error': 'No lodge selected. Please select a lodge first.'
                    }, status=400)
            elif property_type == 'hotel':
                selected_property_id = request.session.get('selected_hotel_property_id')
                if not selected_property_id:
                    return JsonResponse({
                        'success': False,
                        'error': 'No hotel selected. Please select a hotel first.'
                    }, status=400)
            elif property_type == 'venue':
                # For venues, use the venue ID from POST data
                selected_property_id = venue_id
                if not selected_property_id:
                    return JsonResponse({
                        'success': False,
                        'error': 'No venue selected. Please select a venue.'
                    }, status=400)
            
            selected_property_id = validate_property_id(selected_property_id)
            
            if not selected_property_id:
                return JsonResponse({
                    'success': False,
                    'error': f'No {property_type} selected. Please select a {property_type} first.'
                }, status=400)
            
            # Validate property type
            if property_type not in ['hotel', 'lodge', 'venue']:
                return JsonResponse({
                    'success': False,
                    'error': 'This endpoint only supports hotel, lodge, and venue properties.'
                }, status=400)
            
            # CRITICAL: Validate that the selected property matches the property type
            # This ensures lodge NEVER reads hotel data and vice versa
            try:
                selected_property = Property.objects.get(id=selected_property_id)
                actual_property_type = selected_property.property_type.name.lower() if selected_property.property_type else None
                
                if actual_property_type != property_type:
                    return JsonResponse({
                        'success': False,
                        'error': f'Property type mismatch. Expected {property_type}, but property is {actual_property_type}. Lodge and Hotel are completely separate.'
                    }, status=400)
                
                # Double-check with explicit filter
                selected_property = Property.objects.get(id=selected_property_id, property_type__name__iexact=property_type)
            except Property.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': f'Selected {property_type} not found or type mismatch. Lodge and Hotel are completely separate.'
                }, status=400)
            
            # Get form data - different fields for venue vs hotel/lodge
            customer_name = request.POST.get('customer_name', '').strip()
            phone = request.POST.get('phone', '').strip()
            email = request.POST.get('email', '').strip()
            
            # For venue bookings, use different field names
            if property_type == 'venue':
                event_name = request.POST.get('event_name', '').strip()
                event_type = request.POST.get('event_type', '').strip()
                event_date = request.POST.get('event_date', '').strip()
                check_in_date = event_date  # For venues, event_date is check_in_date
                check_out_date = request.POST.get('check_out_date', '').strip() or event_date
                number_of_guests = int(request.POST.get('expected_guests', request.POST.get('number_of_guests', 1)))
                total_amount = float(request.POST.get('total_amount', 0))
                special_requests = request.POST.get('special_requirements', request.POST.get('special_requests', '')).strip()
                room_type = event_type  # Use event_type as room_type for venues
                room_number = None  # Venues don't have rooms
                
                # Validate venue capacity
                if selected_property.capacity and number_of_guests > selected_property.capacity:
                    return JsonResponse({
                        'success': False,
                        'error': f'Expected guests ({number_of_guests}) exceeds venue capacity ({selected_property.capacity}).'
                    }, status=400)
                
                # Validate required fields for venue
                if not all([event_name, customer_name, phone, email, event_type, event_date]):
                    return JsonResponse({
                        'success': False,
                        'error': 'Please fill in all required fields including event name, customer name, phone number, email, event type, and event date.'
                    }, status=400)
                
                # Validate total_amount is provided and greater than 0
                if not total_amount or total_amount <= 0:
                    return JsonResponse({
                        'success': False,
                        'error': 'Total amount must be greater than 0. Please ensure the amount is calculated correctly.'
                    }, status=400)
            else:
                # Hotel/Lodge bookings
                room_type = request.POST.get('room_type', '').strip()
                room_number = request.POST.get('room_number', '').strip()
                check_in_date = request.POST.get('check_in_date', '').strip()
                check_out_date = request.POST.get('check_out_date', '').strip()
                number_of_guests = int(request.POST.get('number_of_guests', 1))
                total_amount = float(request.POST.get('total_amount', 0))
                special_requests = request.POST.get('special_requests', '').strip()
                
                # Validate required fields for hotel/lodge
                if not all([customer_name, phone, email, room_type, check_in_date, check_out_date, total_amount]):
                    return JsonResponse({
                        'success': False,
                        'error': 'Please fill in all required fields including customer name, phone number, and email.'
                    }, status=400)
            
            # Check property availability
            from datetime import datetime
            try:
                check_in = datetime.strptime(check_in_date, '%Y-%m-%d').date()
                check_out = datetime.strptime(check_out_date, '%Y-%m-%d').date()
                
                if not selected_property.is_available_for_booking(check_in, check_out):
                    return JsonResponse({
                        'success': False,
                        'error': 'Property is not available for the selected dates. Please choose different dates.'
                    }, status=400)
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid date format.'
                }, status=400)
            
            # Validate email format
            import re
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            if not re.match(email_pattern, email):
                return JsonResponse({
                    'success': False,
                    'error': 'Please provide a valid email address.'
                }, status=400)
            
            # Parse customer name
            name_parts = customer_name.split(' ', 1)
            first_name = name_parts[0]
            last_name = name_parts[1] if len(name_parts) > 1 else ''
            
            # Create or get customer - email is UNIQUE, so use it as primary identifier
            # Since email is unique, we can safely look up by email and update other fields
            try:
                customer, created = Customer.objects.get_or_create(
                    email=email,
                    defaults={
                        'first_name': first_name,
                        'last_name': last_name,
                        'phone': phone,
                    }
                )
                
                # If customer already existed (not created), update name and phone if different
                if not created:
                    updated = False
                    if customer.first_name != first_name or customer.last_name != last_name:
                        customer.first_name = first_name
                        customer.last_name = last_name
                        updated = True
                    if customer.phone != phone:
                        customer.phone = phone
                        updated = True
                    if updated:
                        customer.save()
            except Exception as e:
                # Handle any database errors (e.g., unique constraint violation)
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error creating/updating customer: {str(e)}")
                return JsonResponse({
                    'success': False,
                    'error': f'Error saving customer information: {str(e)}. Please check that the email is unique.'
                }, status=400)
            
            # Generate booking reference based on property type
            if property_type == 'lodge':
                booking_reference = f"LDG-{Booking.objects.count() + 1:06d}"
            elif property_type == 'venue':
                booking_reference = f"VEN-{Booking.objects.count() + 1:06d}"
            else:
                booking_reference = f"HTL-{Booking.objects.count() + 1:06d}"
            
            # For venue bookings, use event_name as special_requests if provided
            if property_type == 'venue' and event_name:
                special_requests = f"{event_name}" + (f" - {special_requests}" if special_requests else "")
            
            # Create booking - ensure total_amount is Decimal
            from decimal import Decimal
            total_amount_decimal = Decimal(str(total_amount)) if total_amount else Decimal('0')
            
            # Create booking - use parsed date objects (check_in, check_out) instead of strings
            booking = Booking.objects.create(
                property_obj=selected_property,
                customer=customer,
                booking_reference=booking_reference,
                check_in_date=check_in,  # Use parsed date object, not string
                check_out_date=check_out,  # Use parsed date object, not string
                number_of_guests=number_of_guests,
                room_type=room_type,
                total_amount=total_amount_decimal,
                special_requests=special_requests,
                created_by=request.user,
            )
            
            # Handle room assignment - only for hotel/lodge, not venues
            room_assigned = False
            room_message = ""
            
            if property_type != 'venue':
                # Room selection is mandatory for hotel/lodge
                if not room_number:
                    return JsonResponse({
                        'success': False,
                        'error': 'Please select a room number.'
                    }, status=400)
                
                # Validate and assign the selected room
                try:
                    room = Room.objects.get(
                        property_obj=selected_property, 
                        room_number=room_number,
                        room_type=room_type
                    )
                    
                    if room.status == 'available':
                        # Assign the selected room
                        booking.room_number = room_number
                        room.current_booking = booking
                        room.status = 'occupied'
                        room.save()
                        booking.save()
                        room_assigned = True
                        room_message = f"Room {room_number} assigned successfully"
                    else:
                        return JsonResponse({
                            'success': False,
                            'error': f'Room {room_number} is not available (Status: {room.status}). Please select an available room.'
                        }, status=400)
                        
                except Room.DoesNotExist:
                    return JsonResponse({
                        'success': False,
                        'error': f'Room {room_number} not found. Please select a valid room.'
                    }, status=400)
            
            return JsonResponse({
                'success': True,
                'message': f'Booking {booking.booking_reference} created successfully!',
                'booking_reference': booking.booking_reference,
                'room_message': room_message,
                'room_assigned': room_assigned,
                'room_number': booking.room_number or 'Not assigned'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error creating booking: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    }, status=405)


@login_required
def api_available_rooms(request):
    """API endpoint to get available rooms for booking modal - STRICTLY separates hotel and lodge"""
    try:
        # CRITICAL: Determine property type from HTTP_REFERER - this is REQUIRED
        # Lodge and Hotel are COMPLETELY SEPARATE - never mix them
        referer = request.META.get('HTTP_REFERER', '')
        property_type = None
        
        # Check referer URL to determine if this is a hotel or lodge request
        # This is the PRIMARY way to determine context - no fallbacks that mix
        if '/lodge/' in referer:
            property_type = 'lodge'
        elif '/hotel/' in referer:
            property_type = 'hotel'
        
        # If we cannot determine from referer, return error - DO NOT GUESS
        if not property_type:
            return JsonResponse({
                'success': False,
                'error': 'Unable to determine property type. Please access this from a hotel or lodge bookings page.'
            }, status=400)
        
        # Get the correct property ID based on property type - STRICTLY SEPARATE
        if property_type == 'lodge':
            selected_property_id = request.session.get('selected_lodge_property_id')
            # Validate that we're using lodge data, not hotel
            if not selected_property_id:
                return JsonResponse({
                    'success': False,
                    'error': 'No lodge selected. Please select a lodge first.'
                }, status=400)
        elif property_type == 'hotel':
            selected_property_id = request.session.get('selected_hotel_property_id')
            # Validate that we're using hotel data, not lodge
            if not selected_property_id:
                return JsonResponse({
                    'success': False,
                    'error': 'No hotel selected. Please select a hotel first.'
                }, status=400)
        
        selected_property_id = validate_property_id(selected_property_id)
        
        if not selected_property_id:
            return JsonResponse({
                'success': False,
                'error': f'No {property_type} selected. Please select a {property_type} first.'
            }, status=400)
        
        # CRITICAL: Validate that the selected property matches the property type
        # This ensures lodge NEVER reads hotel data and vice versa
        try:
            selected_property = Property.objects.get(id=selected_property_id)
            actual_property_type = selected_property.property_type.name.lower() if selected_property.property_type else None
            
            if actual_property_type != property_type:
                return JsonResponse({
                    'success': False,
                    'error': f'Property type mismatch. Expected {property_type}, but property is {actual_property_type}. Lodge and Hotel are completely separate.'
                }, status=400)
        except Property.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': f'{property_type.capitalize()} property not found.'
            }, status=404)
        
        # Get rooms for hotel or lodge - STRICTLY filtered by property type
        rooms = Room.objects.filter(
            property_obj__property_type__name__iexact=property_type,
            property_obj_id=selected_property_id
        )
        
        rooms_data = []
        for room in rooms:
            # Only include rooms that have a valid base_rate set
            # Each room MUST have its own price - no default/fallback pricing
            if room.base_rate and room.base_rate > 0:
                # Sync room status to ensure it's up-to-date with bookings
                # This fixes cases where current_booking points to cancelled bookings
                room.sync_status_from_bookings()
                
                rooms_data.append({
                    'id': room.id,
                    'room_number': room.room_number,
                    'room_type': room.room_type,
                    'capacity': room.capacity,
                    'base_rate': float(room.base_rate),  # Use ONLY room's base_rate - no fallback
                    'status': room.status,
                    'floor_number': room.floor_number,
                    'bed_type': room.bed_type,
                    'amenities': room.amenities,
                })
            # Skip rooms without a valid base_rate (they shouldn't be bookable)
        
        return JsonResponse({
            'success': True,
            'rooms': rooms_data,
            'property_type': property_type
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def api_hotel_room_status_summary(request, property_id):
    """API endpoint to get hotel room status summary for quick view"""
    from django.http import JsonResponse
    from properties.models import Room, Booking
    from django.utils import timezone
    
    try:
        property_obj = Property.objects.get(id=property_id, property_type__name__iexact='hotel')
        
        # Get all rooms for this hotel
        rooms = Room.objects.filter(property_obj=property_obj, is_active=True).select_related('current_booking', 'current_booking__customer')
        
        # Sync room statuses
        for room in rooms:
            room.sync_status_from_bookings()
        
        # Categorize rooms
        available_rooms = []
        occupied_rooms = []
        maintenance_rooms = []
        out_of_order_rooms = []
        
        for room in rooms:
            room_data = {
                'id': room.id,
                'room_number': room.room_number,
                'room_type': room.room_type,
                'floor_number': room.floor_number,
                'status': room.status,
            }
            
            if room.status == 'available':
                available_rooms.append(room_data)
            elif room.status == 'occupied':
                # Get booking info if available
                booking_info = None
                if room.current_booking:
                    booking_info = {
                        'booking_id': room.current_booking.id,
                        'customer_name': room.current_booking.customer.full_name if room.current_booking.customer else 'N/A',
                        'check_in': room.current_booking.check_in_date.strftime('%b %d, %Y') if room.current_booking.check_in_date else 'N/A',
                        'check_out': room.current_booking.check_out_date.strftime('%b %d, %Y') if room.current_booking.check_out_date else 'N/A',
                    }
                room_data['booking'] = booking_info
                occupied_rooms.append(room_data)
            elif room.status == 'maintenance':
                maintenance_rooms.append(room_data)
            elif room.status == 'out_of_order':
                out_of_order_rooms.append(room_data)
        
        # Calculate statistics
        total_rooms = rooms.count()
        available_count = len(available_rooms)
        occupied_count = len(occupied_rooms)
        maintenance_count = len(maintenance_rooms)
        out_of_order_count = len(out_of_order_rooms)
        
        return JsonResponse({
            'success': True,
            'property': {
                'id': property_obj.id,
                'title': property_obj.title,
                'address': property_obj.address,
            },
            'summary': {
                'total_rooms': total_rooms,
                'available_count': available_count,
                'occupied_count': occupied_count,
                'maintenance_count': maintenance_count,
                'out_of_order_count': out_of_order_count,
            },
            'rooms': {
                'available': available_rooms,
                'occupied': occupied_rooms,
                'maintenance': maintenance_rooms,
                'out_of_order': out_of_order_rooms,
            }
        })
        
    except Property.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Hotel property not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def api_booking_status(request, booking_id):
    """API endpoint to get booking payment status"""
    try:
        booking = Booking.objects.get(id=booking_id)
        return JsonResponse({
            'booking_id': booking.id,
            'booking_reference': booking.booking_reference,
            'total_amount': float(booking.total_amount),
            'paid_amount': float(booking.paid_amount),
            'remaining_balance': float(booking.total_amount - booking.paid_amount),
            'payment_status': booking.payment_status,
            'booking_status': booking.booking_status
        })
    except Booking.DoesNotExist:
        return JsonResponse({'error': 'Booking not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def api_collect_payment(request):
    """API endpoint to collect a new payment"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        # Get booking ID from the request (this should be passed from the frontend)
        booking_id = request.POST.get('booking_id')
        if not booking_id:
            return JsonResponse({'error': 'Booking ID is required'}, status=400)
        
        booking = Booking.objects.get(id=booking_id)
        
        # Get payment data
        amount = request.POST.get('amount')
        payment_method = request.POST.get('payment_method')
        payment_type = request.POST.get('payment_type', 'partial')
        transaction_reference = request.POST.get('transaction_reference', '')
        payment_date_str = request.POST.get('payment_date')
        notes = request.POST.get('notes', '')
        
        # Validate required fields
        if not amount or not payment_method or not payment_date_str:
            return JsonResponse({'error': 'Amount, payment method, and payment date are required'}, status=400)
        
        # Convert amount to decimal for calculations
        from decimal import Decimal
        try:
            payment_amount = Decimal(str(amount))
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid amount format'}, status=400)
        
        # Check if booking is already fully paid
        if booking.payment_status == 'paid':
            return JsonResponse({
                'error': f'Booking {booking.booking_reference} is already fully paid (Tsh{booking.total_amount:,.0f}). Please start a new booking for additional payments.'
            }, status=400)
        
        # Calculate remaining balance (use calculated_total_amount if available)
        calculated_total = booking.calculated_total_amount if booking.calculated_total_amount and booking.calculated_total_amount > 0 else (booking.total_amount or 0)
        remaining_balance = max(Decimal('0'), Decimal(str(calculated_total)) - booking.paid_amount)
        
        # Validate payment amount doesn't exceed remaining balance
        if payment_amount > remaining_balance:
            return JsonResponse({
                'error': f'Payment amount (Tsh{payment_amount}) exceeds remaining balance (Tsh{remaining_balance}). Maximum payment allowed: Tsh{remaining_balance}'
            }, status=400)
        
        # If payment amount equals remaining balance, set payment type to 'full'
        if payment_amount == remaining_balance:
            payment_type = 'full'
        
        # Parse payment date
        from datetime import datetime
        try:
            payment_date = datetime.fromisoformat(payment_date_str.replace('T', ' '))
        except ValueError:
            return JsonResponse({'error': 'Invalid payment date format'}, status=400)
        
        # Create new payment
        # Validate payment amount doesn't exceed remaining balance
        from django.db.models import Sum
        current_paid = Payment.objects.filter(
            booking=booking
        ).exclude(
            payment_type='refund'
        ).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0')
        
        total_refunded = Payment.objects.filter(
            booking=booking,
            payment_type='refund'
        ).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0')
        
        current_net_paid = max(Decimal('0'), current_paid - total_refunded)
        calculated_total = booking.calculated_total_amount if booking.calculated_total_amount and booking.calculated_total_amount > 0 else (booking.total_amount or 0)
        remaining_balance = max(Decimal('0'), Decimal(str(calculated_total)) - current_net_paid)
        
        # Validate payment amount doesn't exceed remaining balance
        if payment_amount > remaining_balance:
            return JsonResponse({
                'success': False,
                'error': f'Payment amount (Tsh{payment_amount:,.2f}) exceeds remaining balance (Tsh{remaining_balance:,.2f}). Maximum allowed: Tsh{remaining_balance:,.2f}'
            }, status=400)
        
        payment = Payment.objects.create(
            booking=booking,
            amount=payment_amount,
            payment_method=payment_method,
            payment_type=payment_type,
            transaction_reference=transaction_reference,
            payment_date=payment_date,
            notes=notes,
            recorded_by=request.user
        )
        
        # Recalculate booking payment status from all payments
        actual_paid = Payment.objects.filter(
            booking=booking
        ).exclude(
            payment_type='refund'
        ).aggregate(
            total=Sum('amount')
        )['total'] or Decimal('0')
        
        net_paid_amount = max(Decimal('0'), actual_paid - total_refunded)
        booking.paid_amount = net_paid_amount
        booking.update_payment_status()
        booking.save()
        
        return JsonResponse({
            'success': True,
            'payment_id': payment.id,
            'message': 'Payment collected successfully'
        })
        
    except Booking.DoesNotExist:
        return JsonResponse({'error': 'Booking not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error creating payment: {str(e)}'}, status=500)


def generate_refund_reference():
    """Generate next refund reference number"""
    last_refund = Payment.objects.filter(
        payment_type='refund',
        transaction_reference__startswith='REF-'
    ).order_by('-transaction_reference').first()
    
    if last_refund and last_refund.transaction_reference:
        try:
            # Extract number from last reference (e.g., "REF-001" -> 1)
            last_number = int(last_refund.transaction_reference.split('-')[1])
            next_number = last_number + 1
        except (ValueError, IndexError):
            next_number = 1
    else:
        next_number = 1
    
    return f"REF-{next_number:03d}"


@login_required
def api_process_refund(request):
    """API endpoint to process a refund"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        payment_id = request.POST.get('payment_id')
        if not payment_id:
            return JsonResponse({'error': 'Payment ID is required'}, status=400)
        
        original_payment = Payment.objects.get(id=payment_id)
        booking = original_payment.booking
        
        # Get refund data
        refund_amount_str = request.POST.get('amount')
        refund_method = request.POST.get('refund_method')
        refund_reference = request.POST.get('refund_reference', '')
        refund_date_str = request.POST.get('refund_date')
        reason = request.POST.get('reason', '')
        
        # Auto-generate refund reference if not provided
        if not refund_reference:
            refund_reference = generate_refund_reference()
        
        # Validate required fields
        if not refund_amount_str or not refund_method or not refund_date_str:
            return JsonResponse({'error': 'Refund amount, method, and date are required'}, status=400)
        
        # Convert refund amount to decimal
        from decimal import Decimal
        try:
            refund_amount = Decimal(str(refund_amount_str))
        except (ValueError, TypeError):
            return JsonResponse({'error': 'Invalid refund amount format'}, status=400)
        
        # Validate refund amount doesn't exceed paid amount
        if refund_amount > booking.paid_amount:
            return JsonResponse({
                'error': f'Refund amount (Tsh{refund_amount}) exceeds paid amount (Tsh{booking.paid_amount}). Maximum refund allowed: Tsh{booking.paid_amount}'
            }, status=400)
        
        # Parse refund date
        from datetime import datetime
        try:
            refund_date = datetime.fromisoformat(refund_date_str.replace('T', ' '))
        except ValueError:
            return JsonResponse({'error': 'Invalid refund date format'}, status=400)
        
        # Create refund payment
        refund_payment = Payment.objects.create(
            booking=booking,
            amount=refund_amount,
            payment_method=refund_method,
            payment_type='refund',
            transaction_reference=refund_reference,
            payment_date=refund_date,
            notes=f"Refund for Payment #{original_payment.id}. Reason: {reason}",
            recorded_by=request.user,
            status='active'  # Refund payments are active
        )
        
        # Mark original payment as refunded
        original_payment.status = 'refunded'
        original_payment.save()
        
        # Update booking payment status
        booking.paid_amount -= refund_amount
        booking.update_payment_status()
        booking.save()
        
        return JsonResponse({
            'success': True,
            'refund_id': refund_payment.id,
            'message': 'Refund processed successfully'
        })
        
    except Payment.DoesNotExist:
        return JsonResponse({'error': 'Payment not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error processing refund: {str(e)}'}, status=500)


# Room Management API Views

@login_required
def api_room_detail(request, room_id):
    """Get room details for modal display"""
    try:
        room = Room.objects.select_related('property_obj').get(id=room_id)
        
        # Check if user has permission to view this room
        if not request.user.has_perm('properties.view_room') and room.property_obj.owner != request.user:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        data = {
            'id': room.id,
            'room_number': room.room_number,
            'room_type': room.room_type,
            'floor_number': room.floor_number,
            'capacity': room.capacity,
            'bed_type': room.bed_type or 'N/A',
            'base_rate': float(room.base_rate),
            'status': room.status,
            'amenities': room.amenities or 'No amenities specified',
            'property_title': room.property_obj.title,
            'is_available': room.is_available,
            'created_at': room.created_at.strftime('%Y-%m-%d %H:%M'),
            'updated_at': room.updated_at.strftime('%Y-%m-%d %H:%M'),
        }
        
        return JsonResponse({'success': True, 'room': data})
        
    except Room.DoesNotExist:
        return JsonResponse({'error': 'Room not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error retrieving room details: {str(e)}'}, status=500)


@login_required
def api_edit_room(request, room_id):
    """Update room details"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        room = Room.objects.get(id=room_id)
        
        # Check if user has permission to edit this room
        if not request.user.has_perm('properties.change_room') and room.property_obj.owner != request.user:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        # Update room fields
        room.room_number = request.POST.get('room_number', room.room_number)
        room.room_type = request.POST.get('room_type', room.room_type)
        room.capacity = int(request.POST.get('capacity', room.capacity))
        
        # Handle floor number
        floor_number = request.POST.get('floor_number')
        if floor_number and floor_number.strip():
            room.floor_number = int(floor_number)
        else:
            room.floor_number = None
            
        room.bed_type = request.POST.get('bed_type', room.bed_type)
        room.base_rate = float(request.POST.get('base_rate', room.base_rate))
        room.status = request.POST.get('status', room.status)
        room.amenities = request.POST.get('amenities', room.amenities)
        
        # Check for duplicate room number within the same property
        if Room.objects.filter(property_obj=room.property_obj, room_number=room.room_number).exclude(id=room.id).exists():
            return JsonResponse({'error': f'Room number {room.room_number} already exists in this hotel'}, status=400)
        
        room.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Room {room.room_number} updated successfully',
            'room': {
                'id': room.id,
                'room_number': room.room_number,
                'room_type': room.room_type,
                'status': room.status,
                'capacity': room.capacity,
                'base_rate': float(room.base_rate),
            }
        })
        
    except Room.DoesNotExist:
        return JsonResponse({'error': 'Room not found'}, status=404)
    except ValueError as e:
        return JsonResponse({'error': f'Invalid data: {str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Error updating room: {str(e)}'}, status=500)


@login_required
def api_update_room_status(request, room_id):
    """Update room status"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        room = Room.objects.get(id=room_id)
        
        # Check if user has permission to edit this room
        if not request.user.has_perm('properties.change_room') and room.property_obj.owner != request.user:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        new_status = request.POST.get('status')
        reason = request.POST.get('reason', '')
        
        if not new_status or new_status not in [choice[0] for choice in Room.ROOM_STATUS_CHOICES]:
            return JsonResponse({'error': 'Invalid status'}, status=400)
        
        old_status = room.status
        room.status = new_status
        room.save()
        
        # Log the status change (optional - you could create a RoomStatusLog model)
        
        return JsonResponse({
            'success': True,
            'message': f'Room {room.room_number} status changed from {old_status} to {new_status}',
            'room': {
                'id': room.id,
                'room_number': room.room_number,
                'old_status': old_status,
                'new_status': new_status,
                'reason': reason,
            }
        })
        
    except Room.DoesNotExist:
        return JsonResponse({'error': 'Room not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error updating room status: {str(e)}'}, status=500)


@login_required
def api_delete_room(request, room_id):
    """Delete a room"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        room = Room.objects.get(id=room_id)
        
        # Check if user has permission to delete this room
        if not request.user.has_perm('properties.delete_room') and room.property_obj.owner != request.user:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        # Check if room has active bookings
        if room.current_booking:
            return JsonResponse({'error': 'Cannot delete room with active booking'}, status=400)
        
        # Check for any future bookings
        from datetime import datetime
        future_bookings = Booking.objects.filter(
            room=room,
            check_in_date__gte=datetime.now().date()
        ).exists()
        
        if future_bookings:
            return JsonResponse({'error': 'Cannot delete room with future bookings'}, status=400)
        
        room_number = room.room_number
        property_title = room.property_obj.title
        room.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Room {room_number} from {property_title} deleted successfully'
        })
        
    except Room.DoesNotExist:
        return JsonResponse({'error': 'Room not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error deleting room: {str(e)}'}, status=500)


# Booking Modal API Views

@login_required
def api_booking_details(request, booking_id):
    """Get booking details for modal display"""
    try:
        booking = Booking.objects.select_related('customer', 'property_obj', 'property_obj__property_type', 'created_by').get(id=booking_id)
        
        # Check if user has permission to view this booking
        if not request.user.has_perm('properties.view_booking') and booking.created_by != request.user:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        # Auto-correct total amount if there is a mismatch with calculated amount
        try:
            # Safely calculate calculated_total_amount
            calculated_total = booking.calculated_total_amount
            if booking.total_amount != calculated_total:
                booking.calculate_and_update_total()
        except Exception as calc_error:
            # Fail silently in details view; leave existing amount if calculation fails
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Could not calculate total for booking {booking_id}: {str(calc_error)}")
        
        # Ensure property has required fields
        if not booking.property_obj:
            return JsonResponse({'error': 'Property not found for this booking'}, status=404)
        
        # Check if property has rent_amount and rent_period
        if booking.property_obj.rent_amount is None:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Property {booking.property_obj.id} has no rent_amount set")
        
        context = {
            'booking': booking,
        }
        return render(request, 'properties/modals/booking_details.html', context)
        
    except Booking.DoesNotExist:
        return JsonResponse({'error': 'Booking not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading booking details for booking {booking_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error loading booking details: {str(e)}'}, status=500)


@login_required
def api_booking_edit(request, booking_id):
    """Get booking edit form for modal display"""
    try:
        booking = Booking.objects.select_related('customer', 'property_obj', 'property_obj__property_type', 'created_by').get(id=booking_id)
        
        # Check if user has permission to edit this booking
        if not request.user.has_perm('properties.change_booking') and booking.created_by != request.user:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        if request.method == 'POST':
            # Handle form submission
            try:
                from datetime import datetime
                check_in_str = request.POST.get('check_in_date')
                check_out_str = request.POST.get('check_out_date')
                
                if check_in_str:
                    booking.check_in_date = datetime.strptime(check_in_str, '%Y-%m-%d').date()
                if check_out_str:
                    booking.check_out_date = datetime.strptime(check_out_str, '%Y-%m-%d').date()
                
                number_of_guests = request.POST.get('number_of_guests')
                if number_of_guests:
                    booking.number_of_guests = int(number_of_guests)
                
                total_amount = request.POST.get('total_amount')
                if total_amount:
                    booking.total_amount = float(total_amount)
                
                booking.special_requests = request.POST.get('special_requests', booking.special_requests)
                booking.save()
                
                return JsonResponse({'success': True, 'message': 'Booking updated successfully'})
            except Exception as save_error:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error saving booking {booking_id}: {str(save_error)}")
                return JsonResponse({'error': f'Error updating booking: {str(save_error)}'}, status=500)
        
        context = {
            'booking': booking,
        }
        return render(request, 'properties/modals/booking_edit.html', context)
        
    except Booking.DoesNotExist:
        return JsonResponse({'error': 'Booking not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading booking edit form for booking {booking_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error loading booking edit form: {str(e)}'}, status=500)


@login_required
def api_booking_confirm(request, booking_id):
    """Confirm booking modal"""
    try:
        booking = Booking.objects.select_related('customer', 'property_obj', 'property_obj__property_type', 'created_by').get(id=booking_id)
        
        # Check if user has permission to confirm this booking
        if not request.user.has_perm('properties.change_booking') and booking.created_by != request.user:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        if request.method == 'POST':
            try:
                # Confirm the booking
                booking.booking_status = 'confirmed'
                booking.confirmed_at = timezone.now()
                booking.save()
                
                return JsonResponse({'success': True, 'message': 'Booking confirmed successfully'})
            except Exception as save_error:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error confirming booking {booking_id}: {str(save_error)}")
                return JsonResponse({'error': f'Error confirming booking: {str(save_error)}'}, status=500)
        
        context = {
            'booking': booking,
        }
        return render(request, 'properties/modals/booking_confirm.html', context)
        
    except Booking.DoesNotExist:
        return JsonResponse({'error': 'Booking not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading booking confirm for booking {booking_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error confirming booking: {str(e)}'}, status=500)


@login_required
def api_booking_checkin(request, booking_id):
    """Check-in booking modal"""
    try:
        booking = Booking.objects.select_related('customer', 'property_obj', 'property_obj__property_type', 'created_by').get(id=booking_id)
        
        # Check if user has permission to check-in this booking
        if not request.user.has_perm('properties.change_booking') and booking.created_by != request.user:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        if request.method == 'POST':
            try:
                # Process check-in
                booking.booking_status = 'checked_in'
                booking.checked_in_at = timezone.now()
                booking.save()
                
                return JsonResponse({'success': True, 'message': 'Guest checked in successfully'})
            except Exception as save_error:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error checking in booking {booking_id}: {str(save_error)}")
                return JsonResponse({'error': f'Error processing check-in: {str(save_error)}'}, status=500)
        
        context = {
            'booking': booking,
        }
        return render(request, 'properties/modals/booking_checkin.html', context)
        
    except Booking.DoesNotExist:
        return JsonResponse({'error': 'Booking not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading booking checkin for booking {booking_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error processing check-in: {str(e)}'}, status=500)


@login_required
def api_booking_checkout(request, booking_id):
    """Check-out booking modal"""
    try:
        booking = Booking.objects.select_related('customer', 'property_obj', 'property_obj__property_type', 'created_by').get(id=booking_id)
        
        # Check if user has permission to check-out this booking
        if not request.user.has_perm('properties.change_booking') and booking.created_by != request.user:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        if request.method == 'POST':
            try:
                # Process check-out
                booking.booking_status = 'checked_out'
                booking.checked_out_at = timezone.now()
                booking.save()
                
                return JsonResponse({'success': True, 'message': 'Guest checked out successfully'})
            except Exception as save_error:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error checking out booking {booking_id}: {str(save_error)}")
                return JsonResponse({'error': f'Error processing check-out: {str(save_error)}'}, status=500)
        
        context = {
            'booking': booking,
        }
        return render(request, 'properties/modals/booking_checkout.html', context)
        
    except Booking.DoesNotExist:
        return JsonResponse({'error': 'Booking not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading booking checkout for booking {booking_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error processing check-out: {str(e)}'}, status=500)


@login_required
def api_booking_payments(request, booking_id):
    """Get booking payment history for modal display"""
    try:
        booking = Booking.objects.select_related('customer', 'property_obj', 'property_obj__property_type', 'created_by').get(id=booking_id)
        
        # Check if user has permission to view payments
        if not request.user.has_perm('properties.view_payment') and booking.created_by != request.user:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        payments = []
        try:
            payments = Payment.objects.filter(booking=booking).select_related('recorded_by').order_by('-payment_date')
        except Exception as payment_error:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Could not load payments for booking {booking_id}: {str(payment_error)}")
        
        context = {
            'booking': booking,
            'payments': payments,
        }
        return render(request, 'properties/modals/booking_payments.html', context)
        
    except Booking.DoesNotExist:
        return JsonResponse({'error': 'Booking not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading payment history for booking {booking_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error loading payment history: {str(e)}'}, status=500)


@login_required
def api_booking_invoice(request, booking_id):
    """Generate booking invoice for modal display"""
    try:
        booking = Booking.objects.select_related('customer', 'property_obj', 'property_obj__property_type', 'created_by').get(id=booking_id)
        
        # Check if user has permission to view invoices
        if not request.user.has_perm('properties.view_booking') and booking.created_by != request.user:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        context = {
            'booking': booking,
        }
        return render(request, 'properties/modals/booking_invoice.html', context)
        
    except Booking.DoesNotExist:
        return JsonResponse({'error': 'Booking not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error generating invoice for booking {booking_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error generating invoice: {str(e)}'}, status=500)


@login_required
def api_booking_invoice_download(request, booking_id):
    """Download booking invoice as PDF"""
    try:
        booking = Booking.objects.select_related('customer', 'property_obj', 'property_obj__property_type', 'created_by').get(id=booking_id)
        
        # Check if user has permission to download invoices
        if not request.user.has_perm('properties.view_booking') and booking.created_by != request.user:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        # Generate PDF using reportlab
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        
        # Create a BytesIO buffer to store the PDF
        buffer = BytesIO()
        
        # Create the PDF document
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        
        # Define custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1,  # Center alignment
            textColor=colors.darkblue
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.darkblue
        )
        
        normal_style = styles['Normal']
        
        # Build the PDF content
        story = []
        
        # Title
        story.append(Paragraph("INVOICE", title_style))
        story.append(Spacer(1, 20))
        
        # Invoice details
        invoice_data = [
            ['Invoice #:', booking.booking_reference],
            ['Date:', booking.created_at.strftime('%B %d, %Y')],
            ['Booking Reference:', booking.booking_reference],
        ]
        
        invoice_table = Table(invoice_data, colWidths=[2*inch, 4*inch])
        invoice_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(invoice_table)
        story.append(Spacer(1, 20))
        
        # Customer and Property Information
        story.append(Paragraph("Bill To:", heading_style))
        customer_data = [
            ['Name:', booking.customer.full_name],
            ['Email:', booking.customer.email],
            ['Phone:', booking.customer.phone],
        ]
        
        customer_table = Table(customer_data, colWidths=[1.5*inch, 4.5*inch])
        customer_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(customer_table)
        story.append(Spacer(1, 15))
        
        story.append(Paragraph("Property Details:", heading_style))
        property_data = [
            ['Property:', booking.property_obj.title],
            ['Address:', booking.property_obj.address],
            ['Type:', booking.property_obj.property_type.name],
        ]
        
        property_table = Table(property_data, colWidths=[1.5*inch, 4.5*inch])
        property_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        
        story.append(property_table)
        story.append(Spacer(1, 20))
        
        # Invoice items table
        story.append(Paragraph("Invoice Items:", heading_style))
        
        items_data = [
            ['Description', 'Period', 'Rate', 'Amount'],
            [
                f'Rental for {booking.number_of_guests} guest{"s" if booking.number_of_guests > 1 else ""}',
                f'{booking.check_in_date.strftime("%b %d, %Y")} - {booking.check_out_date.strftime("%b %d, %Y")}',
                f'Tsh {booking.property_obj.rent_amount:,.0f}/month',
                f'Tsh {booking.total_amount:,.0f}'
            ]
        ]
        
        items_table = Table(items_data, colWidths=[3*inch, 2*inch, 1.5*inch, 1.5*inch])
        items_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        story.append(items_table)
        story.append(Spacer(1, 20))
        
        # Totals table
        totals_data = [
            ['Total Amount:', f'Tsh {booking.total_amount:,.0f}'],
            ['Paid Amount:', f'Tsh {booking.paid_amount:,.0f}'],
            ['Balance Due:', f'Tsh {booking.remaining_amount:,.0f}'],
        ]
        
        totals_table = Table(totals_data, colWidths=[4*inch, 2*inch])
        totals_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ]))
        
        story.append(totals_table)
        story.append(Spacer(1, 30))
        
        # Payment status
        payment_status = booking.payment_status.title()
        if booking.payment_status == 'paid':
            status_color = colors.green
        elif booking.payment_status == 'partial':
            status_color = colors.orange
        else:
            status_color = colors.red
            
        story.append(Paragraph(f"Payment Status: <font color='{status_color}'>{payment_status}</font>", normal_style))
        story.append(Spacer(1, 20))
        
        # Footer
        story.append(Paragraph(f"Generated on: {timezone.now().strftime('%B %d, %Y at %H:%M:%S')}", 
                              ParagraphStyle('Footer', parent=normal_style, fontSize=8, alignment=2)))
        
        # Build PDF
        doc.build(story)
        
        # Get the PDF content
        pdf_content = buffer.getvalue()
        buffer.close()
        
        # Create HTTP response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="invoice_{booking.booking_reference}.pdf"'
        response.write(pdf_content)
        
        return response
        
    except Booking.DoesNotExist:
        return JsonResponse({'error': 'Booking not found'}, status=404)
    except ImportError:
        return JsonResponse({'error': 'PDF generation library not installed. Please install reportlab.'}, status=500)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error generating PDF invoice for booking {booking_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error generating PDF invoice: {str(e)}'}, status=500)


@login_required
def api_booking_invoice_email(request, booking_id):
    """Email booking invoice"""
    try:
        booking = Booking.objects.select_related('customer', 'property_obj', 'property_obj__property_type', 'created_by').get(id=booking_id)
        
        # Check if user has permission to email invoices
        if not request.user.has_perm('properties.view_booking') and booking.created_by != request.user:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        # For now, return success message
        # In a real implementation, you would send an email here
        customer_email = booking.customer.email if booking.customer and booking.customer.email else 'N/A'
        return JsonResponse({'success': True, 'message': f'Invoice sent to {customer_email}'})
        
    except Booking.DoesNotExist:
        return JsonResponse({'error': 'Booking not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error emailing invoice for booking {booking_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error emailing invoice: {str(e)}'}, status=500)


@login_required
def api_booking_cancel(request, booking_id):
    """Cancel booking modal"""
    try:
        booking = Booking.objects.select_related('customer', 'property_obj', 'property_obj__property_type', 'created_by').get(id=booking_id)
        
        # Check if user has permission to cancel this booking
        if not request.user.has_perm('properties.change_booking') and booking.created_by != request.user:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        if request.method == 'POST':
            try:
                # Cancel the booking
                cancellation_reason = request.POST.get('cancellation_reason', 'No reason provided')
                booking.booking_status = 'cancelled'
                booking.special_requests = f"{booking.special_requests or ''}\n\nCancellation Reason: {cancellation_reason}"
                booking.save()
                
                # If booking has a room assigned, sync room status
                # This applies to both hotel and lodge bookings
                if booking.room_number and booking.property_obj:
                    try:
                        room = Room.objects.get(
                            property_obj=booking.property_obj,
                            room_number=booking.room_number
                        )
                        # Use the sync method to properly update room status based on all bookings
                        room.sync_status_from_bookings()
                    except Room.DoesNotExist:
                        # Room might not exist (e.g., for house bookings), that's okay
                        pass
                
                return JsonResponse({'success': True, 'message': 'Booking cancelled successfully'})
            except Exception as save_error:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error cancelling booking {booking_id}: {str(save_error)}")
                return JsonResponse({'error': f'Error cancelling booking: {str(save_error)}'}, status=500)
        
        context = {
            'booking': booking,
        }
        return render(request, 'properties/modals/booking_cancel.html', context)
        
    except Booking.DoesNotExist:
        return JsonResponse({'error': 'Booking not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading booking cancel for booking {booking_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error cancelling booking: {str(e)}'}, status=500)


@login_required
def api_booking_soft_delete(request, booking_id):
    """Soft delete a booking - only allowed for cancelled or completed bookings"""
    try:
        booking = Booking.objects.select_related('customer', 'property_obj', 'property_obj__property_type', 'created_by').get(id=booking_id)
        
        # Check if user has permission (must be admin/staff)
        if not (request.user.is_staff or request.user.is_superuser):
            return JsonResponse({'error': 'Permission denied. Only admins can soft delete bookings.'}, status=403)
        
        if request.method == 'POST':
            try:
                # Only allow soft delete for cancelled or completed (checked_out) bookings
                if booking.booking_status not in ['cancelled', 'checked_out']:
                    return JsonResponse({
                        'error': f'Cannot soft delete booking with status "{booking.get_booking_status_display()}". Only cancelled or completed bookings can be soft deleted.'
                    }, status=400)
                
                # Check if already soft deleted
                if booking.is_deleted:
                    return JsonResponse({
                        'error': 'Booking is already soft deleted.'
                    }, status=400)
                
                # Soft delete the booking
                from django.utils import timezone
                booking.is_deleted = True
                booking.deleted_at = timezone.now()
                booking.deleted_by = request.user
                booking.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Booking {booking.booking_reference} has been soft deleted successfully.'
                })
            except Exception as save_error:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error soft deleting booking {booking_id}: {str(save_error)}")
                return JsonResponse({'error': f'Error soft deleting booking: {str(save_error)}'}, status=500)
        
        # GET request - return booking info for confirmation
        context = {
            'booking': booking,
        }
        return JsonResponse({
            'booking_reference': booking.booking_reference,
            'status': booking.booking_status,
            'can_delete': booking.booking_status in ['cancelled', 'checked_out'] and not booking.is_deleted
        })
        
    except Booking.DoesNotExist:
        return JsonResponse({'error': 'Booking not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading soft delete for booking {booking_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error: {str(e)}'}, status=500)


@login_required
def api_booking_restore(request, booking_id):
    """Restore a soft-deleted booking"""
    try:
        booking = Booking.objects.select_related('customer', 'property_obj', 'property_obj__property_type', 'created_by').get(id=booking_id)
        
        # Check if user has permission (must be admin/staff)
        if not (request.user.is_staff or request.user.is_superuser):
            return JsonResponse({'error': 'Permission denied. Only admins can restore bookings.'}, status=403)
        
        if request.method == 'POST':
            try:
                # Check if booking is soft deleted
                if not booking.is_deleted:
                    return JsonResponse({
                        'error': 'Booking is not soft deleted.'
                    }, status=400)
                
                # Restore the booking
                booking.is_deleted = False
                booking.deleted_at = None
                booking.deleted_by = None
                booking.save()
                
                return JsonResponse({
                    'success': True,
                    'message': f'Booking {booking.booking_reference} has been restored successfully.'
                })
            except Exception as save_error:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error restoring booking {booking_id}: {str(save_error)}")
                return JsonResponse({'error': f'Error restoring booking: {str(save_error)}'}, status=500)
        
        return JsonResponse({'error': 'Invalid request method'}, status=405)
        
    except Booking.DoesNotExist:
        return JsonResponse({'error': 'Booking not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error restoring booking {booking_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error: {str(e)}'}, status=500)


@login_required
def api_booking_update_total(request, booking_id):
    """Update booking total amount based on calculated amount"""
    try:
        booking = Booking.objects.select_related('customer', 'property_obj', 'property_obj__property_type', 'created_by').get(id=booking_id)
        
        # Check if user has permission to update this booking
        if not request.user.has_perm('properties.change_booking') and booking.created_by != request.user:
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        if request.method == 'POST':
            try:
                # Calculate and update the total amount
                old_total = booking.total_amount
                new_total = booking.calculate_and_update_total()
                
                return JsonResponse({
                    'success': True, 
                    'message': f'Booking total updated from Tsh {old_total:,.0f} to Tsh {new_total:,.0f}',
                    'old_total': float(old_total),
                    'new_total': float(new_total)
                })
            except Exception as calc_error:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error calculating/updating total for booking {booking_id}: {str(calc_error)}")
                return JsonResponse({'error': f'Error updating booking total: {str(calc_error)}'}, status=500)
        
        return JsonResponse({'error': 'Method not allowed'}, status=405)
        
    except Booking.DoesNotExist:
        return JsonResponse({'error': 'Booking not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error updating booking total for booking {booking_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error updating booking total: {str(e)}'}, status=500)


# Tenant Management API Views

@login_required
def api_tenant_profile(request, tenant_id):
    """Get tenant profile for modal display"""
    try:
        tenant = Customer.objects.select_related().prefetch_related('customer_bookings__property_obj').get(id=tenant_id)
        
        # Get latest booking for this tenant
        latest_booking = None
        try:
            latest_booking = tenant.customer_bookings.filter(
                property_obj__property_type__name__iexact='house'
            ).select_related('property_obj', 'property_obj__property_type').order_by('-created_at').first()
        except Exception as booking_error:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Could not load latest booking for tenant {tenant_id}: {str(booking_error)}")
        
        context = {
            'tenant': tenant,
            'latest_booking': latest_booking,
        }
        return render(request, 'properties/modals/tenant_profile.html', context)
        
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Tenant not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading tenant profile for tenant {tenant_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error loading tenant profile: {str(e)}'}, status=500)


@login_required
def api_tenant_edit(request, tenant_id):
    """Get tenant edit form for modal display"""
    try:
        tenant = Customer.objects.get(id=tenant_id)
        
        if request.method == 'POST':
            # Handle form submission
            try:
                tenant.first_name = request.POST.get('first_name', tenant.first_name)
                tenant.last_name = request.POST.get('last_name', tenant.last_name)
                tenant.email = request.POST.get('email', tenant.email)
                tenant.phone = request.POST.get('phone', tenant.phone)
                tenant.save()
                
                return JsonResponse({'success': True, 'message': 'Tenant updated successfully'})
            except Exception as save_error:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Error saving tenant {tenant_id}: {str(save_error)}")
                return JsonResponse({'error': f'Error updating tenant: {str(save_error)}'}, status=500)
        
        context = {
            'tenant': tenant,
        }
        return render(request, 'properties/modals/tenant_edit.html', context)
        
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Tenant not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading tenant edit form for tenant {tenant_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error loading tenant edit form: {str(e)}'}, status=500)


@login_required
def api_property_details(request, property_id):
    """Get property details for modal display"""
    try:
        property_obj = Property.objects.select_related('property_type', 'owner').get(id=property_id)
        
        # Get current tenant for this property
        current_booking = None
        try:
            current_booking = Booking.objects.filter(
                property_obj=property_obj,
                booking_status__in=['confirmed', 'checked_in']
            ).select_related('customer').first()
        except Exception as booking_error:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Could not load current booking for property {property_id}: {str(booking_error)}")
        
        context = {
            'property': property_obj,
            'current_booking': current_booking,
        }
        return render(request, 'properties/modals/property_details.html', context)
        
    except Property.DoesNotExist:
        return JsonResponse({'error': 'Property not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading property details for property {property_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error loading property details: {str(e)}'}, status=500)


@login_required
def api_tenant_lease_history(request, tenant_id):
    """Get tenant lease history for modal display"""
    try:
        tenant = Customer.objects.get(id=tenant_id)
        
        # Get all house bookings for this tenant
        lease_history = []
        try:
            lease_history = Booking.objects.filter(
                customer=tenant,
                property_obj__property_type__name__iexact='house'
            ).select_related('property_obj', 'property_obj__property_type').order_by('-created_at')
        except Exception as history_error:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Could not load lease history for tenant {tenant_id}: {str(history_error)}")
        
        context = {
            'tenant': tenant,
            'lease_history': lease_history,
        }
        return render(request, 'properties/modals/tenant_lease_history.html', context)
        
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Tenant not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading lease history for tenant {tenant_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error loading lease history: {str(e)}'}, status=500)


@login_required
def api_tenant_payment_history(request, tenant_id):
    """Get tenant payment history for modal display"""
    try:
        tenant = Customer.objects.get(id=tenant_id)
        
        # Get all payments for this tenant's house bookings
        payments = []
        try:
            payments = Payment.objects.filter(
                booking__customer=tenant,
                booking__property_obj__property_type__name__iexact='house'
            ).select_related('booking', 'booking__property_obj', 'booking__property_obj__property_type').order_by('-payment_date')
        except Exception as payment_error:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Could not load payment history for tenant {tenant_id}: {str(payment_error)}")
        
        context = {
            'tenant': tenant,
            'payments': payments,
        }
        return render(request, 'properties/modals/tenant_payment_history.html', context)
        
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Tenant not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading payment history for tenant {tenant_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'error': f'Error loading payment history: {str(e)}'}, status=500)


@login_required
def api_tenant_make_vip(request, tenant_id):
    """Make tenant VIP modal"""
    try:
        tenant = Customer.objects.get(id=tenant_id)
        
        if request.method == 'POST':
            # Get the latest booking for this tenant
            latest_booking = tenant.customer_bookings.filter(
                property_obj__property_type__name__iexact='house'
            ).order_by('-created_at').first()
            
            if latest_booking:
                # Increase the rent amount to make them VIP (>= 1,000,000)
                if latest_booking.total_amount < 1000000:
                    latest_booking.total_amount = 1000000
                    latest_booking.save()
                    return JsonResponse({'success': True, 'message': 'Tenant made VIP successfully'})
                else:
                    return JsonResponse({'success': False, 'message': 'Tenant is already VIP'})
            else:
                return JsonResponse({'success': False, 'message': 'No booking found for this tenant'})
        
        context = {
            'tenant': tenant,
        }
        return render(request, 'properties/modals/tenant_make_vip.html', context)
        
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Tenant not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error making tenant VIP: {str(e)}'}, status=500)


@login_required
def api_tenant_remove_vip(request, tenant_id):
    """Remove tenant VIP status modal"""
    try:
        tenant = Customer.objects.get(id=tenant_id)
        
        if request.method == 'POST':
            # Get the latest booking for this tenant
            latest_booking = tenant.customer_bookings.filter(
                property_obj__property_type__name__iexact='house'
            ).order_by('-created_at').first()
            
            if latest_booking:
                # Decrease the rent amount to remove VIP status (< 1,000,000)
                if latest_booking.total_amount >= 1000000:
                    latest_booking.total_amount = 500000  # Set to premium level
                    latest_booking.save()
                    return JsonResponse({'success': True, 'message': 'VIP status removed successfully'})
                else:
                    return JsonResponse({'success': False, 'message': 'Tenant is not VIP'})
            else:
                return JsonResponse({'success': False, 'message': 'No booking found for this tenant'})
        
        context = {
            'tenant': tenant,
        }
        return render(request, 'properties/modals/tenant_remove_vip.html', context)
        
    except Customer.DoesNotExist:
        return JsonResponse({'error': 'Tenant not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'Error removing VIP status: {str(e)}'}, status=500)


@login_required
def api_record_payment(request):
    """API endpoint to record a new payment"""
    from decimal import Decimal
    if request.method == 'POST':
        try:
            # Get form data
            booking_id = request.POST.get('booking_id')
            amount = request.POST.get('amount')
            payment_method = request.POST.get('payment_method')
            payment_type = request.POST.get('payment_type')
            payment_date = request.POST.get('payment_date')
            transaction_reference = request.POST.get('transaction_reference', '')
            status = request.POST.get('status', 'active')
            notes = request.POST.get('notes', '')
            
            # Validate required fields
            if not all([booking_id, amount, payment_method, payment_type, payment_date]):
                return JsonResponse({'success': False, 'message': 'Missing required fields'})
            
            # Get the booking
            try:
                booking = Booking.objects.get(id=booking_id)
            except Booking.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Booking not found'})
            
            # Convert amount to Decimal for proper handling
            from decimal import Decimal
            payment_amount = Decimal(str(amount))
            
            # Calculate remaining balance (use calculated_total_amount if available)
            calculated_total = booking.calculated_total_amount if booking.calculated_total_amount and booking.calculated_total_amount > 0 else (booking.total_amount or 0)
            remaining_balance = max(Decimal('0'), Decimal(str(calculated_total)) - booking.paid_amount)
            
            # For deposits and partial payments, allow any amount up to remaining balance
            # For full payments, the amount should equal the remaining balance
            # For refunds, allow negative amounts
            if payment_type == 'refund':
                # Allow refunds (negative amounts or positive amounts to be deducted)
                pass  # No validation for refunds
            elif payment_type == 'full':
                # Full payment must match remaining balance (with small tolerance)
                if abs(payment_amount - remaining_balance) > Decimal('0.01'):
                    return JsonResponse({
                        'success': False, 
                        'message': f'Full payment amount ({payment_amount}) must equal remaining balance ({remaining_balance})'
                    })
            else:
                # For deposit and partial, check that amount doesn't exceed remaining balance
                if payment_amount > remaining_balance:
                    return JsonResponse({
                        'success': False,
                        'message': f'Payment amount ({payment_amount}) exceeds remaining balance ({remaining_balance})'
                    })
            
            # Prevent negative payments for non-refunds
            if payment_type != 'refund' and payment_amount < 0:
                return JsonResponse({
                    'success': False,
                    'message': 'Payment amount cannot be negative for non-refund payments'
                })
            
            # Create the payment
            payment = Payment.objects.create(
                booking=booking,
                amount=payment_amount,
                payment_method=payment_method,
                payment_type=payment_type,
                payment_date=payment_date,
                transaction_reference=transaction_reference,
                status=status,
                notes=notes,
                recorded_by=request.user
            )
            
            # Recalculate booking's paid_amount from all payments (not just add/subtract)
            from django.db.models import Sum
            actual_paid = Payment.objects.filter(
                booking=booking
            ).exclude(
                payment_type='refund'
            ).aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0')
            
            total_refunded = Payment.objects.filter(
                booking=booking,
                payment_type='refund'
            ).aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0')
            
            # Net paid amount = payments - refunds
            net_paid_amount = max(Decimal('0'), actual_paid - total_refunded)
            
            # Update booking's paid_amount
            booking.paid_amount = net_paid_amount
            
            # Update payment status based on actual amounts
            booking.update_payment_status()
            booking.save()
            
            return JsonResponse({
                'success': True, 
                'message': f'Payment of Tsh {float(amount):,.0f} recorded successfully',
                'payment_id': payment.id
            })
            
        except Exception as e:
            return JsonResponse({'success': False, 'message': f'Error recording payment: {str(e)}'})
    
    return JsonResponse({'success': False, 'message': 'Method not allowed'}, status=405)


# Venue Management API Endpoints
@login_required
def api_property_availability(request, property_id):
    """
    AJAX endpoint to get property availability information
    Returns booked dates and next available date for a property
    """
    try:
        from datetime import datetime, timedelta
        from django.utils import timezone
        
        property_obj = get_object_or_404(Property, pk=property_id)
        
        # Get all active bookings (confirmed, checked_in, pending)
        active_bookings = property_obj.property_bookings.filter(
            booking_status__in=['confirmed', 'checked_in', 'pending']
        ).order_by('check_in_date')
        
        # Get active leases
        from documents.models import Lease
        active_leases = Lease.objects.filter(
            property_ref=property_obj,
            status='active'
        ).order_by('start_date')
        
        # Prepare booked dates from bookings
        booked_dates = []
        current_bookings = []
        
        for booking in active_bookings:
            booked_dates.append({
                'check_in': booking.check_in_date.strftime('%Y-%m-%d'),
                'check_out': booking.check_out_date.strftime('%Y-%m-%d'),
                'status': booking.booking_status,
                'booking_reference': booking.booking_reference
            })
            current_bookings.append({
                'id': booking.id,
                'check_in': booking.check_in_date.strftime('%Y-%m-%d'),
                'check_out': booking.check_out_date.strftime('%Y-%m-%d'),
                'status': booking.booking_status,
                'booking_reference': booking.booking_reference,
                'customer': booking.customer.full_name if booking.customer else 'N/A'
            })
        
        # Add lease dates
        for lease in active_leases:
            booked_dates.append({
                'check_in': lease.start_date.strftime('%Y-%m-%d'),
                'check_out': lease.end_date.strftime('%Y-%m-%d'),
                'status': 'lease',
                'booking_reference': f"Lease-{lease.id}"
            })
        
        # Calculate next available date
        today = timezone.now().date()
        next_available_date = None
        
        # Sort booked dates by check_in
        sorted_bookings = sorted(booked_dates, key=lambda x: x['check_in'])
        
        # Find first gap or date after last booking
        if sorted_bookings:
            last_checkout = max(
                datetime.strptime(b['check_out'], '%Y-%m-%d').date() 
                for b in sorted_bookings
            )
            # Next available is day after last checkout
            if last_checkout >= today:
                next_available_date = (last_checkout + timedelta(days=1)).strftime('%Y-%m-%d')
            else:
                next_available_date = today.strftime('%Y-%m-%d')
        else:
            # No bookings, available from today
            next_available_date = today.strftime('%Y-%m-%d')
        
        # Check if property is currently available
        is_available = property_obj.is_available_for_booking()
        
        return JsonResponse({
            'success': True,
            'property_id': property_obj.id,
            'property_title': property_obj.title,
            'is_available': is_available,
            'property_status': property_obj.status,
            'booked_dates': booked_dates,
            'current_bookings': current_bookings,
            'next_available_date': next_available_date,
            'has_active_bookings': len(current_bookings) > 0,
            'has_active_leases': len(active_leases) > 0
        })
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error fetching property availability: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'Failed to fetch availability: {str(e)}'
        }, status=500)


@login_required
def api_venue_availability(request):
    """API endpoint to get venue availability"""
    auto_complete_venue_bookings(request.user)
    from datetime import datetime, timedelta
    from collections import defaultdict
    
    try:
        date_format = '%Y-%m-%d'
        today = timezone.now().date()
        
        # Parse date range
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')
        
        if start_date_str:
            start_date = datetime.strptime(start_date_str, date_format).date()
        else:
            start_date = today
        
        if end_date_str:
            end_date = datetime.strptime(end_date_str, date_format).date()
        else:
            end_date = start_date + timedelta(days=6)
        
        if end_date < start_date:
            start_date, end_date = end_date, start_date
        
        # Limit range to 31 days to prevent heavy payloads
        max_range = 31
        if (end_date - start_date).days > max_range:
            end_date = start_date + timedelta(days=max_range)
        
        # Venue filtering with multi-tenancy
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            venue_queryset = Property.objects.filter(property_type__name__iexact='venue', owner=request.user)
        else:
            venue_queryset = Property.objects.filter(property_type__name__iexact='venue')
        
        venue_id = request.GET.get('venue_id')
        if venue_id:
            venue_queryset = venue_queryset.filter(id=venue_id)
        
        venue_queryset = venue_queryset.order_by('title')
        
        venue_ids = list(venue_queryset.values_list('id', flat=True))
        if not venue_ids:
            return JsonResponse({
                'success': True,
                'dates': [],
                'venues': [],
                'summary': {
                    'available_today': 0,
                    'booked_today': 0,
                    'blocked_today': 0,
                    'total_capacity': 0,
                },
                'today_schedule': [],
                'start_date': start_date.strftime(date_format),
                'end_date': end_date.strftime(date_format),
                'today': today.strftime(date_format),
            })
        
        # Fetch bookings overlapping the date range
        bookings = Booking.objects.filter(
            property_obj_id__in=venue_ids,
            booking_status__in=['pending', 'confirmed', 'checked_in'],
            check_in_date__lte=end_date,
            check_out_date__gte=start_date,
        ).select_related('property_obj', 'customer').order_by('check_in_date')
        
        day_map = defaultdict(list)
        for booking in bookings:
            booking_start = max(booking.check_in_date, start_date)
            booking_end = min(booking.check_out_date, end_date)
            day_cursor = booking_start
            while day_cursor <= booking_end:
                day_map[(booking.property_obj_id, day_cursor)].append(booking)
                day_cursor += timedelta(days=1)
        
        # Prepare date metadata
        date_cursor = start_date
        dates_meta = []
        while date_cursor <= end_date:
            dates_meta.append({
                'date': date_cursor.strftime(date_format),
                'label': date_cursor.strftime('%a %d %b'),
            })
            date_cursor += timedelta(days=1)
        
        # Build venue/day payload
        venues_payload = []
        today_schedule = []
        today_in_range = start_date <= today <= end_date
        available_today = 0
        booked_today = 0
        total_capacity = 0
        
        for venue in venue_queryset:
            total_capacity += venue.capacity or 0
            venue_days = []
            for date_meta in dates_meta:
                day_date = datetime.strptime(date_meta['date'], date_format).date()
                events = []
                for booking in day_map.get((venue.id, day_date), []):
                    event_name = booking.special_requests or booking.booking_reference or venue.title
                    events.append({
                        'booking_id': booking.id,
                        'reference': booking.booking_reference,
                        'event_name': event_name,
                        'status': booking.booking_status,
                        'customer': getattr(booking.customer, 'full_name', str(booking.customer)),
                        'amount': float(booking.total_amount),
                    })
                    
                    if today_in_range and day_date == today:
                        today_schedule.append({
                            'time': 'All day',
                            'venue': venue.title,
                            'event': event_name,
                            'status': booking.booking_status,
                        })
                
                venue_days.append({
                    'date': date_meta['date'],
                    'events': events,
                })
            
            if today_in_range:
                today_events = day_map.get((venue.id, today), [])
                if today_events:
                    booked_today += 1
                else:
                    available_today += 1
            
            venues_payload.append({
                'id': venue.id,
                'name': venue.title,
                'capacity': venue.capacity or 0,
                'days': venue_days,
            })
        
        response = {
            'success': True,
            'start_date': start_date.strftime(date_format),
            'end_date': end_date.strftime(date_format),
            'today': today.strftime(date_format),
            'dates': dates_meta,
            'venues': venues_payload,
            'summary': {
                'available_today': available_today,
                'booked_today': booked_today,
                'blocked_today': 0,  # Placeholder until blocking logic is implemented
                'total_capacity': total_capacity,
            },
            'today_schedule': today_schedule,
        }
        return JsonResponse(response)
    
    except ValueError:
        return JsonResponse({'success': False, 'message': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def api_venue_booking_status(request, booking_id):
    """API endpoint to update venue booking status"""
    try:
        booking = Booking.objects.get(id=booking_id, property_obj__property_type__name__iexact='venue')
        
        if request.method == 'POST':
            new_status = request.POST.get('status')
            action = request.POST.get('action')
            
            if action == 'confirm':
                booking.booking_status = 'confirmed'
                booking.save()
                return JsonResponse({'success': True, 'message': 'Booking confirmed successfully'})
            
            elif action == 'start':
                booking.booking_status = 'checked_in'
                booking.save()
                return JsonResponse({'success': True, 'message': 'Event started successfully'})
            
            elif action == 'end':
                booking.booking_status = 'checked_out'
                booking.save()
                return JsonResponse({'success': True, 'message': 'Event ended successfully'})
            
            elif action == 'cancel':
                booking.booking_status = 'cancelled'
                booking.save()
                
                # If booking has a room assigned, sync room status
                # This applies to both hotel and lodge bookings
                if booking.room_number and booking.property_obj:
                    try:
                        room = Room.objects.get(
                            property_obj=booking.property_obj,
                            room_number=booking.room_number
                        )
                        # Use the sync method to properly update room status based on all bookings
                        room.sync_status_from_bookings()
                    except Room.DoesNotExist:
                        # Room might not exist (e.g., for venue bookings), that's okay
                        pass
                
                return JsonResponse({'success': True, 'message': 'Booking cancelled successfully'})
            
            elif new_status:
                booking.booking_status = new_status
                booking.save()
                return JsonResponse({'success': True, 'message': 'Status updated successfully'})
        
        return JsonResponse({'success': False, 'message': 'Invalid action'})
        
    except Booking.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Booking not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def api_venue_details(request, venue_id):
    """API endpoint to get venue details (price, capacity)"""
    try:
        # MULTI-TENANCY: Ensure owner can only access their own venues
        if request.user.is_authenticated and not (request.user.is_staff or request.user.is_superuser):
            venue = Property.objects.get(id=venue_id, property_type__name__iexact='venue', owner=request.user)
        else:
            venue = Property.objects.get(id=venue_id, property_type__name__iexact='venue')
        
        return JsonResponse({
            'success': True,
            'venue': {
                'id': venue.id,
                'title': venue.title,
                'capacity': venue.capacity or 0,
                'rent_amount': float(venue.rent_amount) if venue.rent_amount else 0,
                'rent_period': venue.rent_period or 'day',
            }
        })
    except Property.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Venue not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
def api_venue_capacity_check(request):
    """API endpoint to check venue capacity availability"""
    from datetime import datetime
    try:
        venue_id = request.GET.get('venue_id')
        event_date = request.GET.get('event_date')
        guest_count = int(request.GET.get('guest_count', 0))
        
        if not venue_id or not event_date:
            return JsonResponse({'success': False, 'message': 'Missing required parameters'})
        
        venue = Property.objects.get(id=venue_id, property_type__name__iexact='venue')
        event_date = datetime.strptime(event_date, '%Y-%m-%d').date()
        
        # Check if venue has enough capacity
        if venue.capacity and guest_count > venue.capacity:
            return JsonResponse({
                'success': False,
                'available': False,
                'message': f'Guest count ({guest_count}) exceeds venue capacity ({venue.capacity})'
            })
        
        # Use the new availability method
        if not venue.is_available_for_booking(event_date, event_date):
            return JsonResponse({
                'success': False,
                'available': False,
                'message': 'Venue is not available for the selected date',
            })
        
        # Check for conflicting bookings (for detailed info)
        conflicting_bookings = Booking.objects.filter(
            property_obj=venue,
            check_in_date=event_date,
            booking_status__in=['confirmed', 'checked_in', 'pending']
        ).exclude(id=request.GET.get('exclude_booking_id'))
        
        if conflicting_bookings.exists():
            return JsonResponse({
                'success': False,
                'available': False,
                'message': 'Venue is already booked for this date',
                'conflicting_bookings': [
                    {
                        'id': booking.id,
                        'event_name': booking.special_requests or 'Event',
                        'time': f"{booking.check_in_date}"
                    } for booking in conflicting_bookings
                ]
            })
        
        return JsonResponse({
            'success': True,
            'available': True,
            'message': 'Venue is available for booking',
            'venue': {
                'id': venue.id,
                'title': venue.title,
                'capacity': venue.capacity
            }
        })
        
    except Property.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Venue not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def api_venue_analytics(request):
    """API endpoint to get venue analytics data"""
    try:
        from datetime import datetime, timedelta
        from django.db.models import Count, Sum, Avg
        
        # Get parameters
        venue_id = request.GET.get('venue_id')
        period = request.GET.get('period', 'month')  # week, month, quarter, year
        
        # Calculate date range based on period
        end_date = datetime.now().date()
        if period == 'week':
            start_date = end_date - timedelta(days=7)
        elif period == 'month':
            start_date = end_date - timedelta(days=30)
        elif period == 'quarter':
            start_date = end_date - timedelta(days=90)
        elif period == 'year':
            start_date = end_date - timedelta(days=365)
        else:
            start_date = end_date - timedelta(days=30)
        
        # Get venues
        venues_query = Property.objects.filter(property_type__name__iexact='venue')
        if venue_id:
            venues_query = venues_query.filter(id=venue_id)
        
        analytics_data = []
        for venue in venues_query:
            # Get bookings in the period
            bookings = Booking.objects.filter(
                property_obj=venue,
                check_in_date__range=[start_date, end_date]
            )
            
            # Calculate metrics
            total_bookings = bookings.count()
            confirmed_bookings = bookings.filter(booking_status='confirmed').count()
            completed_bookings = bookings.filter(booking_status='checked_out').count()
            cancelled_bookings = bookings.filter(booking_status='cancelled').count()
            
            total_revenue = bookings.filter(payment_status='paid').aggregate(
                total=Sum('total_amount')
            )['total'] or 0
            
            avg_guests = bookings.aggregate(
                avg=Avg('number_of_guests')
            )['avg'] or 0
            
            occupancy_rate = 0
            if venue.capacity and total_bookings > 0:
                total_guests = sum(booking.number_of_guests or 0 for booking in bookings.filter(booking_status__in=['confirmed', 'checked_in', 'checked_out']))
                occupancy_rate = min(100, (total_guests / (venue.capacity * total_bookings)) * 100)
            
            venue_analytics = {
                'venue_id': venue.id,
                'venue_title': venue.title,
                'period': period,
                'date_range': {
                    'start': start_date.strftime('%Y-%m-%d'),
                    'end': end_date.strftime('%Y-%m-%d')
                },
                'metrics': {
                    'total_bookings': total_bookings,
                    'confirmed_bookings': confirmed_bookings,
                    'completed_bookings': completed_bookings,
                    'cancelled_bookings': cancelled_bookings,
                    'total_revenue': float(total_revenue),
                    'average_guests': round(avg_guests, 1),
                    'occupancy_rate': round(occupancy_rate, 1)
                }
            }
            analytics_data.append(venue_analytics)
        
        return JsonResponse({
            'success': True,
            'analytics': analytics_data
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# Payment Action API Endpoints
@login_required
def api_payment_view_details(request, payment_id):
    """API endpoint to view payment details - handles both booking payments and visit payments"""
    try:
        payment = None
        payment_type = None
        
        # Try booking payment first (properties.models.Payment)
        try:
            payment = Payment.objects.select_related(
                'booking', 
                'booking__customer', 
                'booking__property_obj',
                'booking__property_obj__property_type',
                'recorded_by'
            ).get(id=payment_id)
            payment_type = 'booking'
        except Payment.DoesNotExist:
            # Try unified payment (payments.models.Payment) - could be visit payment
            from payments.models import Payment as UnifiedPayment
            from .models import PropertyVisitPayment
            try:
                payment = UnifiedPayment.objects.select_related('tenant', 'recorded_by', 'provider').get(id=payment_id)
                # Check if it's a visit payment
                visit_payment = PropertyVisitPayment.objects.filter(payment=payment).select_related('property', 'property__property_type').first()
                if visit_payment:
                    payment_type = 'visit'
                    payment.visit_payment = visit_payment
                    payment.property = visit_payment.property
                else:
                    payment_type = 'unified'
            except UnifiedPayment.DoesNotExist:
                pass
        
        if not payment:
            return JsonResponse({'success': False, 'message': 'Payment not found'}, status=404)
        
        # Add visit payment expiration info to context if it's a visit payment
        context = {'payment': payment, 'payment_type': payment_type}
        if payment_type == 'visit' and hasattr(payment, 'visit_payment'):
            # Add expiration info to context for template access
            visit_payment = payment.visit_payment
            expires_at = visit_payment.expires_at()
            context['visit_payment'] = visit_payment
            context['visit_payment_expires_at'] = expires_at
            context['visit_payment_is_expired'] = visit_payment.is_expired()
            context['visit_payment_is_active'] = visit_payment.is_active()
        
        return render(request, 'properties/modals/payment_view_details.html', context)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading payment details for payment {payment_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'message': f'Error loading payment details: {str(e)}'}, status=500)


@login_required
def api_payment_generate_receipt(request, payment_id):
    """API endpoint to generate payment receipt - handles both booking payments and visit payments"""
    try:
        # Try booking payment first (properties.models.Payment)
        try:
            payment = Payment.objects.select_related('booking', 'booking__customer', 'booking__property_obj').get(id=payment_id)
            payment_type = 'booking'
        except Payment.DoesNotExist:
            # Try unified payment (payments.models.Payment) - could be visit payment
            from payments.models import Payment as UnifiedPayment
            from .models import PropertyVisitPayment
            payment = UnifiedPayment.objects.select_related('tenant', 'recorded_by').get(id=payment_id)
            # Check if it's a visit payment
            visit_payment = PropertyVisitPayment.objects.filter(payment=payment).first()
            if visit_payment:
                payment_type = 'visit'
                payment.visit_payment = visit_payment
                payment.property = visit_payment.property
            else:
                payment_type = 'unified'
        
        context = {'payment': payment, 'payment_type': payment_type}
        return render(request, 'properties/modals/payment_generate_receipt.html', context)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error generating receipt for payment {payment_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'message': f'Error generating receipt: {str(e)}'}, status=500)


@login_required
def api_payment_download_receipt(request, payment_id):
    """API endpoint to download payment receipt as PDF"""
    try:
        payment = Payment.objects.select_related(
            'booking', 'booking__customer', 'booking__property_obj', 'booking__property_obj__property_type', 'recorded_by'
        ).get(id=payment_id)
        
        # Generate PDF receipt using reportlab
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from io import BytesIO
        
        # Create a BytesIO buffer to store the PDF
        buffer = BytesIO()
        
        # Create the PDF document
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        
        # Define custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1F4788'),
            spaceAfter=30,
            alignment=1
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10
        )
        
        # Build the content for the PDF
        story = []
        
        # Title
        story.append(Paragraph("PAYMENT RECEIPT", title_style))
        story.append(Spacer(1, 20))
        
        # Receipt number and date
        story.append(Paragraph(f"Receipt #: {payment.id}", normal_style))
        story.append(Paragraph(f"Date: {payment.payment_date.strftime('%B %d, %Y at %I:%M %p')}", normal_style))
        story.append(Spacer(1, 20))
        
        # Customer information
        story.append(Paragraph("<b>Customer Information</b>", normal_style))
        customer_data = [
            ['Name:', payment.booking.customer.full_name],
            ['Email:', payment.booking.customer.email or 'N/A'],
            ['Phone:', payment.booking.customer.phone or 'N/A'],
        ]
        customer_table = Table(customer_data, colWidths=[2*inch, 4*inch])
        customer_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(customer_table)
        story.append(Spacer(1, 20))
        
        # Payment details
        story.append(Paragraph("<b>Payment Details</b>", normal_style))
        payment_data = [
            ['Amount Paid:', f"Tsh {payment.amount:,.0f}"],
            ['Payment Method:', payment.get_payment_method_display()],
            ['Payment Type:', payment.get_payment_type_display()],
            ['Transaction Reference:', payment.transaction_reference or 'N/A'],
        ]
        payment_table = Table(payment_data, colWidths=[2*inch, 4*inch])
        payment_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(payment_table)
        story.append(Spacer(1, 20))
        
        # Property information
        story.append(Paragraph("<b>Property Information</b>", normal_style))
        property_data = [
            ['Property:', payment.booking.property_obj.title],
            ['Address:', payment.booking.property_obj.address],
            ['Type:', payment.booking.property_obj.property_type.name],
            ['Booking Reference:', payment.booking.booking_reference],
            ['Check-in:', payment.booking.check_in_date.strftime('%B %d, %Y')],
            ['Check-out:', payment.booking.check_out_date.strftime('%B %d, %Y')],
        ]
        property_table = Table(property_data, colWidths=[2*inch, 4*inch])
        property_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ]))
        story.append(property_table)
        story.append(Spacer(1, 30))
        
        # Footer
        story.append(Paragraph(f"Generated on: {timezone.now().strftime('%B %d, %Y at %H:%M:%S')}", 
                              ParagraphStyle('Footer', parent=normal_style, fontSize=8, alignment=2)))
        
        # Build PDF
        doc.build(story)
        
        # Get the PDF content
        pdf_content = buffer.getvalue()
        buffer.close()
        
        # Create HTTP response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="payment_receipt_{payment.id}.pdf"'
        response.write(pdf_content)
        
        return response
        
    except Payment.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Payment not found'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def api_payment_edit(request, payment_id):
    """API endpoint to edit payment - handles both booking payments and visit payments"""
    from decimal import Decimal
    try:
        # Try booking payment first (properties.models.Payment)
        try:
            payment = Payment.objects.select_related('booking', 'booking__customer', 'booking__property_obj').get(id=payment_id)
            payment_type = 'booking'
            is_visit_payment = False
        except Payment.DoesNotExist:
            # Try unified payment (payments.models.Payment) - could be visit payment
            from payments.models import Payment as UnifiedPayment
            from .models import PropertyVisitPayment
            payment = UnifiedPayment.objects.select_related('tenant', 'recorded_by').get(id=payment_id)
            # Check if it's a visit payment
            visit_payment = PropertyVisitPayment.objects.filter(payment=payment).first()
            if visit_payment:
                payment_type = 'visit'
                is_visit_payment = True
                payment.visit_payment = visit_payment
                payment.property = visit_payment.property
            else:
                payment_type = 'unified'
                is_visit_payment = False
        
        if request.method == 'POST':
            # Update payment based on type
            if payment_type == 'booking':
                payment.amount = Decimal(str(request.POST.get('amount', payment.amount)))
                payment.payment_method = request.POST.get('payment_method', payment.payment_method)
                payment.payment_type = request.POST.get('payment_type', payment.payment_type)
                payment.transaction_reference = request.POST.get('transaction_reference', payment.transaction_reference)
                payment.payment_date = request.POST.get('payment_date', payment.payment_date)
                payment.status = request.POST.get('status', payment.status)
                payment.notes = request.POST.get('notes', payment.notes)
            else:  # unified payment (visit or other)
                payment.amount = Decimal(str(request.POST.get('amount', payment.amount)))
                payment.payment_method = request.POST.get('payment_method', payment.payment_method)
                payment.status = request.POST.get('status', payment.status)
                payment.notes = request.POST.get('notes', payment.notes)
                if hasattr(payment, 'paid_date'):
                    from django.utils.dateparse import parse_date
                    paid_date_str = request.POST.get('payment_date') or request.POST.get('paid_date')
                    if paid_date_str:
                        payment.paid_date = parse_date(paid_date_str)
            
            # Validate payment amount doesn't exceed remaining balance (for booking payments)
            if payment_type == 'booking' and payment.booking:
                from django.db.models import Sum
                from decimal import Decimal
                
                # Calculate current paid amount (excluding the payment being edited)
                old_payment_amount = Payment.objects.filter(id=payment_id).values_list('amount', flat=True).first() or Decimal('0')
                
                # Recalculate total paid from all payments (exclude refunds and current payment)
                actual_paid = Payment.objects.filter(
                    booking=payment.booking
                ).exclude(
                    payment_type='refund'
                ).exclude(
                    id=payment_id  # Exclude current payment being edited
                ).aggregate(
                    total=Sum('amount')
                )['total'] or Decimal('0')
                
                # Get total refunded
                total_refunded = Payment.objects.filter(
                    booking=payment.booking,
                    payment_type='refund'
                ).aggregate(
                    total=Sum('amount')
                )['total'] or Decimal('0')
                
                # Calculate remaining balance before this payment
                calculated_total = payment.booking.calculated_total_amount if payment.booking.calculated_total_amount and payment.booking.calculated_total_amount > 0 else (payment.booking.total_amount or 0)
                current_paid = max(Decimal('0'), actual_paid - total_refunded)
                remaining_balance = max(Decimal('0'), Decimal(str(calculated_total)) - current_paid)
                
                # Validate new payment amount doesn't exceed remaining balance
                if payment.amount > remaining_balance:
                    return JsonResponse({
                        'success': False,
                        'message': f'Payment amount (Tsh{payment.amount:,.2f}) exceeds remaining balance (Tsh{remaining_balance:,.2f}). Maximum allowed: Tsh{remaining_balance:,.2f}'
                    }, status=400)
            
            payment.save()
            
            # If this is a booking payment, recalculate booking's paid_amount and remaining amount
            if payment_type == 'booking' and payment.booking:
                from django.db.models import Sum
                from decimal import Decimal
                
                # Recalculate total paid from all payments (exclude refunds)
                actual_paid = Payment.objects.filter(
                    booking=payment.booking
                ).exclude(
                    payment_type='refund'
                ).aggregate(
                    total=Sum('amount')
                )['total'] or Decimal('0')
                
                # Get total refunded
                total_refunded = Payment.objects.filter(
                    booking=payment.booking,
                    payment_type='refund'
                ).aggregate(
                    total=Sum('amount')
                )['total'] or Decimal('0')
                
                # Net paid amount = payments - refunds
                net_paid_amount = max(Decimal('0'), actual_paid - total_refunded)
                
                # Update booking's paid_amount
                payment.booking.paid_amount = net_paid_amount
                
                # Update payment status based on actual amounts
                payment.booking.update_payment_status()
                payment.booking.save()
            
            return JsonResponse({'success': True, 'message': 'Payment updated successfully'})
        
        # GET request - show edit form
        if payment_type == 'booking':
            context = {
                'payment': payment,
                'payment_type': payment_type,
                'payment_method_choices': Payment.PAYMENT_METHOD_CHOICES,
                'payment_type_choices': Payment.PAYMENT_TYPE_CHOICES,
                'payment_status_choices': Payment.PAYMENT_STATUS_CHOICES,
            }
        else:  # unified payment
            from payments.models import Payment as UnifiedPayment
            context = {
                'payment': payment,
                'payment_type': payment_type,
                'is_visit_payment': is_visit_payment,
                'payment_method_choices': UnifiedPayment.PAYMENT_METHOD_CHOICES,
                'payment_status_choices': UnifiedPayment.STATUS_CHOICES,
            }
        
        return render(request, 'properties/modals/payment_edit.html', context)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading payment edit for payment {payment_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'message': f'Error loading payment: {str(e)}'}, status=500)


@login_required
def api_payment_mark_paid(request, payment_id):
    """API endpoint to mark payment as paid - handles both booking payments and visit payments"""
    try:
        # Try booking payment first (properties.models.Payment)
        try:
            payment = Payment.objects.select_related('booking').get(id=payment_id)
            payment_type = 'booking'
        except Payment.DoesNotExist:
            # Try unified payment (payments.models.Payment) - could be visit payment
            from payments.models import Payment as UnifiedPayment
            from .models import PropertyVisitPayment
            payment = UnifiedPayment.objects.select_related('tenant', 'recorded_by').get(id=payment_id)
            # Check if it's a visit payment
            visit_payment = PropertyVisitPayment.objects.filter(payment=payment).first()
            if visit_payment:
                payment_type = 'visit'
                payment.visit_payment = visit_payment
                payment.property = visit_payment.property
            else:
                payment_type = 'unified'
        
        if request.method == 'POST':
            # Update payment status based on type
            if payment_type == 'booking':
                payment.status = 'completed'
                payment.save()
                # Update booking payment status
                booking = payment.booking
                booking.update_payment_status()
            else:  # unified payment (visit or other)
                from django.utils import timezone
                payment.status = 'completed'  # Payment model uses 'completed', not 'successful'
                if hasattr(payment, 'paid_date'):
                    payment.paid_date = timezone.now().date()
                payment.save()
                
                # If it's a visit payment, update the visit payment status too
                if payment_type == 'visit' and hasattr(payment, 'visit_payment'):
                    payment.visit_payment.status = 'completed'
                    payment.visit_payment.paid_at = timezone.now()
                    payment.visit_payment.save()
            
            return JsonResponse({'success': True, 'message': 'Payment marked as paid successfully'})
        
        context = {'payment': payment, 'payment_type': payment_type}
        return render(request, 'properties/modals/payment_mark_paid.html', context)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading payment mark paid for payment {payment_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'message': f'Error loading payment: {str(e)}'}, status=500)


@login_required
def api_payment_booking_details(request, payment_id):
    """API endpoint to view booking details for a payment - only works for booking payments"""
    try:
        payment = Payment.objects.select_related(
            'booking', 
            'booking__customer', 
            'booking__property_obj',
            'booking__property_obj__property_type',
            'booking__created_by'
        ).get(id=payment_id)
        
        if not payment.booking:
            return JsonResponse({'success': False, 'message': 'Payment is not linked to a booking'}, status=404)
        
        context = {'booking': payment.booking}
        return render(request, 'properties/modals/booking_details.html', context)
    except Payment.DoesNotExist:
        # Check if it's a visit payment - these don't have bookings
        from payments.models import Payment as UnifiedPayment
        from .models import PropertyVisitPayment
        try:
            unified_payment = UnifiedPayment.objects.get(id=payment_id)
            visit_payment = PropertyVisitPayment.objects.filter(payment=unified_payment).first()
            if visit_payment:
                return JsonResponse({'success': False, 'message': 'Visit payments do not have booking details'}, status=400)
        except UnifiedPayment.DoesNotExist:
            pass
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Error checking visit payment for payment {payment_id}: {str(e)}")
        return JsonResponse({'success': False, 'message': 'Payment not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading booking details for payment {payment_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'message': f'Error loading booking details: {str(e)}'}, status=500)


@login_required
def api_payment_property_details(request, payment_id):
    """API endpoint to view property details for a payment - handles both booking and visit payments"""
    try:
        property_obj = None
        
        # Try booking payment first (properties.models.Payment)
        try:
            payment = Payment.objects.select_related('booking', 'booking__property_obj', 'booking__property_obj__property_type', 'booking__property_obj__owner').get(id=payment_id)
            if payment.booking and payment.booking.property_obj:
                property_obj = payment.booking.property_obj
        except Payment.DoesNotExist:
            # Try unified payment (payments.models.Payment) - could be visit payment
            from payments.models import Payment as UnifiedPayment
            from .models import PropertyVisitPayment
            try:
                payment = UnifiedPayment.objects.select_related('tenant', 'recorded_by').get(id=payment_id)
                # Check if it's a visit payment
                visit_payment = PropertyVisitPayment.objects.filter(payment=payment).select_related('property', 'property__property_type', 'property__owner').first()
                if visit_payment and visit_payment.property:
                    property_obj = visit_payment.property
                elif hasattr(payment, 'booking') and payment.booking:
                    # Check if unified payment has a booking
                    property_obj = payment.booking.property_obj if payment.booking.property_obj else None
            except UnifiedPayment.DoesNotExist:
                pass
        
        if not property_obj:
            return JsonResponse({'success': False, 'message': 'Payment not linked to a property'}, status=404)
        
        # Get current booking for the property
        current_booking = None
        try:
            current_booking = Booking.objects.filter(
                property_obj=property_obj,
                booking_status__in=['confirmed', 'checked_in']
            ).select_related('customer').first()
        except Exception:
            pass
        
        context = {
            'property': property_obj,
            'current_booking': current_booking,
        }
        return render(request, 'properties/modals/property_details.html', context)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading property details for payment {payment_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'message': f'Error loading property details: {str(e)}'}, status=500)


@login_required
def api_payment_send_reminder(request, payment_id):
    """API endpoint to send payment reminder"""
    try:
        payment = Payment.objects.select_related('booking', 'booking__customer').get(id=payment_id)
        
        if request.method == 'POST':
            # Here you would integrate with email/SMS service
            # For now, we'll just return success
            return JsonResponse({
                'success': True, 
                'message': f'Payment reminder sent to {payment.booking.customer.full_name}'
            })
        
        context = {'payment': payment}
        return render(request, 'properties/modals/payment_send_reminder.html', context)
    except Payment.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Payment not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error loading payment reminder for payment {payment_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'message': f'Error loading payment reminder: {str(e)}'}, status=500)


@login_required
def api_visit_payment_status(request, property_id):
    """Web-friendly endpoint to check visit payment status"""
    try:
        from .models import Property, PropertyVisitPayment
        
        property_obj = get_object_or_404(Property, pk=property_id)
        
        # Only house properties support visit payments
        if property_obj.property_type.name.lower() != 'house':
            return JsonResponse({
                'error': 'Visit payment is only available for house properties.'
            }, status=400)
        
        # Check if user has already paid and if payment is still active
        visit_payment = PropertyVisitPayment.objects.filter(
            property=property_obj,
            user=request.user,
            status='completed'
        ).first()
        
        # Check if payment exists and is still active (not expired)
        has_paid = visit_payment is not None and visit_payment.is_active() if visit_payment else False
        is_expired = visit_payment is not None and visit_payment.is_expired() if visit_payment else False
        
        response_data = {
            'has_paid': has_paid,
            'is_expired': is_expired,
            'visit_cost': float(property_obj.visit_cost) if property_obj.visit_cost else None,
            'property_id': property_obj.id,
            'property_title': property_obj.title
        }
        
        # Add expiration info if payment exists
        if visit_payment:
            expires_at = visit_payment.expires_at()
            response_data['paid_at'] = visit_payment.paid_at.isoformat() if visit_payment.paid_at else None
            response_data['expires_at'] = expires_at.isoformat() if expires_at else None
        
        # If user has paid and payment is active, include contact info
        if has_paid:
            owner = property_obj.owner
            owner_profile = getattr(owner, 'profile', None)
            
            response_data['owner_contact'] = {
                'phone': owner_profile.phone if owner_profile and owner_profile.phone else None,
                'email': owner.email,
                'name': owner.get_full_name() or owner.username
            }
            response_data['location'] = {
                'address': property_obj.address,
                'region': property_obj.region.name if property_obj.region else None,
                'district': property_obj.district.name if property_obj.district else None,
                'latitude': float(property_obj.latitude) if property_obj.latitude else None,
                'longitude': float(property_obj.longitude) if property_obj.longitude else None
            }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Visit payment status error: {str(e)}")
        return JsonResponse({
            'error': f'Failed to check visit status: {str(e)}'
        }, status=500)


@login_required
def api_visit_payment_initiate(request, property_id):
    """Web-friendly endpoint to initiate visit payment"""
    try:
        import json
        from .models import Property, PropertyVisitPayment
        from payments.models import Payment, PaymentProvider, PaymentTransaction
        from django.utils import timezone
        
        property_obj = get_object_or_404(Property, pk=property_id)
        
        # Only house properties support visit payments
        if property_obj.property_type.name.lower() != 'house':
            return JsonResponse({
                'error': 'Visit payment is only available for house properties.'
            }, status=400)
        
        # Check if visit cost is set
        if not property_obj.visit_cost or property_obj.visit_cost <= 0:
            return JsonResponse({
                'error': 'Visit cost is not set for this property.'
            }, status=400)
        
        # Check if user has already paid and if payment is still active
        existing_payment = PropertyVisitPayment.objects.filter(
            property=property_obj,
            user=request.user,
            status='completed'
        ).first()
        
        # If payment exists and is still active (not expired), return error
        if existing_payment and existing_payment.is_active():
            expires_at = existing_payment.expires_at()
            return JsonResponse({
                'error': 'You have already paid for visit access to this property.',
                'has_paid': True,
                'is_expired': False,
                'payment_id': existing_payment.id,
                'expires_at': expires_at.isoformat() if expires_at else None
            }, status=400)
        
        # If payment exists but expired, allow re-payment by updating the existing record
        if existing_payment and existing_payment.is_expired():
            # Reset to pending to allow new payment
            existing_payment.status = 'pending'
            existing_payment.paid_at = None
            existing_payment.transaction_id = None
            existing_payment.gateway_reference = None
            existing_payment.amount = property_obj.visit_cost  # Update amount in case it changed
            existing_payment.save()
            visit_payment = existing_payment
            created = False
        else:
            # Get or create pending visit payment
            visit_payment, created = PropertyVisitPayment.objects.get_or_create(
                property=property_obj,
                user=request.user,
                defaults={
                    'amount': property_obj.visit_cost,
                    'status': 'pending'
                }
            )
        
        # Get or create payment provider (AZAM Pay)
        provider, _ = PaymentProvider.objects.get_or_create(
            name='AZAM Pay',
            defaults={'description': 'AZAM Pay Payment Gateway'}
        )
        
        # Create payment record
        payment = Payment.objects.create(
            tenant=request.user,
            provider=provider,
            amount=visit_payment.amount,
            payment_method='online',
            status='pending',
            notes=f'Visit payment for property: {property_obj.title}',
            recorded_by=request.user
        )
        
        # Link visit payment to unified payment
        visit_payment.payment = payment
        visit_payment.save()
        
        # Initiate gateway payment
        from payments.gateway_service import AZAMPayGateway
        
        # Use configured webhook URL (production) instead of localhost
        from django.conf import settings
        callback_url = getattr(settings, 'AZAM_PAY_WEBHOOK_URL', None)
        if not callback_url:
            # Fallback to BASE_URL if webhook URL not configured
            base_domain = getattr(settings, 'BASE_URL', 'https://portal.maishaapp.co.tz')
            callback_url = f"{base_domain}/api/v1/payments/webhook/azam-pay/"
        gateway_response = AZAMPayGateway.initiate_payment(payment, callback_url=callback_url)
        
        if gateway_response.get('success'):
            # Create payment transaction
            PaymentTransaction.objects.create(
                payment=payment,
                provider=provider,
                gateway_transaction_id=gateway_response.get('transaction_id'),
                azam_reference=gateway_response.get('reference'),
                status='initiated',
                request_payload={'property_id': property_id, 'visit_payment_id': visit_payment.id}
            )
            
            visit_payment.transaction_id = gateway_response.get('transaction_id')
            visit_payment.gateway_reference = gateway_response.get('reference')
            visit_payment.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Payment initiated successfully',
                'payment_link': gateway_response.get('payment_link'),
                'transaction_id': gateway_response.get('transaction_id'),
                'reference': gateway_response.get('reference'),
                'visit_payment_id': visit_payment.id,
                'amount': float(visit_payment.amount)
            })
        else:
            return JsonResponse({
                'error': gateway_response.get('error', 'Failed to initiate payment'),
                'visit_payment_id': visit_payment.id
            }, status=400)
            
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Visit payment initiation error: {str(e)}")
        return JsonResponse({
            'error': f'Failed to initiate visit payment: {str(e)}'
        }, status=500)


@login_required
def api_visit_payment_verify(request, property_id):
    """Web-friendly endpoint to verify visit payment"""
    try:
        import json
        from .models import Property, PropertyVisitPayment
        from django.utils import timezone
        
        property_obj = get_object_or_404(Property, pk=property_id)
        
        data = json.loads(request.body) if request.body else {}
        transaction_id = data.get('transaction_id')
        
        if not transaction_id:
            return JsonResponse({
                'error': 'transaction_id is required'
            }, status=400)
        
        # Get visit payment
        visit_payment = PropertyVisitPayment.objects.filter(
            property=property_obj,
            user=request.user,
            transaction_id=transaction_id
        ).first()
        
        if not visit_payment:
            return JsonResponse({
                'error': 'Visit payment not found'
            }, status=404)
        
        # If already completed and active, return contact info
        if visit_payment.status == 'completed' and visit_payment.is_active():
            owner = property_obj.owner
            owner_profile = getattr(owner, 'profile', None)
            
            expires_at = visit_payment.expires_at()
            return JsonResponse({
                'success': True,
                'message': 'Payment already verified',
                'owner_contact': {
                    'phone': owner_profile.phone if owner_profile and owner_profile.phone else None,
                    'email': owner.email,
                    'name': owner.get_full_name() or owner.username
                },
                'location': {
                    'address': property_obj.address,
                    'region': property_obj.region.name if property_obj.region else None,
                    'district': property_obj.district.name if property_obj.district else None,
                    'latitude': float(property_obj.latitude) if property_obj.latitude else None,
                    'longitude': float(property_obj.longitude) if property_obj.longitude else None
                },
                'paid_at': visit_payment.paid_at.isoformat() if visit_payment.paid_at else None,
                'expires_at': expires_at.isoformat() if expires_at else None,
                'is_expired': False
            })
        
        # If payment exists but expired, prompt user to pay again
        if visit_payment.status == 'completed' and visit_payment.is_expired():
            expires_at = visit_payment.expires_at()
            return JsonResponse({
                'success': False,
                'error': 'Visit payment has expired. Please pay again to access property details.',
                'is_expired': True,
                'paid_at': visit_payment.paid_at.isoformat() if visit_payment.paid_at else None,
                'expires_at': expires_at.isoformat() if expires_at else None
            }, status=400)
        
        # Verify payment with gateway
        if visit_payment.payment:
            from payments.gateway_service import AZAMPayGateway
            
            verification_result = AZAMPayGateway.verify_payment(
                visit_payment.payment,
                transaction_id
            )
            
            if verification_result.get('success') and verification_result.get('verified'):
                # Update visit payment status
                visit_payment.status = 'completed'
                visit_payment.paid_at = timezone.now()
                visit_payment.save()
                
                # Update unified payment
                if visit_payment.payment:
                    visit_payment.payment.status = 'completed'  # Payment model uses 'completed', not 'successful'
                    visit_payment.payment.paid_date = timezone.now().date()
                    visit_payment.payment.save()
                
                # Get owner contact info
                owner = property_obj.owner
                owner_profile = getattr(owner, 'profile', None)
                
                expires_at = visit_payment.expires_at()
                return JsonResponse({
                    'success': True,
                    'message': 'Payment verified successfully',
                    'owner_contact': {
                        'phone': owner_profile.phone if owner_profile and owner_profile.phone else None,
                        'email': owner.email,
                        'name': owner.get_full_name() or owner.username
                    },
                    'location': {
                        'address': property_obj.address,
                        'region': property_obj.region.name if property_obj.region else None,
                        'district': property_obj.district.name if property_obj.district else None,
                        'latitude': float(property_obj.latitude) if property_obj.latitude else None,
                        'longitude': float(property_obj.longitude) if property_obj.longitude else None
                    },
                    'paid_at': visit_payment.paid_at.isoformat(),
                    'expires_at': expires_at.isoformat() if expires_at else None
                })
            else:
                return JsonResponse({
                    'error': 'Payment verification failed',
                    'status': 'failed'
                }, status=400)
        else:
            return JsonResponse({
                'error': 'Payment record not found'
            }, status=404)
            
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Visit payment verification error: {str(e)}")
        return JsonResponse({
            'error': f'Failed to verify visit payment: {str(e)}'
        }, status=500)


@login_required
def api_payment_delete(request, payment_id):
    """API endpoint to delete a payment"""
    try:
        from accounts.models import ActivityLog
        from django.contrib.auth.models import User
        from django.db.models import Q
        
        payment = Payment.objects.select_related('booking', 'booking__customer', 'booking__property_obj', 'booking__property_obj__property_type', 'recorded_by').get(id=payment_id)
        
        if not payment.booking:
            return JsonResponse({'success': False, 'message': 'Payment is not linked to a booking'}, status=404)
        
        if request.method == 'POST':
            # Store payment info for confirmation message
            booking_ref = payment.booking.booking_reference if payment.booking else 'N/A'
            amount = payment.amount or 0
            customer_name = payment.booking.customer.full_name if payment.booking and payment.booking.customer else 'N/A'
            property_title = payment.booking.property_obj.title if payment.booking and payment.booking.property_obj else 'N/A'
            
            # Store booking object before deleting
            booking_obj = payment.booking
            
            # Delete the payment
            payment.delete()
            
            # Recalculate booking's paid_amount and remaining amount after deletion
            if booking_obj:
                from django.db.models import Sum
                from decimal import Decimal
                
                # Recalculate total paid from all remaining payments (exclude refunds)
                actual_paid = Payment.objects.filter(
                    booking=booking_obj
                ).exclude(
                    payment_type='refund'
                ).aggregate(
                    total=Sum('amount')
                )['total'] or Decimal('0')
                
                # Get total refunded
                total_refunded = Payment.objects.filter(
                    booking=booking_obj,
                    payment_type='refund'
                ).aggregate(
                    total=Sum('amount')
                )['total'] or Decimal('0')
                
                # Net paid amount = payments - refunds
                net_paid_amount = max(Decimal('0'), actual_paid - total_refunded)
                
                # Update booking's paid_amount
                booking_obj.paid_amount = net_paid_amount
                
                # Update payment status based on actual amounts
                if hasattr(booking_obj, 'update_payment_status'):
                    booking_obj.update_payment_status()
                booking_obj.save()
            
            # Log the deletion activity
            from decimal import Decimal
            activity_description = f"Payment of Tsh{amount:,.2f} for booking {booking_ref} (Customer: {customer_name}, Property: {property_title})"
            
            ActivityLog.objects.create(
                user=request.user,
                action='delete',
                description=activity_description,
                content_type='Payment',
                object_id=payment_id,
                priority='high',
                amount=Decimal(str(amount)),
                metadata={
                    'booking_reference': booking_ref,
                    'customer_name': customer_name,
                    'property_title': property_title,
                    'deleted_by': request.user.username,
                    'deletion_time': timezone.now().isoformat()
                }
            )
            
            # Send notifications to Admins and Property Managers
            # Find users with Admin or Property Manager roles
            admin_and_manager_users = User.objects.filter(
                Q(is_superuser=True) | 
                Q(groups__name__in=['Admin', 'Manager', 'Property Manager', 'Property manager'])
            ).distinct()
            
            # Create notification logs for each admin/manager
            for admin_user in admin_and_manager_users:
                if admin_user != request.user:  # Don't notify the person who deleted it
                    ActivityLog.objects.create(
                        user=admin_user,
                        action='delete',
                        description=f"⚠️ PAYMENT DELETED: {activity_description} by {request.user.get_full_name() or request.user.username}",
                        content_type='Payment',
                        priority='urgent',
                        amount=Decimal(str(amount)),
                        is_read=False,
                        metadata={
                            'booking_reference': booking_ref,
                            'customer_name': customer_name,
                            'property_title': property_title,
                            'deleted_by': request.user.username,
                            'deleted_by_full_name': request.user.get_full_name() or request.user.username,
                            'deletion_time': timezone.now().isoformat(),
                            'deletion_ip': request.META.get('REMOTE_ADDR', 'Unknown'),
                            'notification_type': 'payment_deletion_alert'
                        }
                    )
            
            return JsonResponse({
                'success': True,
                'message': f'Payment of Tsh{amount} for booking {booking_ref} deleted successfully'
            })
        
        # GET request - return confirmation modal
        context = {'payment': payment}
        return render(request, 'properties/modals/payment_delete_confirm.html', context)
    except Payment.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Payment not found'}, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error deleting payment {payment_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return JsonResponse({'success': False, 'message': f'Error deleting payment: {str(e)}'}, status=500)


# ============================================================================
# PROPERTY APPROVAL MANAGEMENT
# ============================================================================

@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def property_approval_list(request):
    """
    List all property approval requests
    Admin only - shows properties pending approval
    """
    # Handle POST requests for actions
    if request.method == 'POST':
        action = request.POST.get('action')
        property_id = request.POST.get('property_id')
        
        if action == 'get_details':
            try:
                property_obj = Property.objects.get(id=property_id)
                return JsonResponse({
                    'id': property_obj.id,
                    'title': property_obj.title,
                    'description': property_obj.description,
                    'property_type': property_obj.property_type.name if property_obj.property_type else 'N/A',
                    'region': property_obj.region.name if property_obj.region else 'N/A',
                    'address': property_obj.address,
                    'rent_amount': str(property_obj.rent_amount),
                    'owner': property_obj.owner.get_full_name() or property_obj.owner.username,
                    'owner_email': property_obj.owner.email,
                    'created_at': property_obj.created_at.strftime('%B %d, %Y at %I:%M %p'),
                    'is_approved': property_obj.is_approved,
                    'admin_comments': property_obj.admin_comments or '',
                })
            except Property.DoesNotExist:
                return JsonResponse({'error': 'Property not found'}, status=404)
    
    # Get all properties pending approval (only from non-admin owners)
    properties = Property.objects.filter(
        is_approved=False,
        owner__is_staff=False,
        owner__is_superuser=False
    ).select_related(
        'property_type', 'region', 'owner', 'district'
    ).prefetch_related('images', 'amenities').order_by('-created_at')
    
    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter == 'pending':
        properties = properties.filter(is_approved=False, is_active=False)
    elif status_filter == 'rejected':
        # Properties that were rejected (have rejection_reason but not approved)
        properties = properties.exclude(rejection_reason='').filter(is_approved=False)
    
    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        properties = properties.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(owner__username__icontains=search_query) |
            Q(owner__email__icontains=search_query)
        )
    
    # Filter by property type if provided
    property_type_filter = request.GET.get('property_type')
    if property_type_filter:
        properties = properties.filter(property_type_id=property_type_filter)
    
    # Pagination with page size selector
    from django.core.paginator import Paginator
    page_size = request.GET.get('page_size', '10')  # Default to 10 items per page
    try:
        page_size = int(page_size)
        if page_size not in [5, 10, 25, 50, 100]:
            page_size = 10
    except (ValueError, TypeError):
        page_size = 10
    
    paginator = Paginator(properties, page_size)
    page_number = request.GET.get('page')
    
    try:
        page_obj = paginator.get_page(page_number)
    except:
        page_obj = paginator.get_page(1)
    
    # Statistics (only for non-admin owners)
    pending_count = Property.objects.filter(
        is_approved=False, 
        is_active=False,
        owner__is_staff=False,
        owner__is_superuser=False
    ).count()
    rejected_count = Property.objects.filter(
        is_approved=False,
        owner__is_staff=False,
        owner__is_superuser=False
    ).exclude(rejection_reason='').count()
    total_pending = Property.objects.filter(
        is_approved=False,
        owner__is_staff=False,
        owner__is_superuser=False
    ).count()
    
    context = {
        'properties': page_obj,
        'page_obj': page_obj,
        'paginator': paginator,
        'status_choices': [
            ('pending', 'Pending'),
            ('rejected', 'Rejected'),
        ],
        'property_types': PropertyType.objects.all(),
        'current_status': status_filter,
        'current_property_type': property_type_filter,
        'search_query': search_query,
        'current_page_size': page_size,
        'pending_count': pending_count,
        'rejected_count': rejected_count,
        'total_pending': total_pending,
    }

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return render(request, 'properties/property_approval_table.html', context)

    return render(request, 'properties/property_approval_list.html', context)


@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
def property_approval_detail(request, pk):
    """
    View property details for approval
    Admin can view all property details, images, and add comments
    """
    property_obj = get_object_or_404(Property, pk=pk)
    
    # Get property images
    images = property_obj.images.all()
    
    # Get amenities
    amenities = property_obj.amenities.all()
    
    context = {
        'property': property_obj,
        'images': images,
        'amenities': amenities,
    }
    return render(request, 'properties/property_approval_detail.html', context)


@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
@require_http_methods(["POST"])
def approve_property(request, pk):
    """
    Approve a property - makes it visible to the public
    """
    try:
        property_obj = Property.objects.get(pk=pk)
        
        # Get admin comments if provided
        admin_comments = request.POST.get('admin_comments', '').strip()
        
        # Approve the property
        property_obj.is_approved = True
        property_obj.is_active = True  # Also activate it
        property_obj.approved_by = request.user
        property_obj.approved_at = timezone.now()
        property_obj.rejection_reason = None  # Clear any previous rejection reason
        if admin_comments:
            property_obj.admin_comments = admin_comments
        property_obj.save()
        
        messages.success(request, f'Property "{property_obj.title}" has been approved and is now visible to the public.')
        return redirect('properties:property_approval_list')
    except Property.DoesNotExist:
        messages.error(request, 'Property not found.')
        return redirect('properties:property_approval_list')
    except Exception as e:
        messages.error(request, f'Error approving property: {str(e)}')
        return redirect('properties:property_approval_list')


@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
@require_http_methods(["POST"])
def reject_property(request, pk):
    """
    Reject a property with a reason
    """
    try:
        property_obj = Property.objects.get(pk=pk)
        
        # Get rejection reason and admin comments
        rejection_reason = request.POST.get('rejection_reason', '').strip()
        admin_comments = request.POST.get('admin_comments', '').strip()
        
        if not rejection_reason:
            messages.error(request, 'Rejection reason is required.')
            return redirect('properties:property_approval_detail', pk=pk)
        
        # Reject the property
        property_obj.is_approved = False
        property_obj.is_active = False
        property_obj.rejection_reason = rejection_reason
        if admin_comments:
            property_obj.admin_comments = admin_comments
        property_obj.save()
        
        messages.success(request, f'Property "{property_obj.title}" has been rejected.')
        return redirect('properties:property_approval_list')
    except Property.DoesNotExist:
        messages.error(request, 'Property not found.')
        return redirect('properties:property_approval_list')
    except Exception as e:
        messages.error(request, f'Error rejecting property: {str(e)}')
        return redirect('properties:property_approval_list')


@login_required
@user_passes_test(lambda u: u.is_staff or u.is_superuser)
@require_http_methods(["POST"])
def delete_approval_request(request, pk):
    """
    Delete a property approval request (deletes the property)
    """
    try:
        property_obj = Property.objects.get(pk=pk)
        property_title = property_obj.title
        
        # Only allow deletion of unapproved properties
        if property_obj.is_approved:
            messages.error(request, 'Cannot delete an approved property. Please deactivate it instead.')
            return redirect('properties:property_approval_list')
        
        property_obj.delete()
        messages.success(request, f'Property approval request "{property_title}" has been deleted.')
        return redirect('properties:property_approval_list')
    except Property.DoesNotExist:
        messages.error(request, 'Property not found.')
        return redirect('properties:property_approval_list')
    except Exception as e:
        messages.error(request, f'Error deleting property: {str(e)}')
        return redirect('properties:property_approval_list')


# Property Image Management Views

@login_required
@require_http_methods(["POST"])
def property_image_add(request, pk):
    """Add a new image to a property"""
    from django.http import JsonResponse
    
    try:
        property_obj = get_object_or_404(Property, pk=pk)
        
        # Check permissions - only owner or admin can add images
        if not (request.user == property_obj.owner or request.user.is_staff or request.user.is_superuser):
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
        # Check if property already has 10 images (max limit)
        if property_obj.images.count() >= 10:
            return JsonResponse({'success': False, 'error': 'Maximum of 10 images allowed per property'}, status=400)
        
        # Get uploaded file
        if 'image' not in request.FILES:
            return JsonResponse({'success': False, 'error': 'No image file provided'}, status=400)
        
        image_file = request.FILES['image']
        caption = request.POST.get('caption', '')
        is_primary = request.POST.get('is_primary', 'false').lower() == 'true'
        
        # If this is set as primary, unset all other primary images
        if is_primary:
            PropertyImage.objects.filter(property=property_obj).update(is_primary=False)
        
        # Get the highest order number and add 1
        max_order = PropertyImage.objects.filter(property=property_obj).aggregate(
            max_order=Max('order')
        )['max_order'] or 0
        
        # Create new image
        new_image = PropertyImage.objects.create(
            property=property_obj,
            image=image_file,
            caption=caption,
            is_primary=is_primary,
            order=max_order + 1
        )
        new_image.refresh_from_db()
        
        return JsonResponse({
            'success': True,
            'message': 'Image added successfully',
            'image': {
                'id': new_image.id,
                'url': new_image.image.url,
                'caption': new_image.caption or '',
                'is_primary': new_image.is_primary,
                'order': new_image.order
            }
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST", "DELETE"])
def property_image_delete(request, pk, image_id):
    """Delete a property image"""
    from django.http import JsonResponse
    
    try:
        property_obj = get_object_or_404(Property, pk=pk)
        image_obj = get_object_or_404(PropertyImage, pk=image_id, property=property_obj)
        
        # Check permissions - only owner or admin can delete images
        if not (request.user == property_obj.owner or request.user.is_staff or request.user.is_superuser):
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
        # Store image info before deletion
        was_primary = image_obj.is_primary
        
        # Delete the image
        image_obj.delete()
        
        # If the deleted image was primary, set the first remaining image as primary
        if was_primary:
            first_image = PropertyImage.objects.filter(property=property_obj).first()
            if first_image:
                first_image.is_primary = True
                first_image.save(update_fields=['is_primary'])
        
        return JsonResponse({
            'success': True,
            'message': 'Image deleted successfully'
        })
        
    except PropertyImage.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Image not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def geocode_property_address(request):
    """
    AJAX endpoint to geocode an address and return coordinates
    """
    from django.http import JsonResponse
    import json
    
    try:
        data = json.loads(request.body)
        address = data.get('address', '').strip()
        region_id = data.get('region_id')
        
        if not address:
            return JsonResponse({
                'success': False,
                'error': 'Address is required'
            }, status=400)
        
        # Get region name if provided
        region_name = None
        if region_id:
            try:
                region = Region.objects.get(id=region_id)
                region_name = region.name
            except Region.DoesNotExist:
                pass
        
        # Geocode the address
        from .utils import geocode_address
        result = geocode_address(address, region=region_name)
        
        if result:
            latitude, longitude = result
            return JsonResponse({
                'success': True,
                'latitude': float(latitude),
                'longitude': float(longitude)
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Could not find coordinates for this address. Please enter coordinates manually.'
            }, status=404)
            
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Geocoding error: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }, status=500)


@login_required
@require_http_methods(["POST"])
def property_image_set_primary(request, pk, image_id):
    """Set an image as the primary image"""
    from django.http import JsonResponse
    
    try:
        property_obj = get_object_or_404(Property, pk=pk)
        image_obj = get_object_or_404(PropertyImage, pk=image_id, property=property_obj)
        
        # Check permissions - only owner or admin can set primary
        if not (request.user == property_obj.owner or request.user.is_staff or request.user.is_superuser):
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
        # Unset all other primary images
        PropertyImage.objects.filter(property=property_obj).exclude(pk=image_id).update(is_primary=False)
        
        # Set this image as primary - explicitly save with update_fields to ensure database commit
        image_obj.is_primary = True
        image_obj.save(update_fields=['is_primary'])
        
        # Refresh from database to ensure changes are committed
        image_obj.refresh_from_db()
        
        return JsonResponse({
            'success': True,
            'message': 'Primary image updated successfully'
        })
        
    except PropertyImage.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Image not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def property_image_update_caption(request, pk, image_id):
    """Update image caption"""
    from django.http import JsonResponse
    
    try:
        property_obj = get_object_or_404(Property, pk=pk)
        image_obj = get_object_or_404(PropertyImage, pk=image_id, property=property_obj)
        
        # Check permissions - only owner or admin can update caption
        if not (request.user == property_obj.owner or request.user.is_staff or request.user.is_superuser):
            return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
        
        caption = request.POST.get('caption', '')
        image_obj.caption = caption
        image_obj.save(update_fields=['caption'])
        
        # Refresh from database to ensure changes are committed
        image_obj.refresh_from_db()
        
        return JsonResponse({
            'success': True,
            'message': 'Caption updated successfully'
        })
        
    except PropertyImage.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Image not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)