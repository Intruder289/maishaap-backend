from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Avg, Q
from django.utils import timezone
from datetime import datetime, timedelta
from django.core.paginator import Paginator
from .models import ReportTemplate, GeneratedReport, FinancialSummary
from properties.models import Property, Booking, Payment as PropertyPayment
from maintenance.models import MaintenanceRequest
import json


def is_staff_or_admin(user):
    """Check if user is staff or admin"""
    return user.is_staff or user.is_superuser


@login_required
def maintenance_reports_redirect(request):
    """Redirect to the proper maintenance requests page"""
    return redirect('maintenance:request_list')


@login_required
@user_passes_test(is_staff_or_admin)
def reports_dashboard(request):
    """Simple reports dashboard focused on essential reports with filtering"""
    
    # Get date range from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    report_type = request.GET.get('report_type', 'all')
    
    # Set default date range if not provided
    if not start_date:
        start_date = (timezone.now() - timedelta(days=30)).date()
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    if not end_date:
        end_date = timezone.now().date()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Get hotel properties only
    hotel_properties = Property.objects.filter(property_type__name__icontains='hotel')
    
    # Get hotel property IDs for filtering
    hotel_property_ids = hotel_properties.values_list('id', flat=True)
    
    # Payment Reports Data (using correct hotel payment system)
    hotel_payments = PropertyPayment.objects.filter(
        payment_date__date__range=[start_date, end_date],
        booking__property_obj_id__in=hotel_property_ids
    ).order_by('-payment_date')
    
    total_payments = hotel_payments.count()
    total_revenue = hotel_payments.aggregate(total=Sum('amount'))['total'] or 0
    
    # Booking Reports Data (using correct hotel booking system)
    hotel_bookings = Booking.objects.filter(
        check_in_date__range=[start_date, end_date],
        property_obj_id__in=hotel_property_ids
    ).order_by('-check_in_date')
    
    total_bookings = hotel_bookings.count()
    confirmed_bookings = hotel_bookings.filter(booking_status='confirmed').count()
    pending_bookings = hotel_bookings.filter(booking_status='pending').count()
    
    # Property Reports Data
    total_hotels = hotel_properties.count()
    available_hotels = hotel_properties.filter(status='available').count()
    occupied_hotels = hotel_properties.filter(status='rented').count()
    occupancy_rate = (occupied_hotels / total_hotels * 100) if total_hotels > 0 else 0
    
    # Revenue by hotel
    revenue_by_hotel = []
    for hotel in hotel_properties:
        hotel_revenue = PropertyPayment.objects.filter(
            booking__property_obj_id=hotel.id,
            payment_date__date__range=[start_date, end_date]
        ).aggregate(total=Sum('amount'))['total'] or 0
        hotel_booking_count = Booking.objects.filter(
            property_obj_id=hotel.id,
            check_in_date__range=[start_date, end_date]
        ).count()
        
        # Include all hotels, even those with no revenue
        avg_per_booking = hotel_revenue / hotel_booking_count if hotel_booking_count > 0 else 0
        revenue_by_hotel.append({
            'hotel_name': hotel.title,
            'revenue': hotel_revenue,
            'bookings': hotel_booking_count,
            'avg_per_booking': avg_per_booking
        })
    
    # Sort by revenue descending
    revenue_by_hotel.sort(key=lambda x: x['revenue'], reverse=True)
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'report_type': report_type,
        
        # Payment Reports
        'rent_payments': hotel_payments[:20],  # Show last 20 payments
        'total_payments': total_payments,
        'total_revenue': total_revenue,
        
        # Booking Reports
        'bookings': hotel_bookings[:20],  # Show last 20 bookings
        'total_bookings': total_bookings,
        'confirmed_bookings': confirmed_bookings,
        'pending_bookings': pending_bookings,
        
        # Property Reports
        'total_hotels': total_hotels,
        'available_hotels': available_hotels,
        'occupied_hotels': occupied_hotels,
        'occupancy_rate': occupancy_rate,
        'revenue_by_hotel': revenue_by_hotel,
    }
    
    return render(request, 'reports/simple_dashboard.html', context)


@login_required
@user_passes_test(is_staff_or_admin)
def financial_reports(request):
    """Financial reports page"""
    
    # Get date range from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if not start_date:
        start_date = (timezone.now() - timedelta(days=30)).date()
    else:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    
    if not end_date:
        end_date = timezone.now().date()
    else:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    
    # Calculate financial metrics
    rent_payments = RentPayment.objects.filter(
        payment_date__range=[start_date, end_date]
    )
    
    total_rent_collected = rent_payments.aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_rent_collected': total_rent_collected,
        'rent_payments': rent_payments[:10],
    }
    
    return render(request, 'reports/financial.html', context)


@login_required
@user_passes_test(is_staff_or_admin)
def generate_report(request):
    """Generate a new report"""
    
    if request.method == 'POST':
        report_type = request.POST.get('report_type')
        format_type = request.POST.get('format', 'pdf')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        title = request.POST.get('title', f'{report_type.title()} Report')
        
        # Create report record
        report = GeneratedReport.objects.create(
            title=title,
            generated_by=request.user,
            format=format_type,
            start_date=datetime.strptime(start_date, '%Y-%m-%d').date(),
            end_date=datetime.strptime(end_date, '%Y-%m-%d').date(),
            parameters={
                'report_type': report_type,
            }
        )
        
        # For now, mark as completed
        report.status = 'completed'
        report.completed_at = timezone.now()
        report.save()
        
        messages.success(request, f'Report "{title}" generated successfully!')
        return redirect('reports:dashboard')
    
    context = {
        'report_types': [
            ('financial', 'Financial Report'),
            ('maintenance', 'Maintenance Report'),
        ],
        'formats': GeneratedReport.FORMAT_CHOICES,
    }
    
    return render(request, 'reports/generate.html', context)


@login_required
@user_passes_test(is_staff_or_admin)
def download_report(request, report_id):
    """Download a generated report"""
    report = get_object_or_404(GeneratedReport, id=report_id, generated_by=request.user)
    
    if report.status != 'completed' or not report.file_path:
        messages.error(request, 'Report is not ready for download.')
        return redirect('reports:dashboard')
    
    try:
        response = HttpResponse(report.file_path.read(), content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{report.title}.{report.format}"'
        return response
    except Exception as e:
        messages.error(request, f'Error downloading report: {str(e)}')
        return redirect('reports:dashboard')


@login_required
@user_passes_test(is_staff_or_admin)
def view_report(request, report_id):
    """View a generated report"""
    report = get_object_or_404(GeneratedReport, id=report_id, generated_by=request.user)
    
    if report.status != 'completed' or not report.file_path:
        messages.error(request, 'Report is not ready for viewing.')
        return redirect('reports:dashboard')
    
    try:
        response = HttpResponse(report.file_path.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{report.title}.{report.format}"'
        return response
    except Exception as e:
        messages.error(request, f'Error viewing report: {str(e)}')
        return redirect('reports:dashboard')


@login_required
@user_passes_test(is_staff_or_admin)
def generate_financial_report(request):
    """Generate financial report data via AJAX"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)
    
    try:
        import json
        data = json.loads(request.body)
        period = data.get('period', 'monthly')
        date_value = data.get('date')
        
        if not date_value:
            return JsonResponse({'error': 'Date is required'}, status=400)
        
        # Parse date based on period
        from datetime import datetime, timedelta
        from django.utils import timezone
        
        if period == 'daily':
            start_date = datetime.strptime(date_value, '%Y-%m-%d').date()
            end_date = start_date
        elif period == 'monthly':
            year, month = map(int, date_value.split('-'))
            start_date = datetime(year, month, 1).date()
            if month == 12:
                end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        elif period == 'yearly':
            year = int(date_value)
            start_date = datetime(year, 1, 1).date()
            end_date = datetime(year, 12, 31).date()
        else:
            return JsonResponse({'error': 'Invalid period'}, status=400)
        
        # Calculate financial data
        rent_payments = RentPayment.objects.filter(
            payment_date__range=[start_date, end_date]
        )
        
        total_revenue = rent_payments.aggregate(total=Sum('amount'))['total'] or 0
        
        # Get maintenance costs (placeholder since no cost field exists)
        maintenance_costs = 0
        
        # Calculate net income and profit margin
        net_income = total_revenue - maintenance_costs
        profit_margin = (net_income / total_revenue * 100) if total_revenue > 0 else 0
        
        # Generate revenue data for the period
        revenue_data = []
        if period == 'daily':
            revenue_data.append({
                'period': start_date.strftime('%Y-%m-%d'),
                'revenue': total_revenue,
                'costs': maintenance_costs,
                'net_income': net_income,
                'margin': profit_margin
            })
        elif period == 'monthly':
            # Get daily breakdown for the month
            current_date = start_date
            while current_date <= end_date:
                day_revenue = RentPayment.objects.filter(
                    payment_date=current_date
                ).aggregate(total=Sum('amount'))['total'] or 0
                
                revenue_data.append({
                    'period': current_date.strftime('%Y-%m-%d'),
                    'revenue': day_revenue,
                    'costs': 0,
                    'net_income': day_revenue,
                    'margin': 100
                })
                current_date += timedelta(days=1)
        elif period == 'yearly':
            # Get monthly breakdown for the year
            for month in range(1, 13):
                month_start = datetime(year, month, 1).date()
                if month == 12:
                    month_end = datetime(year + 1, 1, 1).date() - timedelta(days=1)
                else:
                    month_end = datetime(year, month + 1, 1).date() - timedelta(days=1)
                
                month_revenue = RentPayment.objects.filter(
                    payment_date__range=[month_start, month_end]
                ).aggregate(total=Sum('amount'))['total'] or 0
                
                revenue_data.append({
                    'period': month_start.strftime('%Y-%m'),
                    'revenue': month_revenue,
                    'costs': 0,
                    'net_income': month_revenue,
                    'margin': 100
                })
        
        return JsonResponse({
            'total_revenue': total_revenue,
            'maintenance_costs': maintenance_costs,
            'net_income': net_income,
            'profit_margin': round(profit_margin, 2),
            'revenue_data': revenue_data
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@user_passes_test(is_staff_or_admin)
def export_payment_report(request):
    """Export payment reports to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        
        # Get date range from request
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        if not start_date:
            start_date = (timezone.now() - timedelta(days=30)).date()
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        
        if not end_date:
            end_date = timezone.now().date()
        else:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Get hotel properties only
        hotel_properties = Property.objects.filter(property_type__name__icontains='hotel')
        hotel_property_ids = hotel_properties.values_list('id', flat=True)
        
        # Get payment data (using correct hotel payment system)
        hotel_payments = PropertyPayment.objects.filter(
            payment_date__date__range=[start_date, end_date],
            booking__property_obj_id__in=hotel_property_ids
        ).order_by('-payment_date')
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payment Report"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add headers
        headers = ['Date', 'Tenant', 'Property', 'Amount (Tsh)', 'Status', 'Payment Method', 'Reference']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        row = 2
        for payment in hotel_payments:
            ws.cell(row=row, column=1, value=payment.payment_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=payment.booking.customer.full_name if payment.booking and payment.booking.customer else 'N/A')
            ws.cell(row=row, column=3, value=payment.booking.property_obj.title if payment.booking and payment.booking.property_obj else 'N/A')
            ws.cell(row=row, column=4, value=float(payment.amount))
            ws.cell(row=row, column=5, value=payment.get_status_display())
            ws.cell(row=row, column=6, value=payment.get_payment_method_display())
            ws.cell(row=row, column=7, value=payment.transaction_reference or 'N/A')
            row += 1
        
        # Add summary row
        total_revenue = hotel_payments.aggregate(total=Sum('amount'))['total'] or 0
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=4, value=float(total_revenue)).font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"payment_report_{start_date}_{end_date}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except ImportError:
        return HttpResponse('Excel generation requires openpyxl package', status=500)
    except Exception as e:
        return HttpResponse(f'Error generating Excel report: {str(e)}', status=500)


@login_required
@user_passes_test(is_staff_or_admin)
def export_booking_report(request):
    """Export booking reports to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        
        # Get date range from request
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        if not start_date:
            start_date = (timezone.now() - timedelta(days=30)).date()
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        
        if not end_date:
            end_date = timezone.now().date()
        else:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Get hotel properties only
        hotel_properties = Property.objects.filter(property_type__name__icontains='hotel')
        hotel_property_ids = hotel_properties.values_list('id', flat=True)
        
        # Get booking data (using correct hotel booking system)
        hotel_bookings = Booking.objects.filter(
            check_in_date__range=[start_date, end_date],
            property_obj_id__in=hotel_property_ids
        ).order_by('-check_in_date')
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Booking Report"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add headers
        headers = ['Date', 'Tenant', 'Property', 'Amount (Tsh)', 'Status', 'Check-in', 'Check-out']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Add data rows
        row = 2
        for booking in hotel_bookings:
            ws.cell(row=row, column=1, value=booking.check_in_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=booking.customer.full_name if booking.customer else 'N/A')
            ws.cell(row=row, column=3, value=booking.property_obj.title if booking.property_obj else 'N/A')
            ws.cell(row=row, column=4, value=float(booking.total_amount))
            ws.cell(row=row, column=5, value=booking.get_booking_status_display())
            ws.cell(row=row, column=6, value=booking.check_in_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=7, value=booking.check_out_date.strftime('%Y-%m-%d'))
            row += 1
        
        # Add summary row
        total_bookings = hotel_bookings.count()
        confirmed_bookings = hotel_bookings.filter(booking_status='confirmed').count()
        ws.cell(row=row, column=1, value="SUMMARY").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"Total: {total_bookings}").font = Font(bold=True)
        ws.cell(row=row, column=3, value=f"Confirmed: {confirmed_bookings}").font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"booking_report_{start_date}_{end_date}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except ImportError:
        return HttpResponse('Excel generation requires openpyxl package', status=500)
    except Exception as e:
        return HttpResponse(f'Error generating Excel report: {str(e)}', status=500)


@login_required
@user_passes_test(is_staff_or_admin)
def financial_excel_report(request):
    """Generate Excel financial report"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        
        period = request.GET.get('period', 'monthly')
        date_value = request.GET.get('date')
        
        if not date_value:
            return HttpResponse('Date is required', status=400)
        
        # Parse date based on period
        from datetime import datetime, timedelta
        
        if period == 'daily':
            start_date = datetime.strptime(date_value, '%Y-%m-%d').date()
            end_date = start_date
        elif period == 'monthly':
            year, month = map(int, date_value.split('-'))
            start_date = datetime(year, month, 1).date()
            if month == 12:
                end_date = datetime(year + 1, 1, 1).date() - timedelta(days=1)
            else:
                end_date = datetime(year, month + 1, 1).date() - timedelta(days=1)
        elif period == 'yearly':
            year = int(date_value)
            start_date = datetime(year, 1, 1).date()
            end_date = datetime(year, 12, 31).date()
        else:
            return HttpResponse('Invalid period', status=400)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Financial Report - {period.title()}"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add headers
        headers = ['Period', 'Revenue (Tsh)', 'Costs (Tsh)', 'Net Income (Tsh)', 'Profit Margin (%)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Get financial data
        rent_payments = RentPayment.objects.filter(
            payment_date__range=[start_date, end_date]
        )
        
        total_revenue = rent_payments.aggregate(total=Sum('amount'))['total'] or 0
        maintenance_costs = 0  # Placeholder
        net_income = total_revenue - maintenance_costs
        profit_margin = (net_income / total_revenue * 100) if total_revenue > 0 else 0
        
        # Add data rows
        row = 2
        if period == 'daily':
            ws.cell(row=row, column=1, value=start_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=total_revenue)
            ws.cell(row=row, column=3, value=maintenance_costs)
            ws.cell(row=row, column=4, value=net_income)
            ws.cell(row=row, column=5, value=profit_margin)
            row += 1
        elif period == 'monthly':
            current_date = start_date
            while current_date <= end_date:
                day_revenue = RentPayment.objects.filter(
                    payment_date=current_date
                ).aggregate(total=Sum('amount'))['total'] or 0
                
                ws.cell(row=row, column=1, value=current_date.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=2, value=day_revenue)
                ws.cell(row=row, column=3, value=0)
                ws.cell(row=row, column=4, value=day_revenue)
                ws.cell(row=row, column=5, value=100)
                row += 1
                current_date += timedelta(days=1)
        elif period == 'yearly':
            for month in range(1, 13):
                month_start = datetime(year, month, 1).date()
                if month == 12:
                    month_end = datetime(year + 1, 1, 1).date() - timedelta(days=1)
                else:
                    month_end = datetime(year, month + 1, 1).date() - timedelta(days=1)
                
                month_revenue = RentPayment.objects.filter(
                    payment_date__range=[month_start, month_end]
                ).aggregate(total=Sum('amount'))['total'] or 0
                
                ws.cell(row=row, column=1, value=month_start.strftime('%Y-%m'))
                ws.cell(row=row, column=2, value=month_revenue)
                ws.cell(row=row, column=3, value=0)
                ws.cell(row=row, column=4, value=month_revenue)
                ws.cell(row=row, column=5, value=100)
                row += 1
        
        # Add summary row
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_revenue).font = Font(bold=True)
        ws.cell(row=row, column=3, value=maintenance_costs).font = Font(bold=True)
        ws.cell(row=row, column=4, value=net_income).font = Font(bold=True)
        ws.cell(row=row, column=5, value=profit_margin).font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"financial_report_{period}_{date_value}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except ImportError:
        return HttpResponse('Excel generation requires openpyxl package', status=500)
    except Exception as e:
        return HttpResponse(f'Error generating Excel report: {str(e)}', status=500)
